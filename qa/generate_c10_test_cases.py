import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PASSWORD = "Admin@123"
DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

OUTPUT_JSON = Path("qa/C-10-Profile-Editability-and-Immutable-Identity-Fields.json")
OUTPUT_XLSX = Path("qa/C-10-Profile-Editability-and-Immutable-Identity-Fields.xlsx")


PROFILE_MODULES = [
    {
        "id": "C10-STU-001",
        "role": "student",
        "route": "/dashboard/student/profile",
        "entityType": "Student Profile",
        "immutableFields": ["Student Name"],
        "editableFieldSamples": ["Phone", "Bio", "Department metadata (if allowed)", "Skills/projects references"],
        "awsElements": ["Profile avatar", "Resume/document upload"],
    },
    {
        "id": "C10-EMP-001",
        "role": "employer",
        "route": "/dashboard/employer/profile",
        "entityType": "Employer Profile",
        "immutableFields": ["Company Name"],
        "editableFieldSamples": ["Website", "Description", "Locations", "Company size/type"],
        "awsElements": ["Company logo upload"],
    },
    {
        "id": "C10-COL-001",
        "role": "college_admin",
        "route": "/dashboard/college/settings",
        "entityType": "College Profile/Settings",
        "immutableFields": ["College Name"],
        "editableFieldSamples": ["Placement officer details", "Website", "Social links", "Season label"],
        "awsElements": ["College logo upload"],
    },
]


def build_editability_case(idx, module):
    role = module["role"]
    return {
        "testCaseId": f"C10-EDIT-{idx:03d}",
        "category": "C-10",
        "type": "editable_fields",
        "title": f"{module['entityType']} editable fields can be updated",
        "role": role,
        "route": module["route"],
        "account": {"email": DEMO_USERS[role], "password": PASSWORD},
        "editableFieldSamples": module["editableFieldSamples"],
        "steps": [
            f"Login as {role} and open {module['route']}",
            "Edit each sample editable field with valid new value",
            "Save changes",
            "Refresh page and re-open profile",
            "Verify updated values persist and render correctly",
        ],
        "expectedResult": [
            "All editable fields accept valid updates",
            "No unauthorized validation blocks for editable fields",
            "Saved values persist after refresh/re-login",
        ],
        "priority": "critical",
    }


def build_immutable_case(idx, module):
    role = module["role"]
    immutable = module["immutableFields"]
    return {
        "testCaseId": f"C10-IMM-{idx:03d}",
        "category": "C-10",
        "type": "immutable_identity_fields",
        "title": f"{module['entityType']} immutable identity fields are protected",
        "role": role,
        "route": module["route"],
        "account": {"email": DEMO_USERS[role], "password": PASSWORD},
        "immutableFields": immutable,
        "steps": [
            f"Login as {role} and open {module['route']}",
            "Attempt to edit immutable identity field(s) through UI",
            "Attempt direct update request via network/API (if feasible in QA)",
            "Save/submit and observe response",
            "Reload and verify immutable values unchanged",
        ],
        "expectedResult": [
            "Immutable field is read-only in UI OR blocked with clear validation",
            "Backend rejects unauthorized mutation attempts",
            "Original immutable identity value remains unchanged",
        ],
        "priority": "critical",
    }


def build_aws_case(idx, module):
    role = module["role"]
    return {
        "testCaseId": f"C10-AWS-{idx:03d}",
        "category": "C-10",
        "type": "aws_stored_elements_update",
        "title": f"{module['entityType']} AWS-backed elements are updateable",
        "role": role,
        "route": module["route"],
        "account": {"email": DEMO_USERS[role], "password": PASSWORD},
        "awsElements": module["awsElements"],
        "steps": [
            f"Login as {role} and open {module['route']}",
            "Upload replacement file/image for AWS-backed element",
            "Complete upload flow (presign + complete, if applicable)",
            "Refresh UI and verify new asset is rendered",
            "Open asset URL to confirm accessibility permissions are correct",
        ],
        "expectedResult": [
            "AWS-backed element upload/update succeeds",
            "New asset reference is persisted and visible in UI",
            "No stale asset after refresh/cache bust",
        ],
        "priority": "high",
    }


def build_cases():
    cases = []
    i = 1
    for mod in PROFILE_MODULES:
        cases.append(build_editability_case(i, mod))
        i += 1
    i = 1
    for mod in PROFILE_MODULES:
        cases.append(build_immutable_case(i, mod))
        i += 1
    i = 1
    for mod in PROFILE_MODULES:
        cases.append(build_aws_case(i, mod))
        i += 1

    cases.append(
        {
            "testCaseId": "C10-AUDIT-001",
            "category": "C-10",
            "type": "profile_field_policy_matrix",
            "title": "Profile field policy matrix: Editable vs Immutable vs AWS-backed",
            "role": "all_roles",
            "route": "all profile modules",
            "account": {"email": "role-wise demo", "password": PASSWORD},
            "steps": [
                "Create full field inventory for student/employer/college profiles",
                "Tag each field as Editable / Immutable / System-managed",
                "Tag each media/document field as AWS-backed or local",
                "Validate UI + API behavior matches policy",
                "Publish discrepancy list",
            ],
            "expectedResult": [
                "Every profile field has explicit governance status",
                "Identity fields remain immutable",
                "AWS-backed profile elements are updateable where intended",
            ],
            "priority": "critical",
        }
    )
    return cases


def write_json(cases):
    payload = {
        "category": "C-10",
        "categoryTitle": "Profile editability with immutable identity constraints",
        "description": "Validate that profile elements are editable except immutable identity fields (student/company/college names), including AWS-backed assets.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-10 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Title",
        "Role",
        "Route",
        "Account Email",
        "Account Password",
        "Editable Field Samples",
        "Immutable Fields",
        "AWS Elements",
        "Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["type"],
                case["title"],
                case["role"],
                case["route"],
                case["account"]["email"],
                case["account"]["password"],
                ", ".join(case.get("editableFieldSamples", [])),
                ", ".join(case.get("immutableFields", [])),
                ", ".join(case.get("awsElements", [])),
                "\n".join(case["steps"]),
                "\n".join(case["expectedResult"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 26,
        "D": 42,
        "E": 14,
        "F": 36,
        "G": 30,
        "H": 14,
        "I": 34,
        "J": 24,
        "K": 28,
        "L": 52,
        "M": 48,
        "N": 10,
        "O": 12,
        "P": 28,
        "Q": 28,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-10 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
