"""Build qa/uat_test_cases_extended_batch_50.xlsx from the companion CSV (no pandas)."""
from __future__ import annotations

import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(
        "openpyxl is required: pip install openpyxl\n"
        f"Original error: {e}"
    ) from e

ROOT = Path(__file__).resolve().parent
CSV_PATH = ROOT / "uat_test_cases_extended_batch_50.csv"
XLSX_PATH = ROOT / "uat_test_cases_extended_batch_50.xlsx"


def main() -> None:
    if not CSV_PATH.is_file():
        raise SystemExit(f"Missing {CSV_PATH}")

    wb = Workbook()
    ws = wb.active
    ws.title = "UAT extended 50"

    with CSV_PATH.open(newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Reasonable column widths (cap width for readability)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        letter = get_column_letter(col_idx)
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = max_len + 2

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH}")


if __name__ == "__main__":
    main()
