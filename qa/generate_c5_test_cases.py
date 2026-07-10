import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}
PASSWORD = "Admin@123"

OUTPUT_JSON = Path("qa/C-5-Tabular-Screens-CSV-Upload-Validation.json")
OUTPUT_XLSX = Path("qa/C-5-Tabular-Screens-CSV-Upload-Validation.xlsx")


# Screens where CSV upload exists or is expected for operational testing.
UPLOAD_CASES = [
    {
        "id": "C5-EMP-001",
        "role": "employer",
        "route": "/dashboard/employer/offers-upload",
        "screen": "Employer Offers Upload",
        "datasetType": "offers",
        "postVerifyRoute": "/dashboard/employer/offers",
    },
    {
        "id": "C5-EMP-002",
        "role": "employer",
        "route": "/dashboard/employer/assessment-uploads",
        "screen": "Employer Assessment Uploads",
        "datasetType": "assessment_results",
        "postVerifyRoute": "/dashboard/employer/assessment-summary",
    },
    {
        "id": "C5-COL-001",
        "role": "college_admin",
        "route": "/dashboard/college/offers-upload",
        "screen": "College Offers Upload",
        "datasetType": "offers",
        "postVerifyRoute": "/dashboard/college/offers",
    },
    {
        "id": "C5-COL-002",
        "role": "college_admin",
        "route": "/dashboard/college/students",
        "screen": "College Students Bulk Upload",
        "datasetType": "students_bulk",
        "postVerifyRoute": "/dashboard/college/students",
    },
]


def build_case(row):
    role = row["role"]
    email = DEMO_USERS[role]
    dataset = row["datasetType"]

    return {
        "testCaseId": row["id"],
        "category": "C-5",
        "title": f"{row['screen']}: CSV upload and data reflection",
        "role": role,
        "route": row["route"],
        "postUploadVerifyRoute": row["postVerifyRoute"],
        "demoAccount": {"email": email, "password": PASSWORD},
        "csvDataStrategy": [
            f"Generate synthetic CSV rows for {dataset} (5-10 records).",
            "Use unique marker (e.g., C5_<timestamp>) in one column for traceability.",
            "Follow platform CSV template columns exactly.",
        ],
        "steps": [
            f"Login as {role} and open {row['route']}",
            "Verify CSV Upload option/button is present and enabled",
            "Generate CSV test file as per template",
            "Upload CSV file",
            "Confirm upload completes with zero row-level errors",
            f"Navigate to {row['postVerifyRoute']}",
            "Search for unique marker rows and validate they are visible",
        ],
        "expectedResult": [
            "CSV upload action is available on screen",
            "Upload request succeeds (no fatal validation/system errors)",
            "Any row-level validation issues are zero for valid generated data",
            "New records appear in target tabular screen after refresh",
        ],
        "priority": "critical",
    }


def build_tabular_presence_audit():
    # Category-level guardrail to enforce the ask:
    # every tabular screen should support CSV upload OR be flagged.
    return {
        "testCaseId": "C5-AUDIT-001",
        "category": "C-5",
        "title": "Tabular screens CSV-upload coverage audit",
        "role": "all_roles",
        "route": "all tabular routes",
        "postUploadVerifyRoute": "N/A",
        "demoAccount": {"email": "use role-wise accounts", "password": PASSWORD},
        "csvDataStrategy": [
            "Create inventory of all tabular screens from navigation + route scan.",
            "Mark each screen as Upload Supported / Upload Missing / Not Applicable.",
            "For supported screens, run upload execution checks with generated data.",
        ],
        "steps": [
            "List all tabular screens role-wise",
            "Check presence of CSV upload control per screen",
            "For each screen with upload control, run upload test",
            "Publish gap list for screens missing upload capability",
        ],
        "expectedResult": [
            "Coverage matrix is complete for all tabular screens",
            "Upload-enabled screens pass functional upload checks",
            "Any tabular screen lacking upload is explicitly identified as a gap",
        ],
        "priority": "critical",
    }


def build_cases():
    cases = [build_tabular_presence_audit()]
    for row in UPLOAD_CASES:
        cases.append(build_case(row))
    return cases


def write_json(cases):
    payload = {
        "category": "C-5",
        "categoryTitle": "Tabular screens CSV upload and reflection validation",
        "description": "Validate CSV upload option presence, successful upload with generated datasets, zero errors, and reflected additional records on screens.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-5 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Role",
        "Screen Route",
        "Post Upload Verify Route",
        "Demo Email",
        "Demo Password",
        "CSV Data Generation Strategy",
        "Test Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["role"],
                case["route"],
                case["postUploadVerifyRoute"],
                case["demoAccount"]["email"],
                case["demoAccount"]["password"],
                "\n".join(case["csvDataStrategy"]),
                "\n".join(case["steps"]),
                "\n".join(case["expectedResult"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 14,
        "B": 10,
        "C": 14,
        "D": 40,
        "E": 40,
        "F": 30,
        "G": 14,
        "H": 52,
        "I": 52,
        "J": 46,
        "K": 10,
        "L": 12,
        "M": 30,
        "N": 30,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-5 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
