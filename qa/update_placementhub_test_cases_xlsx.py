"""Write Status / Actual Result / Executed At into docs/PlacementHub-Test-Cases.xlsx."""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "PlacementHub-Test-Cases.xlsx"
RESULTS = ROOT / "qa" / "data" / "placementhub_test_results.json"

STATUS_FILLS = {
    "Pass": PatternFill("solid", fgColor="C6EFCE"),
    "Fail": PatternFill("solid", fgColor="FFC7CE"),
    "Not Run": PatternFill("solid", fgColor="FFEB9C"),
    "Blocked": PatternFill("solid", fgColor="D9D9D9"),
}
STATUS_FONTS = {
    "Pass": Font(name="Arial", bold=True, color="006100"),
    "Fail": Font(name="Arial", bold=True, color="9C0006"),
    "Not Run": Font(name="Arial", bold=True, color="9C5700"),
    "Blocked": Font(name="Arial", bold=True, color="595959"),
}
HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def clean(text: str) -> str:
    return re.sub(r"[\000-\010\013\014\016-\037]", "", str(text or ""))


def ensure_columns(ws):
    headers = {cell.value: idx for idx, cell in enumerate(ws[2], start=1) if cell.value}
    status_col = headers.get("Status")
    actual_col = headers.get("Actual Result")
    executed_col = headers.get("Executed At")

    next_col = ws.max_column + 1
    if status_col is None:
        status_col = next_col
        ws.cell(2, status_col).value = "Status"
        next_col += 1
    if actual_col is None:
        actual_col = max(ws.max_column + 1, status_col + 1)
        ws.cell(2, actual_col).value = "Actual Result"
    if executed_col is None:
        executed_col = max(ws.max_column + 1, actual_col + 1)
        ws.cell(2, executed_col).value = "Executed At"

    for col in (status_col, actual_col, executed_col):
        cell = ws.cell(2, col)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    return status_col, actual_col, executed_col


def update_index(wb, summary: dict, base_url: str, executed_at: str):
    if "Index" not in wb.sheetnames:
        return
    ws = wb["Index"]
    # Find a free area below existing content for run summary
    start = ws.max_row + 2
    ws.cell(start, 1).value = "Last automated run"
    ws.cell(start, 1).font = Font(name="Arial", bold=True, size=12)
    ws.cell(start + 1, 1).value = "Base URL"
    ws.cell(start + 1, 2).value = base_url
    ws.cell(start + 2, 1).value = "Executed at"
    ws.cell(start + 2, 2).value = executed_at
    ws.cell(start + 3, 1).value = "Pass"
    ws.cell(start + 3, 2).value = summary.get("Pass", 0)
    ws.cell(start + 4, 1).value = "Fail"
    ws.cell(start + 4, 2).value = summary.get("Fail", 0)
    ws.cell(start + 5, 1).value = "Blocked"
    ws.cell(start + 5, 2).value = summary.get("Blocked", 0)
    ws.cell(start + 6, 1).value = "Not Run"
    ws.cell(start + 6, 2).value = summary.get("Not Run", 0)
    for r in range(start + 1, start + 7):
        ws.cell(r, 1).font = BODY_FONT
        ws.cell(r, 2).font = BODY_FONT


def main() -> None:
    if not RESULTS.is_file():
        raise SystemExit(f"Missing results JSON: {RESULTS}")
    payload = json.loads(RESULTS.read_text(encoding="utf-8"))
    by_id = {r["id"]: r for r in payload.get("results", [])}

    wb = load_workbook(XLSX)
    updated = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Index", "Coverage Matrix"):
            continue
        ws = wb[sheet_name]
        status_col, actual_col, executed_col = ensure_columns(ws)
        for row_idx in range(3, ws.max_row + 1):
            tc_id = ws.cell(row_idx, 1).value
            if not tc_id or tc_id not in by_id:
                continue
            r = by_id[tc_id]
            status = r.get("status") or "Not Run"
            status_cell = ws.cell(row_idx, status_col)
            status_cell.value = status
            status_cell.font = STATUS_FONTS.get(status, BODY_FONT)
            status_cell.fill = STATUS_FILLS.get(status, PatternFill())
            status_cell.alignment = Alignment(vertical="center")
            status_cell.border = THIN

            actual_cell = ws.cell(row_idx, actual_col)
            actual_cell.value = clean(r.get("actual", ""))
            actual_cell.font = BODY_FONT
            actual_cell.alignment = Alignment(wrap_text=True, vertical="top")

            exec_cell = ws.cell(row_idx, executed_col)
            exec_cell.value = r.get("executedAt") or payload.get("executedAt") or datetime.utcnow().isoformat()
            exec_cell.font = BODY_FONT
            updated += 1

        ws.column_dimensions["M"].width = 12
        if actual_col:
            ws.column_dimensions[ws.cell(2, actual_col).column_letter].width = 48
        if executed_col:
            ws.column_dimensions[ws.cell(2, executed_col).column_letter].width = 22

    update_index(
        wb,
        payload.get("summary") or {},
        payload.get("baseUrl") or "",
        payload.get("executedAt") or "",
    )
    wb.save(XLSX)
    print(f"Updated {updated} rows in {XLSX}")
    print("Summary:", payload.get("summary"))


if __name__ == "__main__":
    main()
