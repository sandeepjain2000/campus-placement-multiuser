import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}
PASSWORD = "Admin@123"

OUTPUT_JSON = Path("qa/C-4-Non-Hardcoded-Data-Validation.json")
OUTPUT_XLSX = Path("qa/C-4-Non-Hardcoded-Data-Validation.xlsx")


DATA_CHECKS = [
    {
        "id": "C4-STU-001",
        "role": "student",
        "route": "/dashboard/student/profile",
        "dataPoint": "Student identity + academic profile fields",
        "sourceOfTruth": "users + student_profiles tables",
    },
    {
        "id": "C4-STU-002",
        "role": "student",
        "route": "/dashboard/student/drives",
        "dataPoint": "Drive list (title, company, dates, status)",
        "sourceOfTruth": "placement_drives + job_postings + employer_profiles",
    },
    {
        "id": "C4-STU-003",
        "role": "student",
        "route": "/dashboard/student/applications",
        "dataPoint": "Application status timeline",
        "sourceOfTruth": "applications + application_status_log",
    },
    {
        "id": "C4-STU-004",
        "role": "student",
        "route": "/dashboard/student/offers",
        "dataPoint": "Offer amount, deadline, response status",
        "sourceOfTruth": "offers",
    },
    {
        "id": "C4-EMP-001",
        "role": "employer",
        "route": "/dashboard/employer/profile",
        "dataPoint": "Company profile details and logo",
        "sourceOfTruth": "employer_profiles",
    },
    {
        "id": "C4-EMP-002",
        "role": "employer",
        "route": "/dashboard/employer/jobs",
        "dataPoint": "Job postings list and criteria",
        "sourceOfTruth": "job_postings",
    },
    {
        "id": "C4-EMP-003",
        "role": "employer",
        "route": "/dashboard/employer/applications",
        "dataPoint": "Candidate rows and statuses",
        "sourceOfTruth": "applications + student_profiles + shortlists",
    },
    {
        "id": "C4-EMP-004",
        "role": "employer",
        "route": "/dashboard/employer/offers",
        "dataPoint": "Issued offer list",
        "sourceOfTruth": "offers",
    },
    {
        "id": "C4-COL-001",
        "role": "college_admin",
        "route": "/dashboard/college/overview",
        "dataPoint": "Dashboard KPI cards",
        "sourceOfTruth": "aggregated API from drives/applications/offers",
    },
    {
        "id": "C4-COL-002",
        "role": "college_admin",
        "route": "/dashboard/college/students",
        "dataPoint": "Student directory rows",
        "sourceOfTruth": "student_profiles + users",
    },
    {
        "id": "C4-COL-003",
        "role": "college_admin",
        "route": "/dashboard/college/drives",
        "dataPoint": "Drive request/approval data",
        "sourceOfTruth": "placement_drives",
    },
    {
        "id": "C4-COL-004",
        "role": "college_admin",
        "route": "/dashboard/college/employers",
        "dataPoint": "Employer approvals and company data",
        "sourceOfTruth": "employer_approvals + employer_profiles",
    },
    {
        "id": "C4-ADM-001",
        "role": "super_admin",
        "route": "/dashboard/admin/overview",
        "dataPoint": "Platform-wide totals",
        "sourceOfTruth": "aggregated tenant/user/employer/student counts",
    },
    {
        "id": "C4-ADM-002",
        "role": "super_admin",
        "route": "/dashboard/admin/colleges",
        "dataPoint": "Colleges list and metadata",
        "sourceOfTruth": "tenants",
    },
    {
        "id": "C4-ADM-003",
        "role": "super_admin",
        "route": "/dashboard/admin/users",
        "dataPoint": "Users list and roles",
        "sourceOfTruth": "users",
    },
    {
        "id": "C4-ALL-001",
        "role": "all_roles",
        "route": "/dashboard/alerts",
        "dataPoint": "Alert/notification entries",
        "sourceOfTruth": "notifications",
    },
    {
        "id": "C4-ALL-002",
        "role": "all_roles",
        "route": "/dashboard/feedback",
        "dataPoint": "Feedback tickets and updates",
        "sourceOfTruth": "platform_feedback + platform_feedback_replies",
    },
    {
        "id": "C4-ALL-003",
        "role": "all_roles",
        "route": "/dashboard/my-exports",
        "dataPoint": "Export history list",
        "sourceOfTruth": "audit/export tracking records",
    },
]


def account_for_role(role):
    if role == "all_roles":
        return {"email": "use role-wise demo account", "password": PASSWORD}
    return {"email": DEMO_USERS[role], "password": PASSWORD}


def build_cases():
    cases = []
    for row in DATA_CHECKS:
        role = row["role"]
        account = account_for_role(role)

        case = {
            "testCaseId": row["id"],
            "category": "C-4",
            "title": f"{row['dataPoint']} is not hard-coded",
            "role": role,
            "route": row["route"],
            "demoAccount": account,
            "dataPointUnderTest": row["dataPoint"],
            "sourceOfTruth": row["sourceOfTruth"],
            "verificationApproach": [
                "Capture baseline value on UI.",
                "Modify source data (seed/DB/API fixture) for the same record.",
                "Reload page (or re-login) and verify updated value appears.",
                "Confirm old baseline value no longer appears.",
                "Run same check with second record to avoid one-off caching effects.",
            ],
            "expectedResult": [
                "UI values reflect latest source data after refresh.",
                "No constant/static placeholder remains in place of live data.",
                "Different tenants/users show different values where expected.",
            ],
            "negativeSignalsOfHardcoding": [
                "Same value shown regardless of changed DB/API record.",
                "Identical demo text appears for multiple distinct entities.",
                "Value only changes after code change, not data change.",
            ],
            "priority": "critical",
        }
        cases.append(case)
    return cases


def write_json(cases):
    payload = {
        "category": "C-4",
        "categoryTitle": "Non-hardcoded information validation",
        "description": "Validate all major information blocks are dynamic and data-driven, not static hard-coded literals.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-4 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Role",
        "Route",
        "Data Point Under Test",
        "Source Of Truth",
        "Demo Email",
        "Demo Password",
        "Verification Steps",
        "Expected Result",
        "Hardcoding Failure Signals",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["role"],
                case["route"],
                case["dataPointUnderTest"],
                case["sourceOfTruth"],
                case["demoAccount"]["email"],
                case["demoAccount"]["password"],
                "\n".join(case["verificationApproach"]),
                "\n".join(case["expectedResult"]),
                "\n".join(case["negativeSignalsOfHardcoding"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 14,
        "B": 10,
        "C": 14,
        "D": 40,
        "E": 38,
        "F": 38,
        "G": 30,
        "H": 14,
        "I": 56,
        "J": 42,
        "K": 44,
        "L": 10,
        "M": 12,
        "N": 28,
        "O": 28,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-4 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
