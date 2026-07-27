"""Generate docs/PlacementHub-Test-Cases-Import-Export.xlsx"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "PlacementHub-Test-Cases-Import-Export.xlsx"

HEADERS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
    "Actual Result",
    "Executed At",
]

# (sheet_name, module_label, cases[])
# case: (feature, title, priority, type, roles, preconditions, steps, expected, notes, automation)

CASES: list[tuple[str, str, list[tuple]]] = []


def add(sheet: str, module: str, rows: list[tuple]):
    CASES.append((sheet, module, rows))


add(
    "01 Students CSV",
    "01 Students CSV",
    [
        (
            "Template",
            "College can download students import template CSV",
            "P0",
            "Functional",
            "College Admin",
            "Signed in as college admin",
            "1. Open /dashboard/college/students\n2. Open Import menu\n3. Download import template",
            "CSV downloads with BOM; headers match STUDENT_CSV_HEADERS; one example row present",
            "GET /api/college/students/import-template",
            "Candidate",
        ),
        (
            "Import happy path",
            "Valid CSV imports new students and returns processed count > 0",
            "P0",
            "Functional",
            "College Admin",
            "Template filled with unique rolls/emails; Verified Yes or No",
            "1. Fill template with 2 new students\n2. Upload via Import\n3. Observe result toast/API and email",
            "processed ≥ 2; students appear in roster; welcome emails attempted for new accounts",
            "Unique rolls not already in tenant",
            "Candidate",
        ),
        (
            "Import upsert",
            "Re-import same roll updates mutable fields without changing locked name/email",
            "P0",
            "Functional",
            "College Admin",
            "Student already exists with known roll",
            "1. Export or note existing roll\n2. Re-upload row with same roll, changed CGPA/skills\n3. Change name/email in CSV and re-upload",
            "CGPA/skills update; name/email remain locked on existing profile",
            "Document which fields are locked",
            "Manual",
        ),
        (
            "Verified blank",
            "Blank Verified cell defaults to No and still imports",
            "P0",
            "Functional",
            "College Admin",
            "Otherwise valid row with Verified empty",
            "1. Upload CSV with Verified blank\n2. Open student detail",
            "Row imports; verified flag is false / No",
            "Regression for prior 'Verified must be Yes or No' blank failure",
            "Candidate",
        ),
        (
            "Verified aliases",
            "Verified accepts Yes/Y/true/1 and No/N/false/0/pending",
            "P1",
            "Functional",
            "College Admin",
            "Valid base rows",
            "1. Upload rows using each accepted alias\n2. Confirm stored verified boolean",
            "All accepted aliases map correctly; pending → No",
            None,
            "Candidate",
        ),
        (
            "Verified invalid",
            "Invalid Verified value fails that row with got \"…\" message",
            "P1",
            "Negative",
            "College Admin",
            "Valid CSV except Verified=Maybe",
            "1. Upload file\n2. Read errors list / completion email",
            "processed may be 0 for that row; error includes got \"Maybe\"",
            None,
            "Candidate",
        ),
        (
            "Required columns",
            "Missing required header columns is rejected before row processing",
            "P0",
            "Negative",
            "College Admin",
            "CSV missing Email or Roll header",
            "1. Upload malformed header CSV",
            "Clear Missing columns: … error; no partial silent import",
            None,
            "Candidate",
        ),
        (
            "Required cells",
            "Blank required cells (Name/Roll/Email/CGPA/etc.) fail the row",
            "P0",
            "Negative",
            "College Admin",
            "Headers valid; Name blank on line 2",
            "1. Upload\n2. Inspect errors",
            "Error identifies missing fields; other valid rows may still process",
            "Remarks/Photo URL/Verified may be blank",
            "Candidate",
        ),
        (
            "CGPA validation",
            "Out-of-range CGPA is rejected with clear message",
            "P1",
            "Negative",
            "College Admin",
            "CGPA=12 or non-numeric",
            "1. Upload row\n2. Read error",
            "Row fails; message references CGPA rules",
            None,
            "Candidate",
        ),
        (
            "Job/Internship status",
            "Invalid Job Status or Internship Status fails row",
            "P1",
            "Negative",
            "College Admin",
            "Job Status=hired (invalid)",
            "1. Upload\n2. Read error",
            "Error lists allowed statuses",
            "JOB: unplaced,placed,opted_out,higher_studies; INTERN: none,ongoing,completed",
            "Candidate",
        ),
        (
            "Export roster",
            "College can export current students roster CSV",
            "P0",
            "Functional",
            "College Admin",
            "At least one student in tenant",
            "1. Open students page\n2. Export CSV (filtered and/or full)",
            "CSV downloads; columns align with template; Verified is Yes/No",
            "ExportCsvSplitButton",
            "Candidate",
        ),
        (
            "Round-trip",
            "Export → re-import does not create duplicates for same rolls",
            "P1",
            "Regression",
            "College Admin",
            "Existing roster",
            "1. Export full roster\n2. Re-import same file unchanged",
            "No duplicate users; processed as updates; error count reasonable",
            None,
            "Manual",
        ),
        (
            "Authz",
            "Student/employer cannot call students bulk-upload or import-template",
            "P0",
            "Security",
            "Student; Employer",
            "Authed non-college sessions",
            "1. POST /api/college/students/bulk-upload\n2. GET import-template",
            "401/403 for both",
            None,
            "Candidate",
        ),
        (
            "File size",
            "CSV larger than upload limit is rejected",
            "P2",
            "Negative",
            "College Admin",
            "File > MAX_CSV_UPLOAD_BYTES (~5MB)",
            "1. Upload oversized CSV",
            "Clear size/limit error; no hang",
            None,
            "Manual",
        ),
        (
            "Completion notify",
            "Import completion email/alert states success count and lists errors",
            "P1",
            "Functional",
            "College Admin",
            "Import with mix of success and errors",
            "1. Run import\n2. Check email + Alerts",
            "Subject includes success count; body lists errors; 0-success wording is clear",
            "context college_student_bulk_import",
            "Candidate",
        ),
    ],
)

add(
    "02 Calendar ICS CSV",
    "02 Calendar ICS CSV",
    [
        (
            "ICS import",
            "College can import valid .ics and create calendar rows",
            "P0",
            "Functional",
            "College Admin",
            "Sample ICS with unique UIDs",
            "1. Open /dashboard/college/calendar\n2. Import ICS\n3. Refresh calendar",
            "Events appear; API returns imported count; source_uid stored",
            "docs/sample-import-july26.ics; POST /api/college/calendar/import",
            "Automated",
        ),
        (
            "ICS dry-run",
            "Dry-run preview reports clashes without writing when overlapping drives exist",
            "P1",
            "Functional",
            "College Admin",
            "Approved drive on same date as ICS event",
            "1. Import with dryRun=true",
            "Preview/clash warnings returned; no durable duplicate create from dry-run",
            None,
            "Automated",
        ),
        (
            "ICS re-import UID",
            "Re-import same UID does not duplicate events",
            "P0",
            "Regression",
            "College Admin",
            "Prior successful import",
            "1. Import sample ICS twice",
            "Second pass upserts/updates; duplicate count does not grow unchecked",
            None,
            "Automated",
        ),
        (
            "ICS invalid",
            "Invalid/empty ICS rejected with clear error",
            "P0",
            "Negative",
            "College Admin",
            "None",
            "1. Upload non-ICS text as .ics",
            "HTTP 4xx; clear parse error",
            None,
            "Automated",
        ),
        (
            "ICS export full",
            "Full ICS export includes imported and placement drive events",
            "P0",
            "Functional",
            "College Admin",
            "Imported events present",
            "1. Export scope=full\n2. Open file",
            "BEGIN:VCALENDAR; contains imported titles/UIDs and drive events",
            "GET /api/college/calendar/export?scope=full",
            "Automated",
        ),
        (
            "ICS export month",
            "Month-scoped ICS export filters to selected month",
            "P1",
            "Functional",
            "College Admin",
            "Events in Jul 2026 and other months",
            "1. Export scope=month&year=2026&month=7",
            "Only July (overlapping) events included",
            None,
            "Candidate",
        ),
        (
            "CSV export",
            "Calendar CSV export works for month and full scopes from UI",
            "P1",
            "Functional",
            "College Admin",
            "Calendar has events",
            "1. Use Export menu → CSV month/full",
            "CSV downloads with expected columns",
            "Client-side export",
            "Candidate",
        ),
        (
            "Delete imported range",
            "Delete imported removes source_uid rows in date range only",
            "P1",
            "Functional",
            "College Admin",
            "Imported events across months",
            "1. DELETE imported with from/to range\n2. Verify remaining events",
            "Only range cleared; manual/placement events untouched",
            "/api/college/calendar/imported",
            "Automated",
        ),
        (
            "Delete all imported",
            "Delete all imported clears tenant source_uid events",
            "P1",
            "Functional",
            "College Admin",
            "Imported events exist",
            "1. DELETE scope=all",
            "All imported gone; drives remain on calendar export",
            None,
            "Automated",
        ),
        (
            "Filters",
            "All / Placement / Imported / Programs filters change visible set",
            "P1",
            "UI",
            "College Admin",
            "Mix of event types",
            "1. Toggle each filter",
            "Visible events match category",
            None,
            "Automated",
        ),
        (
            "Authz",
            "Non-college roles cannot import/export/delete college calendar ICS",
            "P0",
            "Security",
            "Student; Employer",
            "Authed sessions",
            "1. Call import/export/imported APIs",
            "401/403",
            None,
            "Automated",
        ),
    ],
)

add(
    "03 Employer Assessments",
    "03 Employer Assessments",
    [
        (
            "Template",
            "Employer can download assessment CSV template",
            "P0",
            "Functional",
            "Employer",
            "Signed in employer",
            "1. Open assessment uploads\n2. Download template OR GET /api/employer/assessments/template",
            "CSV with required headers and sample row",
            "Prefer campus export for real rolls",
            "Candidate",
        ),
        (
            "Campus export starter",
            "Employer exports campus students/applicants CSV for a kind+tenant",
            "P0",
            "Functional",
            "Employer",
            "Approved partnership with campus; kind set",
            "1. Select campus and kind (internship/job/drive)\n2. Download export/starter",
            "CSV rolls match eligible students; headers match upload",
            "GET /api/employer/assessments/export",
            "Candidate",
        ),
        (
            "Upload happy path",
            "Valid assessment CSV commits results for matching rolls",
            "P0",
            "Functional",
            "Employer",
            "Export file filled with valid outcomes",
            "1. Upload CSV\n2. Open hiring assessment view",
            "Rows committed; college can see results; no needsReview",
            "POST /api/employer/assessments/upload",
            "Candidate",
        ),
        (
            "Needs review",
            "Invalid/unknown rolls return needsReview staging session",
            "P0",
            "Functional",
            "Employer",
            "CSV with bad roll",
            "1. Upload\n2. Open review UI\n3. Fix and accept",
            "422/needsReview; session editable; accept commits fixed rows",
            "/api/employer/assessments/import/[sessionId]",
            "Candidate",
        ),
        (
            "Campus required",
            "Upload without approved campus tenant fails clearly",
            "P0",
            "Negative",
            "Employer",
            "No campus selected / not approved",
            "1. Attempt upload",
            "Error requires campus/tenant context",
            None,
            "Candidate",
        ),
        (
            "Kind matrix",
            "Upload works for internship, jobs, drive, and projects kinds",
            "P1",
            "Functional",
            "Employer",
            "Fixtures per kind",
            "1. Upload one CSV per kind",
            "Each kind stores against correct opportunity context",
            None,
            "Manual",
        ),
        (
            "Hiring assessment export",
            "Employer hiring-assessment page CSV export downloads",
            "P1",
            "Functional",
            "Employer",
            "Committed assessment data",
            "1. Export from hiring assessment",
            "CSV reflects visible/filtered rows",
            None,
            "Candidate",
        ),
        (
            "College view export",
            "College hiring-assessment view can export employer-uploaded results",
            "P1",
            "Functional",
            "College Admin",
            "Employer uploaded results for tenant",
            "1. Open /dashboard/college/hiring-assessment\n2. Export CSV",
            "CSV contains assessment rows for campus",
            None,
            "Candidate",
        ),
        (
            "Authz",
            "College/student cannot upload employer assessments",
            "P0",
            "Security",
            "College Admin; Student",
            "Authed",
            "1. POST /api/employer/assessments/upload",
            "401/403",
            None,
            "Candidate",
        ),
    ],
)

add(
    "04 Offers Files",
    "04 Offers Files",
    [
        (
            "College upload removed",
            "College offer CSV upload returns 410 removed",
            "P0",
            "Regression",
            "College Admin",
            "Signed in",
            "1. POST /api/college/offers/upload with any CSV\n2. Or use offers-upload UI",
            "HTTP 410; message that CSV import was removed",
            "UI may still exist — document residue",
            "Candidate",
        ),
        (
            "Employer upload removed",
            "Employer offer CSV upload returns 410 removed",
            "P0",
            "Regression",
            "Employer",
            "Signed in",
            "1. POST /api/employer/offers/upload",
            "HTTP 410; guidance to use templates + bulk generate",
            None,
            "Candidate",
        ),
        (
            "College blank template",
            "College can download blank offers CSV template from UI helper",
            "P2",
            "Functional",
            "College Admin",
            "Offers-upload page reachable",
            "1. Download blank template",
            "CSV downloads (even if upload path is retired)",
            "Client collegeOffersCsvTemplate",
            "Manual",
        ),
        (
            "Assessment starter",
            "College/employer can download all-students offers assessment starter CSV",
            "P1",
            "Functional",
            "College Admin; Employer",
            "Students on campus master list",
            "1. GET assessment-starter for role\n2. Open CSV",
            "Rolls present; company may be prefilled from latest assessment",
            "/api/college|employer/offers/assessment-starter",
            "Candidate",
        ),
        (
            "Bulk generate",
            "Employer bulk-generate creates formal offers from template + selections",
            "P0",
            "Functional",
            "Employer",
            "Offer template + selected students on drive/job",
            "1. Preview bulk generate\n2. Commit generate\n3. Student sees offers",
            "Offers created with rendered_letter_html; students notified",
            "Not CSV — HTML generation path",
            "Candidate",
        ),
        (
            "Offers list export",
            "Employer and college can export offers list CSV",
            "P1",
            "Functional",
            "Employer; College Admin",
            "At least one offer",
            "1. Export from offers page",
            "CSV downloads with key offer fields",
            None,
            "Candidate",
        ),
        (
            "Student letter view",
            "Student can open formal offer letter HTML view",
            "P0",
            "Functional",
            "Student",
            "Pending formal offer issued",
            "1. Open My Offers\n2. Open letter",
            "Letter HTML renders; optional download URL works if present",
            "/api/student/offers/[id]/letter",
            "Candidate",
        ),
    ],
)

add(
    "05 List CSV Exports",
    "05 List CSV Exports",
    [
        (
            "College drives",
            "College drives list CSV export",
            "P1",
            "Functional",
            "College Admin",
            "≥1 drive",
            "1. /dashboard/college/drives → Export",
            "CSV downloads",
            None,
            "Candidate",
        ),
        (
            "College drive report JSON",
            "College can download post-drive report as JSON",
            "P1",
            "Functional",
            "College Admin",
            "Drive with report data",
            "1. Trigger report download for a drive",
            ".json file with stats/selected students",
            "GET /api/college/drives/[id]/report",
            "Candidate",
        ),
        (
            "College reports multi-CSV",
            "College reports page exports each analytics CSV",
            "P1",
            "Functional",
            "College Admin",
            "Desktop viewport",
            "1. Open /dashboard/college/reports\n2. Export each menu item",
            "Dept placement, salary, recruiters, YoY, events CSVs download",
            "Mobile may toast to use desktop",
            "Candidate",
        ),
        (
            "College interviews",
            "College interviews schedule and results CSV exports",
            "P1",
            "Functional",
            "College Admin",
            "Interview data present",
            "1. Export schedule\n2. Export results",
            "Two distinct CSVs download",
            None,
            "Candidate",
        ),
        (
            "College internship suite",
            "Internship results/feedback/guides/PPO CSV exports",
            "P1",
            "Functional",
            "College Admin",
            "Pages have data or empty state handled",
            "1. Export from each internship-* page",
            "CSV downloads without crash (empty file or headers OK)",
            None,
            "Candidate",
        ),
        (
            "College guest & clarifications",
            "Guest engagements and clarifications exports",
            "P2",
            "Functional",
            "College Admin",
            "Optional data",
            "1. Export guest engagements\n2. Export clarifications (CSV/TXT)",
            "Files download",
            None,
            "Manual",
        ),
        (
            "Employer lists",
            "Employer applications/drives/interviews/internships/calendar CSV exports",
            "P1",
            "Functional",
            "Employer",
            "Signed in; campus selected where required",
            "1. Export from each list page",
            "Each export succeeds",
            "Calendar is CSV not ICS for employer",
            "Candidate",
        ),
        (
            "Employer supervisors",
            "Internship supervisors CSV export",
            "P2",
            "Functional",
            "Employer",
            "Supervisors page",
            "1. Export",
            "CSV downloads",
            None,
            "Manual",
        ),
        (
            "Student opportunities",
            "Student jobs/internships/not-processed CSV exports",
            "P1",
            "Functional",
            "Student",
            "Visible opportunities",
            "1. Export from each page",
            "CSV matches loaded rows",
            "studentOpportunityCsvExport.js",
            "Candidate",
        ),
        (
            "Student applications",
            "Student applications CSV export by type",
            "P1",
            "Functional",
            "Student",
            "≥1 application",
            "1. Export from applications tabs",
            "CSV downloads",
            None,
            "Candidate",
        ),
        (
            "Student clarifications",
            "Student clarifications CSV/TXT export",
            "P2",
            "Functional",
            "Student",
            "Optional data",
            "1. Export",
            "File downloads",
            None,
            "Manual",
        ),
        (
            "Admin lists",
            "Super admin users/colleges/employers/pending/listings/feedback CSV exports",
            "P1",
            "Functional",
            "Super Admin",
            "Signed in",
            "1. Export from each admin list\n2. Overview section CSV",
            "All downloads succeed",
            None,
            "Candidate",
        ),
        (
            "Filtered vs full",
            "Export split button respects filtered vs full list where offered",
            "P1",
            "Functional",
            "College Admin; Employer",
            "Filters active reducing row count",
            "1. Apply filter\n2. Export filtered\n3. Export all",
            "Filtered file smaller/subset; full includes unfiltered rows",
            "ExportCsvSplitButton",
            "Candidate",
        ),
    ],
)

add(
    "06 Audit My Data",
    "06 Audit My Data",
    [
        (
            "Audit export college",
            "College audit report export emails S3 download link",
            "P0",
            "Functional",
            "College Admin",
            "S3 configured; date range with logs",
            "1. /dashboard/college/audit-reports\n2. Choose range\n3. Export",
            "Email with time-limited CSV link; file opens; tenant-scoped rows only",
            "POST /api/audit/reports/export",
            "Candidate",
        ),
        (
            "Audit export admin",
            "Super admin audit export can be platform-wide",
            "P1",
            "Functional",
            "Super Admin",
            "S3 configured",
            "1. Admin audit-reports export",
            "CSV delivered; scope matches selection",
            None,
            "Candidate",
        ),
        (
            "Audit missing range",
            "Audit export without date range fails validation",
            "P1",
            "Negative",
            "College Admin",
            "None",
            "1. Submit export with empty dates",
            "Validation error; no email",
            None,
            "Candidate",
        ),
        (
            "My data export student",
            "Student My Exports creates personal data CSV and notifies",
            "P0",
            "Functional",
            "Student",
            "user_data_exports migration applied",
            "1. /dashboard/my-exports\n2. Request export\n3. Download/open attachment",
            "Multi-section CSV; history row stored; email/alert sent",
            "POST /api/user/data-export",
            "Candidate",
        ),
        (
            "My data export roles",
            "Employer, college, and super admin can request role-appropriate data export",
            "P1",
            "Functional",
            "Employer; College Admin; Super Admin",
            "Migration applied",
            "1. Request export as each role\n2. Inspect payload sections",
            "Payload matches role; college does not dump full student PII",
            None,
            "Manual",
        ),
        (
            "My data history",
            "Prior exports appear in My Exports history",
            "P2",
            "Functional",
            "Student",
            "At least one prior export",
            "1. Open my-exports",
            "History lists previous requests/status",
            None,
            "Candidate",
        ),
        (
            "Migration missing",
            "Data export returns 503 when tables missing",
            "P2",
            "Negative",
            "Any",
            "Env without user_data_exports (staging only)",
            "1. POST data-export",
            "503 with migration guidance",
            "Skip on prod if migrated",
            "Manual",
        ),
    ],
)

add(
    "07 CVs Documents",
    "07 CVs Documents",
    [
        (
            "Student CV upload",
            "Student can upload PDF/DOC/DOCX CV",
            "P0",
            "Functional",
            "Student",
            "Signed in student",
            "1. My CVs → upload valid PDF",
            "Upload succeeds; appears in list; S3 object stored",
            "Magic-byte validation",
            "Candidate",
        ),
        (
            "Invalid file type",
            "Non-document upload is rejected",
            "P0",
            "Negative",
            "Student",
            "None",
            "1. Upload .exe or .txt disguised",
            "Clear validation error",
            None,
            "Candidate",
        ),
        (
            "Student CV download",
            "Student can view and download own CV",
            "P0",
            "Functional",
            "Student",
            "CV uploaded",
            "1. View\n2. Download attachment mode",
            "Presigned URL works; download=1 forces attachment when supported",
            None,
            "Candidate",
        ),
        (
            "College CV view",
            "College can list/view student CVs for tenant students",
            "P0",
            "Functional",
            "College Admin",
            "Student with CV in tenant",
            "1. Open student → CVs\n2. View",
            "List loads without 500; view works",
            "Regression for cv_soft_failure / SP_ACTIVE_CLAUSE",
            "Candidate",
        ),
        (
            "Employer resume",
            "Employer can view applicant resume when permitted",
            "P0",
            "Functional",
            "Employer",
            "Application with resume; partnership",
            "1. Applications → open resume",
            "Presigned view/download succeeds",
            "/api/employer/applications/resume",
            "Candidate",
        ),
        (
            "Authz cross-tenant",
            "College cannot view CV for other tenant's student",
            "P0",
            "Security",
            "College Admin",
            "Two campuses",
            "1. Call other-tenant student CV APIs",
            "401/403/404",
            None,
            "Manual",
        ),
        (
            "Documents upload",
            "Student document upload and list works",
            "P1",
            "Functional",
            "Student",
            "Documents page",
            "1. Upload allowed type\n2. List",
            "Stored and listed",
            None,
            "Candidate",
        ),
        (
            "Avatar/logo upload",
            "Student avatar and college/employer logo uploads succeed",
            "P2",
            "Functional",
            "Student; College Admin; Employer",
            "S3 configured",
            "1. Upload image via profile/settings",
            "Image displays after complete",
            "Not spreadsheet — binary upload",
            "Manual",
        ),
    ],
)

add(
    "08 Cross Cutting",
    "08 Cross Cutting",
    [
        (
            "5MB CSV limit",
            "Shared CSV upload max size enforced across import endpoints",
            "P1",
            "Negative",
            "College Admin; Employer",
            "Oversized CSV",
            "1. Hit students bulk-upload and assessment upload with >5MB",
            "Both reject with size error",
            "MAX_CSV_UPLOAD_BYTES",
            "Manual",
        ),
        (
            "Encoding BOM",
            "CSV with UTF-8 BOM imports correctly",
            "P1",
            "Functional",
            "College Admin",
            "Template from API (has BOM)",
            "1. Edit in Excel and save UTF-8 CSV\n2. Re-import",
            "Headers recognized; no Missing columns false positive",
            None,
            "Manual",
        ),
        (
            "Excel save quirks",
            "Excel-edited template retains Yes/No Verified and status enums",
            "P1",
            "Regression",
            "College Admin",
            "Excel available",
            "1. Open template in Excel\n2. Save as CSV\n3. Import",
            "Import succeeds; TRUE/FALSE coerced if present",
            None,
            "Manual",
        ),
        (
            "Concurrent import",
            "Two overlapping student imports for same rolls do not corrupt accounts",
            "P2",
            "Reliability",
            "College Admin",
            "Same CSV",
            "1. Fire two uploads close together",
            "No duplicate users; final state consistent",
            None,
            "Manual",
        ),
        (
            "Audit trail",
            "Successful student bulk import writes audit_logs entry",
            "P1",
            "Functional",
            "College Admin",
            "Import success",
            "1. Import\n2. Check audit / admin logs",
            "action student_bulk_import with counts",
            None,
            "Candidate",
        ),
        (
            "Not file-based",
            "Bulk notifications are not CSV import — document as N/A",
            "P2",
            "Documentation",
            "College Admin",
            "None",
            "1. Open bulk-notifications\n2. Confirm filter/send UI only",
            "No file upload control for notifications roster",
            "Mark N/A in coverage",
            "Manual",
        ),
        (
            "No XLSX runtime",
            "App has no end-user XLSX import/export — only CSV/ICS/PDF",
            "P2",
            "Documentation",
            "All",
            "None",
            "1. Scan Import/Export menus across roles",
            "No .xlsx upload/download in product UI",
            "QA xlsx scripts are tooling only",
            "Manual",
        ),
    ],
)


def style_header(ws, row=2):
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin


def autosize(ws):
    widths = {
        1: 12,
        2: 18,
        3: 18,
        4: 48,
        5: 8,
        6: 12,
        7: 18,
        8: 28,
        9: 40,
        10: 40,
        11: 32,
        12: 12,
        13: 10,
        14: 24,
        15: 18,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{ws.max_row}"


def main():
    wb = Workbook()
    # Index
    idx = wb.active
    idx.title = "Index"
    title_font = Font(name="Arial", bold=True, size=16, color="1E3A8A")
    head_font = Font(name="Arial", bold=True, size=11)
    body = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="2563EB", underline="single")
    idx["A1"] = "PlacementHub — Import / Export Test Cases"
    idx["A1"].font = title_font
    idx["A2"] = (
        "Covers CSV/ICS/PDF document import & export across College, Employer, Student, and Super Admin. "
        "Source: product import/export inventory (students, calendar, assessments, offers, list exports, audit, my-data, CVs)."
    )
    idx["A2"].font = body
    idx["A2"].alignment = Alignment(wrap_text=True)
    idx.merge_cells("A2:G2")
    idx.row_dimensions[2].height = 36

    headers_idx = ["#", "Module / Tab", "Case Count", "P0", "P1", "P2+", "Open"]
    for c, h in enumerate(headers_idx, 1):
        cell = idx.cell(4, c, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")

    total = 0
    p0 = p1 = p2 = 0
    for i, (sheet, module, rows) in enumerate(CASES, 1):
        counts = {"P0": 0, "P1": 0, "P2": 0}
        for r in rows:
            pr = r[2]
            if pr.startswith("P0"):
                counts["P0"] += 1
            elif pr.startswith("P1"):
                counts["P1"] += 1
            else:
                counts["P2"] += 1
        n = len(rows)
        total += n
        p0 += counts["P0"]
        p1 += counts["P1"]
        p2 += counts["P2"]
        row = 4 + i
        idx.cell(row, 1, i).font = body
        idx.cell(row, 2, sheet).font = body
        idx.cell(row, 3, n).font = body
        idx.cell(row, 4, counts["P0"]).font = body
        idx.cell(row, 5, counts["P1"]).font = body
        idx.cell(row, 6, counts["P2"]).font = body
        link = idx.cell(row, 7, "Open ›")
        link.font = link_font
        link.hyperlink = f"#'{sheet}'!A1"

    sum_row = 5 + len(CASES)
    idx.cell(sum_row, 1, "").font = body
    idx.cell(sum_row, 2, "TOTAL").font = head_font
    idx.cell(sum_row, 3, total).font = head_font
    idx.cell(sum_row, 4, p0).font = head_font
    idx.cell(sum_row, 5, p1).font = head_font
    idx.cell(sum_row, 6, p2).font = head_font

    idx.cell(sum_row + 2, 1, "How to use").font = head_font
    idx.cell(
        sum_row + 3,
        1,
        "1) Execute by module tab. 2) Fill Status / Actual Result / Executed At. "
        "3) Priority: P0 blocker, P1 important, P2 polish. "
        "4) Retired offer CSV uploads are included as expected-410 regressions.",
    ).font = body
    idx.merge_cells(start_row=sum_row + 3, start_column=1, end_row=sum_row + 3, end_column=7)
    idx.row_dimensions[sum_row + 3].height = 48

    for c, w in enumerate([6, 28, 12, 8, 8, 8, 10], 1):
        idx.column_dimensions[get_column_letter(c)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    font = Font(name="Arial", size=10)
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for sheet, module, rows in CASES:
        ws = wb.create_sheet(sheet)
        back = ws.cell(1, 1, "« Back to Index")
        back.font = link_font
        back.hyperlink = "#Index!A1"
        style_header(ws, 2)
        prefix = sheet.split(" ", 1)[0]
        for i, r in enumerate(rows, 1):
            feature, title, priority, typ, roles, pre, steps, expected, notes, auto = r
            values = [
                f"TC-IE-{prefix}-{i:03d}",
                module,
                feature,
                title,
                priority,
                typ,
                roles,
                pre,
                steps,
                expected,
                notes,
                auto,
                "",
                "",
                "",
            ]
            row_idx = 2 + i
            for col, val in enumerate(values, 1):
                cell = ws.cell(row_idx, col, val)
                cell.font = font
                cell.alignment = wrap
                cell.border = thin
            ws.row_dimensions[row_idx].height = 72
        autosize(ws)

    wb.save(OUT)
    print(f"Wrote {OUT} with {total} cases ({p0} P0 / {p1} P1 / {p2} P2+)")


if __name__ == "__main__":
    main()
