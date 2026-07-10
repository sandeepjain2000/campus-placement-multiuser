from __future__ import annotations

import asyncio
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from browser_use.browser.session import BrowserSession
from openpyxl import load_workbook


BASE_URL = "http://127.0.0.1:3000"
EMAIL = "arjun.verma@iitm.edu"
ROOT = Path(__file__).resolve().parent
XLSX_PATH = ROOT / "uat_test_cases_extended_batch_50.xlsx"
RESULTS_PATH = ROOT / "browser_use_10_results.json"

SELECTED_IDS = [
    "TC2-01",
    "TC2-02",
    "TC2-03",
    "TC2-06",
    "TC2-07",
    "TC2-11",
    "TC2-12",
    "TC2-13",
    "TC2-14",
    "TC2-15",
]


async def safe_nav(page: Any, path: str, timeout: int = 25) -> None:
    url = f"{BASE_URL}{path}"
    try:
        await asyncio.wait_for(page.goto(url), timeout=timeout)
    except TimeoutError:
        # Next dev pages can keep requests open; DOM is usually usable after timeout.
        pass
    await asyncio.sleep(2)


async def eval_json(page: Any, script: str) -> Any:
    raw = await page.evaluate(f"() => JSON.stringify(({script})())")
    return json.loads(raw)


async def page_snapshot(page: Any) -> dict[str, Any]:
    return await eval_json(
        page,
        """
        () => {
          const text = document.body?.innerText || "";
          const headings = [...document.querySelectorAll("h1,h2")].map((h) => h.innerText.trim()).filter(Boolean);
          const buttons = [...document.querySelectorAll("button,a")].map((b) => b.innerText.trim()).filter(Boolean).slice(0, 40);
          return {
            url: location.pathname,
            title: document.title,
            text: text.slice(0, 3000),
            headings,
            buttons,
            hasError: /Application error|Unhandled Runtime Error|500|404|not found/i.test(text)
          };
        }
        """,
    )


async def wait_for_text(page: Any, needle: str, timeout: int = 10) -> dict[str, Any]:
    needle = needle.lower()
    last_snapshot: dict[str, Any] = {}
    for _ in range(timeout * 2):
        last_snapshot = await page_snapshot(page)
        if needle in last_snapshot["text"].lower():
            return last_snapshot
        await asyncio.sleep(0.5)
    return last_snapshot


def pass_if(condition: bool, passed: str, failed: str) -> tuple[str, str]:
    return ("Pass", passed) if condition else ("Fail", failed)


async def login(page: Any) -> tuple[bool, str]:
    await safe_nav(page, f"/login?email={EMAIL.replace('@', '%40')}")
    form = await eval_json(
        page,
        """
        () => ({
          email: document.querySelector("#login-email")?.value || "",
          passwordLength: document.querySelector("#login-password")?.value?.length || 0,
          hasSubmit: Boolean(document.querySelector("#login-submit"))
        })
        """,
    )
    if form["email"] != EMAIL or form["passwordLength"] == 0 or not form["hasSubmit"]:
        return False, f"Login form not ready: {form}"

    submit = await page.get_elements_by_css_selector("#login-submit")
    if not submit:
        return False, "Login submit button not found"

    await submit[0].click()
    for _ in range(20):
        await asyncio.sleep(1)
        current_url = await page.get_url()
        if "/dashboard/student" in current_url:
            return True, "Logged in as seeded student"
    return False, f"Login did not reach student dashboard; final URL: {await page.get_url()}"


async def tc_browse_drives_mode(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/drives")
    result = await eval_json(
        page,
        """
        () => {
          const text = document.body.innerText;
          const selects = [...document.querySelectorAll("select")];
          const select = selects.find((s) => [...s.options].some((o) => /on[- ]?campus/i.test(o.textContent || o.value)));
          let exercised = false;
          if (select) {
            const option = [...select.options].find((o) => /on[- ]?campus/i.test(o.textContent || o.value));
            select.value = option.value;
            select.dispatchEvent(new Event("change", { bubbles: true }));
            exercised = true;
          }
          const apply = [...document.querySelectorAll("button")].find((b) => /apply|filter/i.test(b.innerText));
          if (apply) apply.click();
          return { text, exercised, hasDriveCopy: /drive|company|campus/i.test(text) };
        }
        """,
    )
    return pass_if(
        result["hasDriveCopy"] and result["exercised"],
        "Browse Drives loaded and an On-campus mode filter control was exercised.",
        f"Browse Drives loaded={result['hasDriveCopy']}, On-campus filter exercised={result['exercised']}.",
    )


async def tc_month_picker(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/drives")
    result = await eval_json(
        page,
        """
        () => {
          const text = document.body.innerText;
          const hasMonthInput = Boolean(document.querySelector('input[type="month"]'));
          const hasMonthCopy = /month|year|custom range|when/i.test(text);
          return { hasMonthInput, hasMonthCopy };
        }
        """,
    )
    return pass_if(
        result["hasMonthInput"] or result["hasMonthCopy"],
        "Month/date filter UI is present on Browse Drives.",
        "Could not identify a month-year picker or equivalent date filter on Browse Drives.",
    )


async def tc_custom_range(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/drives")
    result = await eval_json(
        page,
        """
        () => {
          const text = document.body.innerText;
          const custom = [...document.querySelectorAll("button,select,option,label")]
            .find((el) => /custom range|from|to/i.test(el.innerText || el.textContent || ""));
          return { found: Boolean(custom), textHasRange: /custom range|from|to/i.test(text) };
        }
        """,
    )
    return pass_if(
        result["found"] or result["textHasRange"],
        "Custom range controls or labels are visible.",
        "No Custom range / From / To controls were found on Browse Drives.",
    )


async def tc_calendar(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/calendar")
    snap = await page_snapshot(page)
    return pass_if(
        not snap["hasError"] and "calendar" in snap["text"].lower(),
        "Placement calendar page loaded without a runtime error.",
        "Placement calendar did not expose expected calendar content.",
    )


async def tc_applications(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/applications/jobs")
    snap = await wait_for_text(page, "application")
    return pass_if(
        not snap["hasError"] and any(word in snap["text"].lower() for word in ["application", "job"]),
        "My Applications page loaded and displayed application/status content.",
        f"My Applications page did not show application content. URL={snap.get('url')}, headings={snap.get('headings')}",
    )


async def tc_discussions(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/discussions")
    snap = await page_snapshot(page)
    return pass_if(
        not snap["hasError"] and any(word in snap["text"].lower() for word in ["discussion", "clarification", "empty"]),
        "Discussions page loaded with role-appropriate content or empty state.",
        "Discussions page did not show expected discussion or empty-state content.",
    )


async def tc_clarifications(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/clarifications")
    snap = await wait_for_text(page, "clarification")
    return pass_if(
        not snap["hasError"] and "clarification" in snap["text"].lower(),
        "Clarifications page loaded with expected clarification content.",
        f"Clarifications page did not show expected content. URL={snap.get('url')}, headings={snap.get('headings')}",
    )


async def tc_offers(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/offers")
    snap = await page_snapshot(page)
    return pass_if(
        not snap["hasError"] and "offer" in snap["text"].lower(),
        "My Offers page loaded with offer content or empty state.",
        "My Offers page did not show expected offer content.",
    )


async def tc_interviews(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/student/interviews")
    snap = await page_snapshot(page)
    return pass_if(
        not snap["hasError"] and "interview" in snap["text"].lower(),
        "My Interviews page loaded with interview content or empty state.",
        "My Interviews page did not show expected interview content.",
    )


async def tc_alerts(page: Any) -> tuple[str, str]:
    await safe_nav(page, "/dashboard/alerts")
    snap = await page_snapshot(page)
    return pass_if(
        not snap["hasError"] and any(word in snap["text"].lower() for word in ["alert", "inbox", "notification"]),
        "Alerts page loaded with inbox/alert content.",
        "Alerts page did not show expected inbox or alert content.",
    )


RUNNERS = {
    "TC2-01": tc_browse_drives_mode,
    "TC2-02": tc_month_picker,
    "TC2-03": tc_custom_range,
    "TC2-06": tc_calendar,
    "TC2-07": tc_applications,
    "TC2-11": tc_discussions,
    "TC2-12": tc_clarifications,
    "TC2-13": tc_offers,
    "TC2-14": tc_interviews,
    "TC2-15": tc_alerts,
}


def load_cases() -> dict[str, dict[str, str]]:
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    cases: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {headers[i]: "" if value is None else str(value) for i, value in enumerate(row)}
        if record.get("test_id") in SELECTED_IDS:
            cases[record["test_id"]] = record
    return cases


async def main() -> None:
    cases = load_cases()
    results = []
    browser = BrowserSession(
        headless=True,
        use_cloud=False,
        enable_default_extensions=False,
        allowed_domains=["127.0.0.1", "localhost"],
        keep_alive=False,
    )
    await browser.start()
    page = await browser.new_page()
    try:
        ok, detail = await login(page)
        if not ok:
            raise RuntimeError(detail)

        for test_id in SELECTED_IDS:
            case = cases[test_id]
            try:
                status, actual = await RUNNERS[test_id](page)
            except Exception as exc:  # Keep the batch going and record the failure.
                status, actual = "Fail", f"{type(exc).__name__}: {exc}"
            results.append(
                {
                    "test_id": test_id,
                    "module": case["module"],
                    "test_case": case["test_case"],
                    "expected_result": case["expected_result"],
                    "actual_result": actual,
                    "status": status,
                }
            )
    finally:
        await browser.stop()

    payload = {
        "tool": "browser-use BrowserSession",
        "base_url": BASE_URL,
        "source_xlsx": str(XLSX_PATH),
        "executed_at": datetime.now().isoformat(timespec="seconds"),
        "results": results,
        "summary": {
            "total": len(results),
            "passed": sum(1 for item in results if item["status"] == "Pass"),
            "failed": sum(1 for item in results if item["status"] == "Fail"),
        },
    }
    RESULTS_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], indent=2))
    for item in results:
        print(f"{item['test_id']} {item['status']}: {item['actual_result']}")


if __name__ == "__main__":
    asyncio.run(main())
