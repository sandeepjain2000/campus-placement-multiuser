import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PASSWORD = "Admin@123"
DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

OUTPUT_JSON = Path("qa/C-6-PDF-Report-Availability-and-Correctness.json")
OUTPUT_XLSX = Path("qa/C-6-PDF-Report-Availability-and-Correctness.xlsx")


TRANSACTIONAL_SCREENS = [
    {"role": "student", "route": "/dashboard/student/applications", "module": "Student Applications"},
    {"role": "student", "route": "/dashboard/student/interviews", "module": "Student Interviews"},
    {"role": "student", "route": "/dashboard/student/offers", "module": "Student Offers"},
    {"role": "employer", "route": "/dashboard/employer/jobs", "module": "Employer Jobs"},
    {"role": "employer", "route": "/dashboard/employer/drives", "module": "Employer Drives"},
    {"role": "employer", "route": "/dashboard/employer/applications", "module": "Employer Applications"},
    {"role": "employer", "route": "/dashboard/employer/interviews", "module": "Employer Interviews"},
    {"role": "employer", "route": "/dashboard/employer/offers", "module": "Employer Offers"},
    {"role": "employer", "route": "/dashboard/employer/assessment-summary", "module": "Employer Assessment Summary"},
    {"role": "college_admin", "route": "/dashboard/college/drives", "module": "College Drives"},
    {"role": "college_admin", "route": "/dashboard/college/applications", "module": "College Applications"},
    {"role": "college_admin", "route": "/dashboard/college/interviews", "module": "College Interviews"},
    {"role": "college_admin", "route": "/dashboard/college/offers", "module": "College Offers"},
    {"role": "college_admin", "route": "/dashboard/college/reports", "module": "College Reports"},
    {"role": "college_admin", "route": "/dashboard/college/audit-reports", "module": "College Audit Reports"},
    {"role": "super_admin", "route": "/dashboard/admin/overview", "module": "Admin Overview"},
    {"role": "super_admin", "route": "/dashboard/admin/audit-reports", "module": "Admin Audit Reports"},
]


def make_presence_case(idx, screen):
    role = screen["role"]
    email = DEMO_USERS[role]
    return {
        "testCaseId": f"C6-PDF-PRESENCE-{idx:03d}",
        "category": "C-6",
        "type": "pdf_presence",
        "title": f"{screen['module']} has PDF report option",
        "role": role,
        "route": screen["route"],
        "demoAccount": {"email": email, "password": PASSWORD},
        "steps": [
            f"Login as {role}",
            f"Open {screen['route']}",
            "Identify report/export actions",
            "Verify a PDF-specific action exists (label or format selector includes PDF)",
        ],
        "expectedResult": [
            "PDF report/download option is visible and enabled",
            "Option is accessible for relevant records/filters",
        ],
        "priority": "critical",
    }


def make_functional_case(idx, screen):
    role = screen["role"]
    email = DEMO_USERS[role]
    return {
        "testCaseId": f"C6-PDF-FUNC-{idx:03d}",
        "category": "C-6",
        "type": "pdf_generation_and_sanity",
        "title": f"{screen['module']} PDF generates and is correct",
        "role": role,
        "route": screen["route"],
        "demoAccount": {"email": email, "password": PASSWORD},
        "steps": [
            f"Login as {role} and open {screen['route']}",
            "Apply filter/record selection for deterministic output",
            "Trigger PDF report generation/download",
            "Open downloaded file and confirm valid PDF opens",
            "Verify report includes correct title/date and 3-5 sampled values matching UI",
        ],
        "expectedResult": [
            "PDF file downloads without server/client errors",
            "File opens successfully and is not corrupted/blank",
            "Sampled values in PDF match UI data",
            "Report metadata (module/date/context) is correct",
        ],
        "priority": "high",
    }


def build_cases():
    cases = []
    for i, screen in enumerate(TRANSACTIONAL_SCREENS, start=1):
        cases.append(make_presence_case(i, screen))
    for i, screen in enumerate(TRANSACTIONAL_SCREENS, start=1):
        cases.append(make_functional_case(i, screen))

    cases.append(
        {
            "testCaseId": "C6-GAP-001",
            "category": "C-6",
            "type": "gap_register",
            "title": "PDF coverage gap register for transactional screens",
            "role": "all_roles",
            "route": "all transactional routes",
            "demoAccount": {"email": "use role-wise demo account", "password": PASSWORD},
            "steps": [
                "Compile transactional screen list role-wise",
                "Mark each screen as PDF Available / Missing",
                "For Missing, capture reason and suggested implementation point",
            ],
            "expectedResult": [
                "Complete PDF availability matrix exists",
                "All missing PDF options are explicitly tracked as actionable gaps",
            ],
            "priority": "critical",
        }
    )
    return cases


def write_json(cases):
    payload = {
        "category": "C-6",
        "categoryTitle": "PDF report option availability and correctness",
        "description": "Validate every transactional screen for PDF report option presence and correctness of generated PDF content.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-6 Manual Cases"
    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Role",
        "Route",
        "Demo Email",
        "Demo Password",
        "Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["type"],
                case["role"],
                case["route"],
                case["demoAccount"]["email"],
                case["demoAccount"]["password"],
                "\n".join(case["steps"]),
                "\n".join(case["expectedResult"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 20,
        "B": 10,
        "C": 24,
        "D": 14,
        "E": 42,
        "F": 30,
        "G": 14,
        "H": 54,
        "I": 50,
        "J": 10,
        "K": 12,
        "L": 30,
        "M": 30,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-6 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
