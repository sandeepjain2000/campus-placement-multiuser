import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PASSWORD = "Admin@123"
DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

OUTPUT_JSON = Path("qa/C-12-Use-Case-Flow-Executability.json")
OUTPUT_XLSX = Path("qa/C-12-Use-Case-Flow-Executability.xlsx")


USE_CASES = [
    {
        "id": "UC-001",
        "name": "Employer onboarding and campus partnership",
        "actors": ["employer", "college_admin"],
        "primaryActor": "employer",
        "preconditions": ["Employer account exists", "Target college exists"],
        "mainFlow": [
            "Employer logs in and opens campus selection",
            "Employer sends partnership request",
            "College admin reviews request",
            "College admin approves request",
            "Employer sees campus as approved/active",
        ],
        "successEvidence": ["Approval status visible to both roles", "Partnership appears in respective lists"],
    },
    {
        "id": "UC-002",
        "name": "Create and publish placement drive",
        "actors": ["employer", "college_admin", "student"],
        "primaryActor": "employer",
        "preconditions": ["Employer approved for campus", "Job posting exists or can be created"],
        "mainFlow": [
            "Employer creates drive request",
            "College admin reviews and approves/schedules drive",
            "Student opens drives list",
            "Student can view approved drive details",
        ],
        "successEvidence": ["Drive state transitions correctly", "Drive is visible to eligible students"],
    },
    {
        "id": "UC-003",
        "name": "Student application lifecycle",
        "actors": ["student", "employer", "college_admin"],
        "primaryActor": "student",
        "preconditions": ["Approved drive visible to student"],
        "mainFlow": [
            "Student applies to drive",
            "Employer sees application in pipeline",
            "Employer updates status (shortlist/reject/progress)",
            "Student and college see updated status",
        ],
        "successEvidence": ["Application record appears across role views", "Status is synchronized"],
    },
    {
        "id": "UC-004",
        "name": "Interview scheduling and visibility",
        "actors": ["employer", "student", "college_admin"],
        "primaryActor": "employer",
        "preconditions": ["At least one shortlisted/in-progress candidate"],
        "mainFlow": [
            "Employer creates interview slot",
            "Student sees interview schedule",
            "College admin sees interview schedule",
            "Employer edits schedule and all viewers see update",
        ],
        "successEvidence": ["Interview appears in all intended calendars/lists", "Edit propagates"],
    },
    {
        "id": "UC-005",
        "name": "Offer release and response",
        "actors": ["employer", "student", "college_admin"],
        "primaryActor": "employer",
        "preconditions": ["Candidate reaches offer stage"],
        "mainFlow": [
            "Employer releases offer",
            "Student sees offer details",
            "Student accepts/rejects offer",
            "Employer and college see final response state",
        ],
        "successEvidence": ["Offer visibility across roles", "Response persisted and reflected"],
    },
    {
        "id": "UC-006",
        "name": "Assessment upload to decision support",
        "actors": ["employer", "college_admin"],
        "primaryActor": "employer",
        "preconditions": ["Assessment template available"],
        "mainFlow": [
            "Employer uploads assessment CSV",
            "Upload rows are accepted",
            "Assessment summary reflects uploaded records",
            "College hiring-assessment view reflects mapped outcomes",
        ],
        "successEvidence": ["No processing errors", "Rows visible in downstream views"],
    },
    {
        "id": "UC-007",
        "name": "College offers upload and student visibility",
        "actors": ["college_admin", "student"],
        "primaryActor": "college_admin",
        "preconditions": ["Offer CSV template available"],
        "mainFlow": [
            "College uploads offers CSV",
            "Offers appear in college offers table",
            "Concerned student sees offer",
        ],
        "successEvidence": ["Uploaded rows persist", "Student sees applicable records"],
    },
    {
        "id": "UC-008",
        "name": "Clarification and discussion communication",
        "actors": ["college_admin", "student", "employer"],
        "primaryActor": "college_admin",
        "preconditions": ["Communication modules enabled"],
        "mainFlow": [
            "College publishes clarification",
            "Student sees clarification entry",
            "Employer sees related discussion/communication context",
        ],
        "successEvidence": ["Message visibility aligns with audience"],
    },
    {
        "id": "UC-009",
        "name": "Profile completion and document handling",
        "actors": ["student"],
        "primaryActor": "student",
        "preconditions": ["Student logged in"],
        "mainFlow": [
            "Student updates profile fields",
            "Student uploads resume/document",
            "Profile and documents persist after refresh",
        ],
        "successEvidence": ["Updated profile fields visible", "Document listed and retrievable"],
    },
    {
        "id": "UC-010",
        "name": "Admin governance and audit visibility",
        "actors": ["super_admin"],
        "primaryActor": "super_admin",
        "preconditions": ["Platform has multi-tenant data"],
        "mainFlow": [
            "Admin reviews platform overview",
            "Admin checks users/colleges/employers lists",
            "Admin inspects audit reports",
            "Admin validates traceability for key operations",
        ],
        "successEvidence": ["Governance screens load with current data", "Audit artifacts accessible"],
    },
]


def build_cases():
    cases = []
    for idx, uc in enumerate(USE_CASES, start=1):
        primary = uc["primaryActor"]
        cases.append(
            {
                "testCaseId": f"C12-UC-{idx:03d}",
                "category": "C-12",
                "type": "use_case_flow_executability",
                "useCaseId": uc["id"],
                "useCaseName": uc["name"],
                "actors": uc["actors"],
                "primaryActor": primary,
                "primaryActorAccount": {"email": DEMO_USERS.get(primary, "n/a"), "password": PASSWORD},
                "preconditions": uc["preconditions"],
                "steps": uc["mainFlow"],
                "expectedResult": [
                    "Entire use case flow executes without blocking errors",
                    "All mandatory handoffs between actors succeed",
                    "Final success evidence is observable in UI/data",
                ]
                + uc["successEvidence"],
                "priority": "critical" if idx <= 7 else "high",
            }
        )

    cases.append(
        {
            "testCaseId": "C12-AUDIT-001",
            "category": "C-12",
            "type": "use_case_catalog_governance",
            "useCaseId": "UC-AUDIT",
            "useCaseName": "Use case catalog coverage and gap register",
            "actors": ["all_roles"],
            "primaryActor": "qa",
            "primaryActorAccount": {"email": "n/a", "password": "n/a"},
            "preconditions": ["Current feature list available"],
            "steps": [
                "Prepare complete product use-case inventory",
                "Map each use case to at least one executable test case",
                "Run smoke pass for each use case flow",
                "Log failed handoffs and missing paths",
            ],
            "expectedResult": [
                "No core use case remains untested",
                "Gap list exists for non-executable or partially executable flows",
            ],
            "priority": "critical",
        }
    )
    return cases


def write_json(cases):
    payload = {
        "category": "C-12",
        "categoryTitle": "Use case flow executability",
        "description": "Validate that each defined business use case can be executed end-to-end across relevant actors.",
        "useCaseCatalog": USE_CASES,
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-12 Use Case Flows"

    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Use Case ID",
        "Use Case Name",
        "Actors",
        "Primary Actor",
        "Primary Actor Email",
        "Preconditions",
        "Flow Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for c in cases:
        ws.append(
            [
                c["testCaseId"],
                c["category"],
                c["type"],
                c["useCaseId"],
                c["useCaseName"],
                ", ".join(c["actors"]),
                c["primaryActor"],
                c["primaryActorAccount"]["email"],
                "\n".join(c["preconditions"]),
                "\n".join(c["steps"]),
                "\n".join(c["expectedResult"]),
                c["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    widths = {
        "A": 14,
        "B": 10,
        "C": 26,
        "D": 12,
        "E": 34,
        "F": 26,
        "G": 14,
        "H": 30,
        "I": 34,
        "J": 44,
        "K": 44,
        "L": 10,
        "M": 12,
        "N": 28,
        "O": 28,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-12 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
