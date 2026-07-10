import json
import re
import openpyxl
from pathlib import Path

def get_test_status(results):
    status_map = {}
    
    def traverse(node):
        if isinstance(node, dict):
            if 'title' in node and 'C4-' in node['title'] and 'tests' in node:
                match = re.search(r'(C4-[A-Z]+-\d+)', node['title'])
                if match:
                    test_id = match.group(1)
                    # A test might have multiple runs (e.g. retries), get the outcome
                    tests = node.get('tests', [])
                    if tests:
                        # Find the status of the last run
                        results_array = tests[0].get('results', [])
                        if results_array:
                            last_status = results_array[-1].get('status')
                            
                            # Note: Playwright json reporter maps status to 'passed', 'failed', 'timedOut', etc.
                            if last_status == 'passed':
                                status_map[test_id] = 'Passed'
                            elif last_status in ('failed', 'timedOut'):
                                status_map[test_id] = 'Failed'
                            else:
                                status_map[test_id] = 'Not Run'
            
            for v in node.values():
                traverse(v)
        elif isinstance(node, list):
            for item in node:
                traverse(item)
                
    traverse(results)
    return status_map

def update_excel():
    json_file = 'c4_results.json'
    excel_file = 'C-4-Non-Hardcoded-Data-Validation.xlsx'

    if not Path(json_file).exists():
        print(f"Results file {json_file} not found.")
        return

    try:
        with open(json_file, 'r', encoding='utf-16') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Failed to load json: {e}")
        return

    results_map = get_test_status(data)
    print(f"Parsed {len(results_map)} results from JSON: {list(results_map.keys())}")

    if not Path(excel_file).exists():
        print(f"Excel file {excel_file} not found.")
        return

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    headers = {cell.value: idx for idx, cell in enumerate(sheet[1]) if cell.value}
    
    id_col = headers.get('testCaseId', headers.get('Test Case ID'))
    if id_col is None:
        for idx, cell in enumerate(sheet[1]):
            if str(cell.value).lower().replace(' ', '') == 'testcaseid':
                id_col = idx
                break
    
    if id_col is None:
        print("Could not find Test Case ID column.")
        return

    status_col = None
    notes_col = None
    
    for k, v in headers.items():
        if k and 'status' in str(k).lower(): status_col = v
        if k and 'note' in str(k).lower() or 'comment' in str(k).lower(): notes_col = v

    if status_col is None:
        status_col = len(headers)
        sheet.cell(row=1, column=status_col+1).value = 'Status'
    if notes_col is None:
        notes_col = len(headers) + 1
        sheet.cell(row=1, column=notes_col+1).value = 'Execution Notes'

    updated_count = 0
    for row in sheet.iter_rows(min_row=2):
        row_id = row[id_col].value
        if row_id in results_map:
            row[status_col].value = results_map[row_id]
            row[notes_col].value = "Automated test execution"
            updated_count += 1

    wb.save(excel_file)
    print(f"Successfully updated {updated_count} test cases in {excel_file}")

if __name__ == '__main__':
    update_excel()
