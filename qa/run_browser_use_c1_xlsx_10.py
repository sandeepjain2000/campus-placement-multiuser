from __future__ import annotations

import asyncio
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from browser_use.browser.session import BrowserSession
from openpyxl import load_workbook


BASE_URL = "http://127.0.0.1:3000"
ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = ROOT / "C-1-Role-Based-Screen-Access.xlsx"
RESULTS_PATH = ROOT / "browser_use_c1_xlsx_10_results.json"


async def safe_nav(page: Any, path: str, timeout: int = 25) -> None:
    try:
        await asyncio.wait_for(page.goto(f"{BASE_URL}{path}"), timeout=timeout)
    except TimeoutError:
        pass
    await asyncio.sleep(1.5)


async def eval_json(page: Any, script: str) -> Any:
    raw = await page.evaluate(f"() => JSON.stringify(({script})())")
    return json.loads(raw)


async def snapshot(page: Any) -> dict[str, Any]:
    return await eval_json(
        page,
        """
        () => {
          const text = document.body?.innerText || "";
          return {
            url: location.pathname,
            title: document.title,
            text: text.slice(0, 2500),
            headings: [...document.querySelectorAll("h1,h2")].map((h) => h.innerText.trim()).filter(Boolean),
            hasRuntimeError: /Application error|Unhandled Runtime Error|TypeError|ReferenceError|500|404|not found/i.test(text),
            authBlocked: /unauthorized|access denied|sign in|login/i.test(text) && location.pathname.includes("/login")
          };
        }
        """,
    )


async def wait_for_rendered_snapshot(page: Any, timeout: int = 15) -> dict[str, Any]:
    last: dict[str, Any] = {}
    for _ in range(timeout * 2):
        last = await snapshot(page)
        if last["hasRuntimeError"] or len(last["text"].strip()) > 80:
            return last
        await asyncio.sleep(0.5)
    return last


async def login(page: Any, email: str) -> None:
    encoded = email.replace("@", "%40")
    await safe_nav(page, f"/login?email={encoded}")
    ready = {"email": "", "passwordLength": 0}
    for _ in range(20):
        ready = await eval_json(
            page,
            """
            () => ({
              email: document.querySelector("#login-email")?.value || "",
              passwordLength: document.querySelector("#login-password")?.value?.length || 0
            })
            """,
        )
        if ready["email"] == email and ready["passwordLength"] > 0:
            break
        await asyncio.sleep(0.5)
    else:
        raise RuntimeError(f"Login form was not prefilled for {email}: {ready}")
    buttons = await page.get_elements_by_css_selector("#login-submit")
    if not buttons:
        raise RuntimeError("Login submit button not found")
    await buttons[0].click()
    for _ in range(20):
        await asyncio.sleep(1)
        if "/dashboard" in await page.get_url():
            return
    raise RuntimeError(f"Login failed for {email}; final URL={await page.get_url()}")


def load_first_10_cases() -> list[dict[str, str]]:
    wb = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[i]: "" if value is None else str(value) for i, value in enumerate(values)}
        if row["Type"] == "positive_navigation":
            rows.append(row)
        if len(rows) == 10:
            return rows
    return rows


async def main() -> None:
    cases = load_first_10_cases()
    if len(cases) < 10:
        raise RuntimeError(f"Expected 10 positive navigation cases, found {len(cases)}")

    browser = BrowserSession(
        headless=True,
        use_cloud=False,
        enable_default_extensions=False,
        allowed_domains=["127.0.0.1", "localhost"],
        keep_alive=False,
    )
    await browser.start()
    page = await browser.new_page()
    results: list[dict[str, str]] = []
    current_email = ""
    try:
        for case in cases:
            email = case["Demo Email"]
            if email != current_email:
                await login(page, email)
                current_email = email

            await safe_nav(page, case["Target Route"])
            info = await wait_for_rendered_snapshot(page)
            passed = (
                not info["hasRuntimeError"]
                and not info["authBlocked"]
                and info["url"] != "/login"
                and len(info["text"].strip()) > 80
            )
            results.append(
                {
                    "test_case_id": case["Test Case ID"],
                    "role": case["Role"],
                    "route": case["Target Route"],
                    "expected_result": case["Expected Result"],
                    "actual_result": (
                        f"Loaded {info['url']} with headings={info['headings'][:3]}"
                        if passed
                        else f"Route check failed: {info}"
                    ),
                    "status": "Pass" if passed else "Fail",
                }
            )
    finally:
        await browser.stop()

    payload = {
        "tool": "browser-use BrowserSession",
        "base_url": BASE_URL,
        "source_xlsx": str(SOURCE_XLSX),
        "executed_at": datetime.now().isoformat(timespec="seconds"),
        "results": results,
        "summary": {
            "total": len(results),
            "passed": sum(1 for item in results if item["status"] == "Pass"),
            "failed": sum(1 for item in results if item["status"] == "Fail"),
        },
    }
    RESULTS_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], indent=2))
    for item in results:
        line = f"{item['test_case_id']} {item['status']}: {item['route']} -> {item['actual_result']}"
        print(line.encode("ascii", "backslashreplace").decode("ascii"))


if __name__ == "__main__":
    asyncio.run(main())
