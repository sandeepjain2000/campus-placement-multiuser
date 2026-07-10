import json
import openpyxl

def update_excel():
    json_file = 'c1_150_260_results.json'
    excel_file = 'C-1-Role-Based-Screen-Access.xlsx'

    with open(json_file, 'r') as f:
        results = json.load(f)

    # Convert results array to a dictionary keyed by ID
    results_map = {item['id']: item for item in results}

    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}
    
    # Locate column indices
    id_col = headers.get('testCaseId', headers.get('Test Case ID'))
    if id_col is None:
        # Fallback to searching first row visually if names differ
        for idx, cell in enumerate(sheet[1]):
            if str(cell.value).lower().replace(' ', '') == 'testcaseid':
                id_col = idx
                break
    
    if id_col is None:
        print("Could not find Test Case ID column.")
        return

    status_col = None
    notes_col = None
    
    for k, v in headers.items():
        if k and 'status' in str(k).lower(): status_col = v
        if k and 'note' in str(k).lower() or 'comment' in str(k).lower(): notes_col = v

    # Add columns if they don't exist
    if status_col is None:
        status_col = len(headers)
        sheet.cell(row=1, column=status_col+1).value = 'Status'
    if notes_col is None:
        notes_col = len(headers) + 1
        sheet.cell(row=1, column=notes_col+1).value = 'Execution Notes'

    updated_count = 0
    for row in sheet.iter_rows(min_row=2):
        row_id = row[id_col].value
        if row_id in results_map:
            result = results_map[row_id]
            row[status_col].value = result['status']
            import re
            clean_notes = re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', str(result['notes']))
            row[notes_col].value = clean_notes
            updated_count += 1

    wb.save(excel_file)
    print(f"Successfully updated {updated_count} test cases in {excel_file}")

if __name__ == '__main__':
    update_excel()
