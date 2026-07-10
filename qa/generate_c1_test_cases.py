import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


ROLE_SCREENS = {
    "student": [
        "/dashboard/student/overview",
        "/dashboard/student/profile",
        "/dashboard/my-exports",
        "/dashboard/alerts",
        "/dashboard/student/drives",
        "/dashboard/student/internships",
        "/dashboard/student/projects",
        "/dashboard/student/calendar",
        "/dashboard/student/applications",
        "/dashboard/student/interviews",
        "/dashboard/student/offers",
        "/dashboard/student/clarifications",
        "/dashboard/student/discussions",
        "/dashboard/student/documents",
        "/dashboard/feedback",
    ],
    "employer": [
        "/dashboard/employer/overview",
        "/dashboard/employer/select-campus",
        "/dashboard/my-exports",
        "/dashboard/alerts",
        "/dashboard/employer/calendar",
        "/dashboard/employer/profile",
        "/dashboard/employer/sponsorships",
        "/dashboard/employer/campus-guest-needs",
        "/dashboard/employer/internships",
        "/dashboard/employer/projects",
        "/dashboard/employer/jobs",
        "/dashboard/employer/drives",
        "/dashboard/employer/hiring-assessment",
        "/dashboard/employer/interviews",
        "/dashboard/employer/assessment-summary",
        "/dashboard/employer/applications",
        "/dashboard/employer/assessment-uploads",
        "/dashboard/employer/offers",
        "/dashboard/employer/offers-upload",
        "/dashboard/employer/discussions",
        "/dashboard/feedback",
    ],
    "college_admin": [
        "/dashboard/college/overview",
        "/dashboard/my-exports",
        "/dashboard/alerts",
        "/dashboard/college/employers",
        "/dashboard/college/employers/requests",
        "/dashboard/college/drives",
        "/dashboard/college/internships",
        "/dashboard/college/internship-results",
        "/dashboard/college/sponsorships",
        "/dashboard/college/students",
        "/dashboard/college/enrollment-key",
        "/dashboard/college/applications",
        "/dashboard/college/offers",
        "/dashboard/college/offers-upload",
        "/dashboard/college/hiring-assessment",
        "/dashboard/college/interviews",
        "/dashboard/college/clarifications",
        "/dashboard/college/discussions",
        "/dashboard/college/calendar",
        "/dashboard/college/events",
        "/dashboard/college/guest-engagements",
        "/dashboard/college/rules",
        "/dashboard/college/infrastructure",
        "/dashboard/college/settings",
        "/dashboard/college/reports",
        "/dashboard/college/audit-reports",
        "/dashboard/feedback",
    ],
    "super_admin": [
        "/dashboard/admin/overview",
        "/dashboard/my-exports",
        "/dashboard/admin/colleges",
        "/dashboard/admin/pending-registrations",
        "/dashboard/admin/employers",
        "/dashboard/admin/users",
        "/dashboard/admin/feedback",
        "/dashboard/admin/audit-reports",
        "/dashboard/admin/settings",
    ],
}

DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

PASSWORD = "Admin@123"
OUTPUT_JSON = Path("qa/C-1-Role-Based-Screen-Access.json")
OUTPUT_XLSX = Path("qa/C-1-Role-Based-Screen-Access.xlsx")


def build_cases():
    roles = list(ROLE_SCREENS.keys())
    all_routes = sorted({route for routes in ROLE_SCREENS.values() for route in routes})
    cases = []
    case_num = 1

    for role in roles:
        allowed = set(ROLE_SCREENS[role])
        blocked = [r for r in all_routes if r not in allowed]
        demo_email = DEMO_USERS[role]

        for route in sorted(allowed):
            cases.append(
                {
                    "testCaseId": f"C1-VISIT-{case_num:03d}",
                    "category": "C-1",
                    "title": f"{role} can open {route}",
                    "type": "positive_navigation",
                    "role": role,
                    "demoAccount": {"email": demo_email, "password": PASSWORD},
                    "steps": [
                        "Login with demo account",
                        f"Navigate to {route}",
                        "Wait for page to load completely",
                    ],
                    "expected": [
                        "Target screen opens successfully",
                        "No runtime error banner/toast/modal is shown",
                        "No crash/blank screen is observed",
                    ],
                    "forbiddenRoutesSample": blocked[:5],
                    "priority": "high",
                }
            )
            case_num += 1

        for blocked_route in blocked:
            cases.append(
                {
                    "testCaseId": f"C1-BLOCK-{case_num:03d}",
                    "category": "C-1",
                    "title": f"{role} cannot open {blocked_route}",
                    "type": "negative_authorization",
                    "role": role,
                    "demoAccount": {"email": demo_email, "password": PASSWORD},
                    "steps": [
                        "Login with demo account",
                        f"Directly open URL {blocked_route}",
                    ],
                    "expected": [
                        "Access is denied OR user is redirected to role-appropriate screen",
                        "Wrong screen is not accessible",
                        "No server/client error is shown",
                    ],
                    "priority": "high",
                }
            )
            case_num += 1

    return cases


def write_json(cases):
    payload = {
        "category": "C-1",
        "categoryTitle": "Role-based screen access and open stability",
        "description": "Validate each role can open only allowed screens, all allowed screens load without errors, and disallowed screens are blocked.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-1 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Role",
        "Demo Email",
        "Demo Password",
        "Target Route",
        "Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        route = ""
        for step in case["steps"]:
            if "Navigate to " in step:
                route = step.replace("Navigate to ", "").strip()
            if "Directly open URL " in step:
                route = step.replace("Directly open URL ", "").strip()
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["type"],
                case["role"],
                case["demoAccount"]["email"],
                case["demoAccount"]["password"],
                route,
                "\n".join(case["steps"]),
                "\n".join(case["expected"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 24,
        "D": 14,
        "E": 32,
        "F": 14,
        "G": 42,
        "H": 48,
        "I": 52,
        "J": 10,
        "K": 12,
        "L": 30,
        "M": 30,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
