import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PASSWORD = "Admin@123"
DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

OUTPUT_JSON = Path("qa/C-9-Transactional-Lifecycle-Add-Edit-Delete-View-Cancel.json")
OUTPUT_XLSX = Path("qa/C-9-Transactional-Lifecycle-Add-Edit-Delete-View-Cancel.xlsx")


ENTITIES = [
    {
        "entity": "Placement Drive",
        "ownerRole": "employer",
        "ownerRoute": "/dashboard/employer/drives",
        "impactedRoles": ["college_admin", "student"],
        "impactedRoutes": ["/dashboard/college/drives", "/dashboard/student/drives"],
        "deletePermissible": "conditional",
    },
    {
        "entity": "Job Posting",
        "ownerRole": "employer",
        "ownerRoute": "/dashboard/employer/jobs",
        "impactedRoles": ["college_admin", "student"],
        "impactedRoutes": ["/dashboard/college/drives", "/dashboard/student/drives"],
        "deletePermissible": "conditional",
    },
    {
        "entity": "Application",
        "ownerRole": "student",
        "ownerRoute": "/dashboard/student/applications",
        "impactedRoles": ["employer", "college_admin"],
        "impactedRoutes": ["/dashboard/employer/applications", "/dashboard/college/applications"],
        "deletePermissible": "no_hard_delete_expected",
    },
    {
        "entity": "Interview Schedule",
        "ownerRole": "employer",
        "ownerRoute": "/dashboard/employer/interviews",
        "impactedRoles": ["student", "college_admin"],
        "impactedRoutes": ["/dashboard/student/interviews", "/dashboard/college/interviews"],
        "deletePermissible": "conditional",
    },
    {
        "entity": "Offer",
        "ownerRole": "employer",
        "ownerRoute": "/dashboard/employer/offers",
        "impactedRoles": ["student", "college_admin"],
        "impactedRoutes": ["/dashboard/student/offers", "/dashboard/college/offers"],
        "deletePermissible": "restricted_after_acceptance",
    },
]


def make_crudv_case(i, row):
    owner = row["ownerRole"]
    return {
        "testCaseId": f"C9-CRUDV-{i:03d}",
        "category": "C-9",
        "type": "add_edit_delete_view_lifecycle",
        "title": f"{row['entity']} supports Add/Edit/View and governed Delete",
        "entity": row["entity"],
        "ownerRole": owner,
        "ownerRoute": row["ownerRoute"],
        "ownerAccount": {"email": DEMO_USERS[owner], "password": PASSWORD},
        "deletePermissibleRule": row["deletePermissible"],
        "steps": [
            f"Login as {owner} and open {row['ownerRoute']}",
            f"Add new {row['entity']} record",
            "View created record in listing/details",
            "Edit at least one business field and save",
            "Verify updated value is visible",
            "Attempt Delete and verify behavior follows integrity rule",
        ],
        "expectedResult": [
            "Add works and record is persisted",
            "View shows latest saved record",
            "Edit updates are reflected correctly",
            "Delete is allowed only where business integrity permits",
        ],
        "priority": "critical",
    }


def make_cancel_vs_delete_case(i, row):
    owner = row["ownerRole"]
    return {
        "testCaseId": f"C9-CANCEL-DELETE-{i:03d}",
        "category": "C-9",
        "type": "cancel_vs_delete_behavior",
        "title": f"{row['entity']}: Cancel keeps visibility, Delete hides from non-admin",
        "entity": row["entity"],
        "ownerRole": owner,
        "ownerRoute": row["ownerRoute"],
        "ownerAccount": {"email": DEMO_USERS[owner], "password": PASSWORD},
        "impactedRoles": row["impactedRoles"],
        "impactedRoutes": row["impactedRoutes"],
        "steps": [
            f"Create or pick an active {row['entity']} record",
            "Run Cancel action (or equivalent status change)",
            "Verify record remains visible to intended users with cancelled status",
            "Run Delete action (if permissible)",
            "Verify record is removed from non-admin user views",
            "Login as Super Admin and verify deleted record is still visible in admin/audit view",
        ],
        "expectedResult": [
            "Cancel does not remove record from normal role visibility",
            "Delete removes record from non-admin UI",
            "Super Admin retains deleted-data visibility",
        ],
        "priority": "critical",
    }


def make_alert_impact_case(i, row):
    owner = row["ownerRole"]
    return {
        "testCaseId": f"C9-ALERT-IMPACT-{i:03d}",
        "category": "C-9",
        "type": "impact_alert_propagation",
        "title": f"{row['entity']} deletion/cancel informs impacted users",
        "entity": row["entity"],
        "ownerRole": owner,
        "ownerRoute": row["ownerRoute"],
        "ownerAccount": {"email": DEMO_USERS[owner], "password": PASSWORD},
        "impactedRoles": row["impactedRoles"],
        "impactedRoutes": row["impactedRoutes"],
        "steps": [
            f"Perform cancel/delete action for {row['entity']}",
            "Capture transaction reference and timestamp",
            "Login as each impacted role",
            "Check Alerts/Notifications and impacted module list",
            "Open alert and verify context (what changed, when, impact)",
        ],
        "expectedResult": [
            "Impacted users receive alert for cancel/delete impact",
            "Alert message is role-appropriate and context-rich",
            "No irrelevant users receive this impact alert",
        ],
        "priority": "high",
    }


def build_cases():
    cases = []
    idx = 1
    for row in ENTITIES:
        cases.append(make_crudv_case(idx, row))
        idx += 1
    idx = 1
    for row in ENTITIES:
        cases.append(make_cancel_vs_delete_case(idx, row))
        idx += 1
    idx = 1
    for row in ENTITIES:
        cases.append(make_alert_impact_case(idx, row))
        idx += 1

    cases.append(
        {
            "testCaseId": "C9-AUDIT-001",
            "category": "C-9",
            "type": "policy_coverage_matrix",
            "title": "Lifecycle policy matrix: Add/Edit/Delete/View/Cancel by entity",
            "entity": "all transactional entities",
            "ownerRole": "all_roles",
            "ownerRoute": "all transactional modules",
            "ownerAccount": {"email": "role-wise demo", "password": PASSWORD},
            "steps": [
                "Prepare entity-wise lifecycle matrix (Add/Edit/Delete/View/Cancel)",
                "Mark delete permission and constraints per entity",
                "Validate Super Admin deleted-data visibility path",
                "Validate impacted-user alert requirements",
                "Publish gaps where behavior diverges from policy",
            ],
            "expectedResult": [
                "Complete lifecycle governance coverage exists",
                "Delete vs Cancel semantics are consistent across entities",
                "Admin retention and user impact alerts are validated",
            ],
            "priority": "critical",
        }
    )
    return cases


def write_json(cases):
    payload = {
        "category": "C-9",
        "categoryTitle": "Transactional lifecycle and delete/cancel governance",
        "description": "Validate Add/Edit/Delete/View capabilities, delete permission boundaries, cancel vs delete semantics, Super Admin deleted-data visibility, and impacted-user alerts.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-9 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Title",
        "Entity",
        "Owner Role",
        "Owner Route",
        "Owner Email",
        "Owner Password",
        "Delete Rule",
        "Impacted Roles",
        "Impacted Routes",
        "Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["type"],
                case["title"],
                case["entity"],
                case["ownerRole"],
                case["ownerRoute"],
                case["ownerAccount"]["email"],
                case["ownerAccount"]["password"],
                case.get("deletePermissibleRule", ""),
                ", ".join(case.get("impactedRoles", [])),
                ", ".join(case.get("impactedRoutes", [])),
                "\n".join(case["steps"]),
                "\n".join(case["expectedResult"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 26,
        "D": 42,
        "E": 22,
        "F": 14,
        "G": 34,
        "H": 30,
        "I": 14,
        "J": 24,
        "K": 24,
        "L": 40,
        "M": 56,
        "N": 52,
        "O": 10,
        "P": 12,
        "Q": 28,
        "R": 28,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-9 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
