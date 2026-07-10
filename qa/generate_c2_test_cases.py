import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

PASSWORD = "Admin@123"

# Screen inventory for Category-2:
# pages that are tabular/list-heavy and should expose CSV export.
TABULAR_SCREENS = [
    {"role": "student", "route": "/dashboard/student/applications", "screen": "Student Applications"},
    {"role": "student", "route": "/dashboard/student/offers", "screen": "Student Offers"},
    {"role": "student", "route": "/dashboard/student/interviews", "screen": "Student Interviews"},
    {"role": "student", "route": "/dashboard/student/documents", "screen": "Student Documents"},
    {"role": "employer", "route": "/dashboard/employer/applications", "screen": "Employer Applications"},
    {"role": "employer", "route": "/dashboard/employer/offers", "screen": "Employer Offers"},
    {"role": "employer", "route": "/dashboard/employer/drives", "screen": "Employer Drives"},
    {"role": "employer", "route": "/dashboard/employer/interviews", "screen": "Employer Interviews"},
    {"role": "employer", "route": "/dashboard/employer/assessment-summary", "screen": "Employer Assessment Summary"},
    {"role": "employer", "route": "/dashboard/employer/assessment-uploads", "screen": "Employer Assessment Uploads"},
    {"role": "employer", "route": "/dashboard/employer/calendar", "screen": "Employer Calendar"},
    {"role": "college_admin", "route": "/dashboard/college/students", "screen": "College Students"},
    {"role": "college_admin", "route": "/dashboard/college/applications", "screen": "College Applications"},
    {"role": "college_admin", "route": "/dashboard/college/offers", "screen": "College Offers"},
    {"role": "college_admin", "route": "/dashboard/college/drives", "screen": "College Drives"},
    {"role": "college_admin", "route": "/dashboard/college/interviews", "screen": "College Interviews"},
    {"role": "college_admin", "route": "/dashboard/college/internship-results", "screen": "College Internship Results"},
    {"role": "college_admin", "route": "/dashboard/college/employers", "screen": "College Employers"},
    {"role": "college_admin", "route": "/dashboard/college/employers/requests", "screen": "Employer Partnership Requests"},
    {"role": "college_admin", "route": "/dashboard/college/calendar", "screen": "College Calendar"},
    {"role": "college_admin", "route": "/dashboard/college/reports", "screen": "College Reports"},
    {"role": "super_admin", "route": "/dashboard/admin/colleges", "screen": "Admin Colleges"},
    {"role": "super_admin", "route": "/dashboard/admin/users", "screen": "Admin Users"},
    {"role": "super_admin", "route": "/dashboard/admin/employers", "screen": "Admin Employers"},
    {"role": "super_admin", "route": "/dashboard/admin/pending-registrations", "screen": "Pending Registrations"},
    {"role": "super_admin", "route": "/dashboard/admin/feedback", "screen": "Admin Feedback Inbox"},
    {"role": "super_admin", "route": "/dashboard/admin/audit-reports", "screen": "Admin Audit Reports"},
]

OUTPUT_JSON = Path("qa/C-2-Tabular-Screens-CSV-Export.json")
OUTPUT_XLSX = Path("qa/C-2-Tabular-Screens-CSV-Export.xlsx")


def build_cases():
    cases = []
    for idx, entry in enumerate(TABULAR_SCREENS, start=1):
        role = entry["role"]
        route = entry["route"]
        screen = entry["screen"]
        email = DEMO_USERS[role]

        cases.append(
            {
                "testCaseId": f"C2-CSV-{idx:03d}",
                "category": "C-2",
                "title": f"{screen}: CSV export availability and sanity",
                "role": role,
                "route": route,
                "demoAccount": {"email": email, "password": PASSWORD},
                "preConditions": [
                    "Seed data loaded",
                    "User logged in with role-specific demo account",
                    "Screen contains at least one row of tabular/list data",
                ],
                "checks": [
                    {
                        "name": "Export option visible",
                        "expectation": "Export CSV button/link/menu action is visible and enabled on the screen.",
                    },
                    {
                        "name": "Export action works",
                        "expectation": "CSV file downloads successfully after triggering export.",
                    },
                    {
                        "name": "High-level data sanity",
                        "expectation": "CSV has header row, at least one data row (when UI has rows), and obvious key fields align with on-screen values for 2-3 sampled records.",
                    },
                ],
                "steps": [
                    f"Open {route}",
                    "Locate tabular/list section",
                    "Click Export CSV (or equivalent export action)",
                    "Open downloaded CSV",
                    "Compare 2-3 sampled rows with current UI values",
                ],
                "expectedResult": [
                    "Export option exists",
                    "CSV is downloaded without errors",
                    "CSV structure and sampled values are logically correct",
                ],
                "priority": "high",
            }
        )
    return cases


def write_json(cases):
    payload = {
        "category": "C-2",
        "categoryTitle": "Tabular screens CSV export checks",
        "description": "For each tabular data screen, validate CSV export presence, functionality, and high-level output sanity.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-2 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Role",
        "Screen",
        "Route",
        "Demo Email",
        "Demo Password",
        "Checks",
        "Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        checks_text = "\n".join([f"- {c['name']}: {c['expectation']}" for c in case["checks"]])
        expected_text = "\n".join(case["expectedResult"])
        steps_text = "\n".join(case["steps"])
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["role"],
                case["title"].split(":")[0],
                case["route"],
                case["demoAccount"]["email"],
                case["demoAccount"]["password"],
                checks_text,
                steps_text,
                expected_text,
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 14,
        "D": 28,
        "E": 44,
        "F": 32,
        "G": 14,
        "H": 60,
        "I": 48,
        "J": 46,
        "K": 10,
        "L": 12,
        "M": 30,
        "N": 30,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-2 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
