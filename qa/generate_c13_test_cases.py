import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_JSON = Path("qa/C-13-Visual-Style-Consistency-Against-Design-Tokens.json")
OUTPUT_XLSX = Path("qa/C-13-Visual-Style-Consistency-Against-Design-Tokens.xlsx")

# Source of truth: src/app/globals.css (:root tokens, .btn-*, .card, typography)
ESTABLISHED_STYLE = {
    "fonts": {
        "sans": "Inter (--font-sans)",
        "display": "Fraunces (--font-display) for headings/display only",
    },
    "semanticColors": "text-primary/secondary/tertiary, --primary-*, --success-*, --warning-*, --danger-*, --info-*, --gray-*",
    "surfaces": "--bg-primary/secondary/tertiary/elevated, --border-default/strong",
    "spacing": "--space-xs through --space-3xl",
    "radius": "--radius-sm through --radius-full",
    "shadows": "--shadow-sm/md/lg/xl",
    "components": "Prefer .btn, .btn-primary, .btn-secondary, .card, .form-input, .form-select per globals.css",
}

DEVIATION_SIGNALS = [
    "Hard-coded hex/rgb in inline style or arbitrary class (e.g. #333, rgb()) instead of CSS variables",
    "Font family or font-size set inline instead of body/heading scale in globals",
    "Button that is not .btn / .btn-* variant but custom-styled <button>",
    "Border radius or shadow that does not match --radius-* / --shadow-* scale",
    "Text color as raw gray hex instead of var(--text-*) or token",
    "Inconsistent heading level styling (h1/h2) vs .page-header patterns",
    "Mix of emoji-heavy headers with pages that use plain text (document pattern per screen type)",
]

SCREEN_SAMPLES = [
    {"route": "/dashboard/layout (shell)", "role": "all", "focus": "Sidebar, top bar, nav density, focus rings"},
    {"route": "/dashboard/student/overview", "role": "student", "focus": "Cards, stats grid, typography"},
    {"route": "/dashboard/student/profile", "role": "student", "focus": "Forms, labels, file inputs"},
    {"route": "/dashboard/college/overview", "role": "college_admin", "focus": "KPI cards, charts placeholders"},
    {"route": "/dashboard/college/offers", "role": "college_admin", "focus": "Table, modals, action buttons"},
    {"route": "/dashboard/college/drives", "role": "college_admin", "focus": "Status chips, report CTA"},
    {"route": "/dashboard/employer/overview", "role": "employer", "focus": "Dashboard cards"},
    {"route": "/dashboard/employer/offers", "role": "employer", "focus": "Dense table, filters"},
    {"route": "/dashboard/employer/select-campus", "role": "employer", "focus": "Cards, inline styles hot spot"},
    {"route": "/dashboard/admin/overview", "role": "super_admin", "focus": "Admin chrome consistency"},
    {"route": "/dashboard/alerts", "role": "multi", "focus": "List rows, empty states"},
    {"route": "/dashboard/help", "role": "multi", "focus": "Prose, accordions, spacing"},
    {"route": "/login", "role": "public", "focus": "Auth layout vs dashboard tokens"},
    {"route": "/register", "role": "public", "focus": "Auth layout vs dashboard tokens"},
    {"route": "/", "role": "public", "focus": "Landing vs app shell consistency"},
]


def build_cases():
    cases = []
    for i, row in enumerate(SCREEN_SAMPLES, start=1):
        cases.append(
            {
                "testCaseId": f"C13-VIS-{i:03d}",
                "category": "C-13",
                "type": "visual_token_audit",
                "title": f"Visual consistency audit: {row['route']}",
                "route": row["route"],
                "role": row["role"],
                "uiFocus": row["focus"],
                "establishedStyleReference": ESTABLISHED_STYLE,
                "deviationSignalsToFlag": DEVIATION_SIGNALS,
                "steps": [
                    "Open route in light and dark theme (if supported)",
                    "Compare typography: body uses Inter; display headings use Fraunces where intended",
                    "Verify text colors use semantic tokens (no one-off grays)",
                    "Verify primary CTAs use .btn-primary (or equivalent token-backed variant)",
                    "Verify secondary actions use .btn-secondary",
                    "Verify cards use .card or token-backed surfaces",
                    "Verify spacing follows 4/8px rhythm (--space-*)",
                    "Verify border radius uses token scale",
                    "Inspect DevTools: flag inline styles that duplicate token values",
                ],
                "expectedResult": [
                    "No blocking visual deviation from globals.css token system",
                    "Deviations are documented with screenshot + selector + suggested token",
                ],
                "priority": "P1" if i <= 5 else "P2",
            }
        )

    cases.append(
        {
            "testCaseId": "C13-AUTO-001",
            "category": "C-13",
            "type": "automated_heuristic",
            "title": "Repo scan: inline style and hard-coded color hotspots",
            "route": "src/app (dashboard + public)",
            "role": "qa_engineering",
            "uiFocus": "Grep / lint for style={{ and #[0-9a-fA-F]{3,6}",
            "establishedStyleReference": ESTABLISHED_STYLE,
            "deviationSignalsToFlag": DEVIATION_SIGNALS,
            "steps": [
                "Search codebase for inline style objects on dashboard pages",
                "Search for hex colors in JSX outside theme definition files",
                "Triage: acceptable (charts/third-party) vs should migrate to token",
            ],
            "expectedResult": [
                "Inventory of files with highest deviation risk",
                "Trend line improves sprint-over-sprint",
            ],
            "priority": "P1",
        }
    )

    cases.append(
        {
            "testCaseId": "C13-AUDIT-001",
            "category": "C-13",
            "type": "full_route_matrix",
            "title": "Complete dashboard route visual consistency sign-off",
            "route": "all /dashboard/** routes from menuConfig",
            "role": "all",
            "uiFocus": "100% route coverage checklist",
            "establishedStyleReference": ESTABLISHED_STYLE,
            "deviationSignalsToFlag": DEVIATION_SIGNALS,
            "steps": [
                "Enumerate every route from dashboardMenu.js + screenRegistry",
                "Mark each Pass / Minor deviation / Major deviation",
                "File bugs for Major; backlog Minor to token migration",
            ],
            "expectedResult": [
                "Signed matrix attached to release",
                "Zero undocumented Major deviations at release gate",
            ],
            "priority": "P0",
        }
    )
    return cases


def write_json(cases):
    payload = {
        "category": "C-13",
        "categoryTitle": "Visual style consistency vs established app tokens",
        "description": "Detect deviation in font, color, size, spacing, radius, and components from globals.css design system.",
        "sourceOfTruthFile": "src/app/globals.css",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-13 Visual audit"
    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Priority",
        "Route",
        "Role",
        "UI focus",
        "Steps (summary)",
        "Deviation signals",
        "Status",
        "Finding",
        "Suggested fix",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    for c in cases:
        steps = c.get("steps", [])
        ws.append(
            [
                c["testCaseId"],
                c["category"],
                c["type"],
                c.get("priority", ""),
                c["route"],
                c["role"],
                c.get("uiFocus", ""),
                steps[0] if steps else "",
                "; ".join(c.get("deviationSignalsToFlag", [])[:3]) + "…",
                "Not Run",
                "",
                "",
            ]
        )
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    for col, w in zip("ABCDEFGHIJKL", [14, 8, 22, 8, 36, 12, 28, 40, 36, 12, 24, 24]):
        ws.column_dimensions[col].width = w
    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-13 test cases")
    print(OUTPUT_JSON)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
