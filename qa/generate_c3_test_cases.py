import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}
PASSWORD = "Admin@123"

OUTPUT_JSON = Path("qa/C-3-Transactional-Flow-and-Cross-Role-Visibility.json")
OUTPUT_XLSX = Path("qa/C-3-Transactional-Flow-and-Cross-Role-Visibility.xlsx")


FLOWS = [
    {
        "id": "DRIVE_CREATE_APPROVE_PUBLISH",
        "title": "Create placement drive -> approve -> visible to student",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/drives",
        "transactionAction": "Create new placement drive",
        "downstreamChecks": [
            {
                "role": "college_admin",
                "route": "/dashboard/college/drives",
                "expectation": "New drive appears in requested/pending list and can be reviewed.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/drives",
                "expectation": "After approval, drive status is approved/scheduled.",
            },
            {
                "role": "student",
                "route": "/dashboard/student/drives",
                "expectation": "Approved drive is visible for eligible students.",
            },
        ],
    },
    {
        "id": "JOB_POSTING_TO_STUDENT_VISIBILITY",
        "title": "Create job posting -> visible in student opportunities",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/jobs",
        "transactionAction": "Create job posting with eligibility criteria",
        "downstreamChecks": [
            {
                "role": "college_admin",
                "route": "/dashboard/college/drives",
                "expectation": "Posting/linked drive can be tracked by placement office.",
            },
            {
                "role": "student",
                "route": "/dashboard/student/drives",
                "expectation": "Eligible student can discover posting/drive.",
            },
        ],
    },
    {
        "id": "STUDENT_APPLY_TO_EMPLOYER_PIPELINE",
        "title": "Student applies -> application visible to employer and college",
        "transactionOwnerRole": "student",
        "transactionRoute": "/dashboard/student/applications",
        "transactionAction": "Apply for an internship/drive",
        "downstreamChecks": [
            {
                "role": "employer",
                "route": "/dashboard/employer/applications",
                "expectation": "New application appears in employer candidate pipeline.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/applications",
                "expectation": "Application appears in college applications list with same status.",
            },
        ],
    },
    {
        "id": "EMPLOYER_SHORTLIST_PROPAGATION",
        "title": "Employer shortlist decision reflects to student and college",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/applications",
        "transactionAction": "Mark candidate as shortlisted",
        "downstreamChecks": [
            {
                "role": "student",
                "route": "/dashboard/student/applications",
                "expectation": "Student sees updated status as shortlisted.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/applications",
                "expectation": "College sees shortlisted status for candidate.",
            },
        ],
    },
    {
        "id": "INTERVIEW_SCHEDULE_VISIBILITY",
        "title": "Interview schedule visible to student, employer, and college",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/interviews",
        "transactionAction": "Create or update interview slot",
        "downstreamChecks": [
            {
                "role": "student",
                "route": "/dashboard/student/interviews",
                "expectation": "Scheduled interview appears with date/time details.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/interviews",
                "expectation": "Interview appears in college scheduling board.",
            },
        ],
    },
    {
        "id": "OFFER_RELEASE_VISIBILITY",
        "title": "Employer releases offer -> visible to student and college",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/offers",
        "transactionAction": "Issue offer to selected student",
        "downstreamChecks": [
            {
                "role": "student",
                "route": "/dashboard/student/offers",
                "expectation": "Offer appears with salary/location/deadline details.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/offers",
                "expectation": "Offer appears in college offer tracking.",
            },
        ],
    },
    {
        "id": "STUDENT_OFFER_RESPONSE_PROPAGATION",
        "title": "Student accepts/rejects offer -> reflected back to employer and college",
        "transactionOwnerRole": "student",
        "transactionRoute": "/dashboard/student/offers",
        "transactionAction": "Accept or reject offer",
        "downstreamChecks": [
            {
                "role": "employer",
                "route": "/dashboard/employer/offers",
                "expectation": "Offer response status updates to accepted/rejected.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/offers",
                "expectation": "College sees final offer response state.",
            },
        ],
    },
    {
        "id": "ASSESSMENT_UPLOAD_PROPAGATION",
        "title": "Assessment upload and result mapping visible across stakeholders",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/assessment-uploads",
        "transactionAction": "Upload assessment CSV/results",
        "downstreamChecks": [
            {
                "role": "employer",
                "route": "/dashboard/employer/assessment-summary",
                "expectation": "Uploaded records appear in assessment summary.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/hiring-assessment",
                "expectation": "College can view uploaded assessment outcome mapping.",
            },
        ],
    },
    {
        "id": "COLLEGE_OFFERS_UPLOAD_PROPAGATION",
        "title": "College bulk offers upload visible to student and employer views",
        "transactionOwnerRole": "college_admin",
        "transactionRoute": "/dashboard/college/offers-upload",
        "transactionAction": "Upload offers CSV from placement office",
        "downstreamChecks": [
            {
                "role": "college_admin",
                "route": "/dashboard/college/offers",
                "expectation": "Uploaded offers appear in college offers list.",
            },
            {
                "role": "student",
                "route": "/dashboard/student/offers",
                "expectation": "Concerned student sees uploaded offer entry.",
            },
        ],
    },
    {
        "id": "EMPLOYER_PARTNERSHIP_REQUEST_FLOW",
        "title": "Employer partnership request -> visible to college review queue",
        "transactionOwnerRole": "employer",
        "transactionRoute": "/dashboard/employer/select-campus",
        "transactionAction": "Send partnership/tie-up request to a college",
        "downstreamChecks": [
            {
                "role": "college_admin",
                "route": "/dashboard/college/employers/requests",
                "expectation": "New request appears in requests inbox.",
            },
            {
                "role": "college_admin",
                "route": "/dashboard/college/employers",
                "expectation": "After approval, employer appears in approved partners.",
            },
        ],
    },
    {
        "id": "CALENDAR_EVENT_FLOW",
        "title": "College placement event creation visible on student calendar",
        "transactionOwnerRole": "college_admin",
        "transactionRoute": "/dashboard/college/calendar",
        "transactionAction": "Create placement event/interview block",
        "downstreamChecks": [
            {
                "role": "student",
                "route": "/dashboard/student/calendar",
                "expectation": "Event appears on student placement calendar.",
            },
            {
                "role": "employer",
                "route": "/dashboard/employer/calendar",
                "expectation": "Relevant event visibility aligns with linked hiring activity.",
            },
        ],
    },
    {
        "id": "DISCUSSION_CLARIFICATION_FLOW",
        "title": "Clarification/discussion created by college visible to student and employer",
        "transactionOwnerRole": "college_admin",
        "transactionRoute": "/dashboard/college/clarifications",
        "transactionAction": "Publish clarification entry for a company/drive",
        "downstreamChecks": [
            {
                "role": "student",
                "route": "/dashboard/student/clarifications",
                "expectation": "Clarification appears in student communication module.",
            },
            {
                "role": "employer",
                "route": "/dashboard/employer/discussions",
                "expectation": "Associated communication thread is visible to employer.",
            },
        ],
    },
]


def expand_cases():
    cases = []
    serial = 1
    for flow in FLOWS:
        owner_role = flow["transactionOwnerRole"]
        owner_email = DEMO_USERS[owner_role]
        case = {
            "testCaseId": f"C3-TXN-{serial:03d}",
            "category": "C-3",
            "flowId": flow["id"],
            "title": flow["title"],
            "transactionOwnerRole": owner_role,
            "transactionOwnerAccount": {"email": owner_email, "password": PASSWORD},
            "transactionRoute": flow["transactionRoute"],
            "transactionAction": flow["transactionAction"],
            "steps": [
                f"Login as {owner_role} ({owner_email})",
                f"Open {flow['transactionRoute']}",
                f"Execute transaction: {flow['transactionAction']}",
                "Capture transaction reference (name/id/status/timestamp)",
                "Login as each downstream role and validate visibility",
            ],
            "downstreamChecks": flow["downstreamChecks"],
            "expectedResult": [
                "Transaction executes successfully without errors",
                "Transaction state persists after page refresh",
                "Transaction is visible in all intended cross-role locations",
                "Key fields (status, title/name, candidate/company, dates) stay consistent across roles",
            ],
            "priority": "critical",
        }
        cases.append(case)
        serial += 1
    return cases


def write_json(cases):
    payload = {
        "category": "C-3",
        "categoryTitle": "Cross-role transactional execution and visibility",
        "description": "Validate that business transactions can be executed by owning role and become visible to all intended roles/screens.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-3 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Flow ID",
        "Title",
        "Owner Role",
        "Owner Email",
        "Owner Password",
        "Transaction Route",
        "Transaction Action",
        "Cross-Role Validation Points",
        "Execution Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        cross_role_text = "\n".join(
            [
                f"- [{c['role']}] {c['route']} -> {c['expectation']}"
                for c in case["downstreamChecks"]
            ]
        )
        steps_text = "\n".join(case["steps"])
        expected_text = "\n".join(case["expectedResult"])
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["flowId"],
                case["title"],
                case["transactionOwnerRole"],
                case["transactionOwnerAccount"]["email"],
                case["transactionOwnerAccount"]["password"],
                case["transactionRoute"],
                case["transactionAction"],
                cross_role_text,
                steps_text,
                expected_text,
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 32,
        "D": 46,
        "E": 14,
        "F": 30,
        "G": 14,
        "H": 38,
        "I": 36,
        "J": 70,
        "K": 52,
        "L": 58,
        "M": 10,
        "N": 12,
        "O": 32,
        "P": 32,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = expand_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-3 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
