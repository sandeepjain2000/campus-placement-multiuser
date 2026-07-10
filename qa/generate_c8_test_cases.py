import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PASSWORD = "Admin@123"
DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

OUTPUT_JSON = Path("qa/C-8-Alerts-Notifications-Trigger-Visibility-and-Control.json")
OUTPUT_XLSX = Path("qa/C-8-Alerts-Notifications-Trigger-Visibility-and-Control.xlsx")


ALERT_TRIGGER_CASES = [
    {
        "id": "C8-ALERT-001",
        "transaction": "Employer creates drive request",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/drives",
        "intendedViewerRole": "college_admin",
        "intendedViewerRoute": "/dashboard/alerts",
        "expectedAlertType": "Drive request",
    },
    {
        "id": "C8-ALERT-002",
        "transaction": "College admin approves drive",
        "actorRole": "college_admin",
        "actorRoute": "/dashboard/college/drives",
        "intendedViewerRole": "employer",
        "intendedViewerRoute": "/dashboard/alerts",
        "expectedAlertType": "Drive approval",
    },
    {
        "id": "C8-ALERT-003",
        "transaction": "Student applies to internship/drive",
        "actorRole": "student",
        "actorRoute": "/dashboard/student/drives",
        "intendedViewerRole": "employer",
        "intendedViewerRoute": "/dashboard/alerts",
        "expectedAlertType": "New application",
    },
    {
        "id": "C8-ALERT-004",
        "transaction": "Employer shortlists/rejects candidate",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/applications",
        "intendedViewerRole": "student",
        "intendedViewerRoute": "/dashboard/alerts",
        "expectedAlertType": "Application status changed",
    },
    {
        "id": "C8-ALERT-005",
        "transaction": "Interview gets scheduled/updated",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/interviews",
        "intendedViewerRole": "student",
        "intendedViewerRoute": "/dashboard/student/interviews",
        "expectedAlertType": "Interview schedule",
    },
    {
        "id": "C8-ALERT-006",
        "transaction": "Offer released to student",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/offers",
        "intendedViewerRole": "student",
        "intendedViewerRoute": "/dashboard/alerts",
        "expectedAlertType": "Offer released",
    },
    {
        "id": "C8-ALERT-007",
        "transaction": "Student accepts/rejects offer",
        "actorRole": "student",
        "actorRoute": "/dashboard/student/offers",
        "intendedViewerRole": "employer",
        "intendedViewerRoute": "/dashboard/alerts",
        "expectedAlertType": "Offer response",
    },
    {
        "id": "C8-ALERT-008",
        "transaction": "Employer sends partnership request",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/select-campus",
        "intendedViewerRole": "college_admin",
        "intendedViewerRoute": "/dashboard/college/employers/requests",
        "expectedAlertType": "Partnership request",
    },
    {
        "id": "C8-ALERT-009",
        "transaction": "College publishes clarification",
        "actorRole": "college_admin",
        "actorRoute": "/dashboard/college/clarifications",
        "intendedViewerRole": "student",
        "intendedViewerRoute": "/dashboard/student/clarifications",
        "expectedAlertType": "Clarification posted",
    },
]


PREFERENCE_CASES = [
    {
        "id": "C8-PREF-001",
        "title": "User can view alert/notification preference settings",
        "role": "student",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C8-PREF-002",
        "title": "User can disable one alert category",
        "role": "employer",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C8-PREF-003",
        "title": "Disabled alert category suppresses future alert creation/visibility",
        "role": "college_admin",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C8-PREF-004",
        "title": "User can re-enable alert category and receive alerts again",
        "role": "college_admin",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C8-PREF-005",
        "title": "Alert preferences are isolated per user",
        "role": "super_admin",
        "route": "/dashboard/admin/settings",
    },
]


def build_trigger_cases():
    cases = []
    for row in ALERT_TRIGGER_CASES:
        actor_role = row["actorRole"]
        viewer_role = row["intendedViewerRole"]
        cases.append(
            {
                "testCaseId": row["id"],
                "category": "C-8",
                "type": "alert_trigger_and_visibility",
                "title": f"{row['transaction']} creates intended alert",
                "actorRole": actor_role,
                "actorRoute": row["actorRoute"],
                "actorAccount": {"email": DEMO_USERS[actor_role], "password": PASSWORD},
                "intendedViewerRole": viewer_role,
                "intendedViewerRoute": row["intendedViewerRoute"],
                "intendedViewerAccount": {"email": DEMO_USERS[viewer_role], "password": PASSWORD},
                "expectedAlertType": row["expectedAlertType"],
                "steps": [
                    f"Login as {actor_role} and execute transaction at {row['actorRoute']}",
                    "Capture transaction reference (id/title/status/time)",
                    f"Login as {viewer_role} and open {row['intendedViewerRoute']}",
                    "Verify new alert/notification entry is visible",
                    "Open alert target and verify deep-link/context is correct",
                ],
                "expectedResult": [
                    "Alert entry is created for intended viewer(s)",
                    "Alert content matches transaction context",
                    "Unauthorized roles do not see this alert",
                ],
                "priority": "critical",
            }
        )
    return cases


def build_preference_cases():
    cases = []
    for row in PREFERENCE_CASES:
        role = row["role"]
        cases.append(
            {
                "testCaseId": row["id"],
                "category": "C-8",
                "type": "alert_preference_control",
                "title": row["title"],
                "actorRole": role,
                "actorRoute": row["route"],
                "actorAccount": {"email": DEMO_USERS[role], "password": PASSWORD},
                "steps": [
                    f"Login as {role} and open {row['route']}",
                    "Locate alert/notification controls",
                    "Toggle category ON/OFF and save",
                    "Trigger mapped transaction",
                    "Verify alert visibility follows configured preference",
                ],
                "expectedResult": [
                    "Preference control is available and persists after refresh/login",
                    "Suppressed categories do not show new alerts",
                    "Enabled categories show new alerts",
                ],
                "priority": "high",
            }
        )
    return cases


def build_meta_case():
    return [
        {
            "testCaseId": "C8-AUDIT-001",
            "category": "C-8",
            "type": "coverage_matrix",
            "title": "All transaction/change events mapped to alert rules",
            "actorRole": "all_roles",
            "actorRoute": "all transactional/change routes",
            "actorAccount": {"email": "role-wise demo", "password": PASSWORD},
            "steps": [
                "Build inventory of all transaction + status-change events",
                "Map each event to expected alert recipients",
                "Execute sample events and verify alert creation",
                "Publish gap register for missing/incorrect alert behavior",
            ],
            "expectedResult": [
                "Complete mapping exists for alert coverage",
                "Any missing alert triggers or wrong recipient mapping is logged",
            ],
            "priority": "critical",
        }
    ]


def build_cases():
    return build_trigger_cases() + build_preference_cases() + build_meta_case()


def write_json(cases):
    payload = {
        "category": "C-8",
        "categoryTitle": "Alert/notification trigger, visibility, and preference control",
        "description": "Validate alert creation for each transaction/change event, intended cross-role visibility, and user-level control over alert categories.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-8 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Title",
        "Actor Role",
        "Actor Route",
        "Actor Email",
        "Actor Password",
        "Intended Viewer Role",
        "Intended Viewer Route",
        "Intended Viewer Email",
        "Expected Alert Type",
        "Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        intended_viewer = case.get("intendedViewerAccount", {})
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["type"],
                case["title"],
                case["actorRole"],
                case["actorRoute"],
                case["actorAccount"]["email"],
                case["actorAccount"]["password"],
                case.get("intendedViewerRole", ""),
                case.get("intendedViewerRoute", ""),
                intended_viewer.get("email", ""),
                case.get("expectedAlertType", ""),
                "\n".join(case["steps"]),
                "\n".join(case["expectedResult"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 26,
        "D": 40,
        "E": 14,
        "F": 36,
        "G": 30,
        "H": 14,
        "I": 20,
        "J": 34,
        "K": 30,
        "L": 24,
        "M": 52,
        "N": 48,
        "O": 10,
        "P": 12,
        "Q": 28,
        "R": 28,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-8 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
