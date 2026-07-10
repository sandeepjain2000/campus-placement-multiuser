import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font


PASSWORD = "Admin@123"
DEMO_USERS = {
    "student": "arjun.verma@iitm.edu",
    "employer": "hr@techcorp.com",
    "college_admin": "admin@iitm.edu",
    "super_admin": "admin@placementhub.com",
}

OUTPUT_JSON = Path("qa/C-7-Email-Notification-and-Preference-Control.json")
OUTPUT_XLSX = Path("qa/C-7-Email-Notification-and-Preference-Control.xlsx")


TRANSACTION_EMAIL_CASES = [
    {
        "id": "C7-EMAIL-001",
        "transaction": "Employer creates placement drive request",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/drives",
        "expectedRecipientRole": "college_admin",
        "expectedEmailType": "Drive request notification",
    },
    {
        "id": "C7-EMAIL-002",
        "transaction": "College admin approves drive",
        "actorRole": "college_admin",
        "actorRoute": "/dashboard/college/drives",
        "expectedRecipientRole": "employer",
        "expectedEmailType": "Drive approval update",
    },
    {
        "id": "C7-EMAIL-003",
        "transaction": "Student applies to internship/drive",
        "actorRole": "student",
        "actorRoute": "/dashboard/student/drives",
        "expectedRecipientRole": "employer",
        "expectedEmailType": "New application alert",
    },
    {
        "id": "C7-EMAIL-004",
        "transaction": "Employer shortlists candidate",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/applications",
        "expectedRecipientRole": "student",
        "expectedEmailType": "Shortlist status update",
    },
    {
        "id": "C7-EMAIL-005",
        "transaction": "Employer schedules interview",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/interviews",
        "expectedRecipientRole": "student",
        "expectedEmailType": "Interview schedule notification",
    },
    {
        "id": "C7-EMAIL-006",
        "transaction": "Employer releases offer",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/offers",
        "expectedRecipientRole": "student",
        "expectedEmailType": "Offer release email",
    },
    {
        "id": "C7-EMAIL-007",
        "transaction": "Student accepts/rejects offer",
        "actorRole": "student",
        "actorRoute": "/dashboard/student/offers",
        "expectedRecipientRole": "employer",
        "expectedEmailType": "Offer response update",
    },
    {
        "id": "C7-EMAIL-008",
        "transaction": "Employer sends campus partnership request",
        "actorRole": "employer",
        "actorRoute": "/dashboard/employer/select-campus",
        "expectedRecipientRole": "college_admin",
        "expectedEmailType": "Partnership request notification",
    },
    {
        "id": "C7-EMAIL-009",
        "transaction": "College publishes clarification",
        "actorRole": "college_admin",
        "actorRoute": "/dashboard/college/clarifications",
        "expectedRecipientRole": "student",
        "expectedEmailType": "Clarification broadcast",
    },
]


PREFERENCE_CASES = [
    {
        "id": "C7-PREF-001",
        "title": "User can view email notification preferences",
        "role": "student",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C7-PREF-002",
        "title": "User can disable a notification category",
        "role": "employer",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C7-PREF-003",
        "title": "Disabled category suppresses future email",
        "role": "college_admin",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C7-PREF-004",
        "title": "User can re-enable email category",
        "role": "college_admin",
        "route": "/dashboard/alerts",
    },
    {
        "id": "C7-PREF-005",
        "title": "Preference isolation across users/roles",
        "role": "super_admin",
        "route": "/dashboard/admin/settings",
    },
]


def build_transaction_cases():
    cases = []
    for row in TRANSACTION_EMAIL_CASES:
        actor = row["actorRole"]
        actor_email = DEMO_USERS[actor]
        recipient_role = row["expectedRecipientRole"]
        recipient_email = DEMO_USERS.get(recipient_role, "configured recipient email")

        cases.append(
            {
                "testCaseId": row["id"],
                "category": "C-7",
                "type": "transaction_email",
                "title": f"{row['transaction']} triggers email",
                "actorRole": actor,
                "actorRoute": row["actorRoute"],
                "actorAccount": {"email": actor_email, "password": PASSWORD},
                "expectedRecipientRole": recipient_role,
                "expectedRecipientEmail": recipient_email,
                "expectedEmailType": row["expectedEmailType"],
                "mailVerificationMethod": [
                    "Use working mailbox credentials from JSON config.",
                    "Capture baseline unread count / latest timestamp.",
                    "Execute transaction in app.",
                    "Check inbox for new email within SLA (e.g., 0-5 min).",
                    "Validate subject/body include transaction reference (name/id/status).",
                ],
                "expectedResult": [
                    "Email is delivered to intended recipient mailbox.",
                    "Content corresponds to executed transaction.",
                    "No duplicate or wrong-recipient email is sent.",
                ],
                "priority": "critical",
            }
        )
    return cases


def build_preference_cases():
    cases = []
    for row in PREFERENCE_CASES:
        role = row["role"]
        cases.append(
            {
                "testCaseId": row["id"],
                "category": "C-7",
                "type": "email_preference_control",
                "title": row["title"],
                "actorRole": role,
                "actorRoute": row["route"],
                "actorAccount": {"email": DEMO_USERS[role], "password": PASSWORD},
                "mailVerificationMethod": [
                    "Open notification/email preference controls.",
                    "Toggle target category ON/OFF and save.",
                    "Trigger mapped transaction.",
                    "Verify email delivered/suppressed as configured.",
                ],
                "expectedResult": [
                    "Preference control exists and persists.",
                    "Email behavior follows selected preference.",
                ],
                "priority": "high",
            }
        )
    return cases


def build_meta_cases():
    return [
        {
            "testCaseId": "C7-AUDIT-001",
            "category": "C-7",
            "type": "coverage_matrix",
            "title": "Email trigger coverage for all transactions and state changes",
            "actorRole": "all_roles",
            "actorRoute": "all transactional routes",
            "actorAccount": {"email": "role-wise demo", "password": PASSWORD},
            "mailVerificationMethod": [
                "Create transaction inventory from business flows.",
                "Mark each flow as Email Expected / Not Expected by design.",
                "Run trigger test for each Email Expected flow.",
                "Publish gap list for missing or incorrect emails.",
            ],
            "expectedResult": [
                "Complete email trigger coverage exists.",
                "Any missing email trigger is explicitly logged as defect/gap.",
            ],
            "priority": "critical",
        }
    ]


def build_cases():
    return build_transaction_cases() + build_preference_cases() + build_meta_cases()


def write_json(cases):
    payload = {
        "category": "C-7",
        "categoryTitle": "Email notifications and user preference control",
        "description": "Validate transaction-triggered emails, mailbox delivery verification, and user-level control over email categories.",
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-7 Manual Cases"

    headers = [
        "Test Case ID",
        "Category",
        "Type",
        "Title",
        "Actor Role",
        "Actor Route",
        "Actor Email",
        "Actor Password",
        "Expected Recipient Role",
        "Expected Recipient Email",
        "Expected Email Type",
        "Mail Verification Steps",
        "Expected Result",
        "Priority",
        "Status",
        "Actual Result",
        "Comments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    for case in cases:
        ws.append(
            [
                case["testCaseId"],
                case["category"],
                case["type"],
                case["title"],
                case["actorRole"],
                case["actorRoute"],
                case["actorAccount"]["email"],
                case["actorAccount"]["password"],
                case.get("expectedRecipientRole", ""),
                case.get("expectedRecipientEmail", ""),
                case.get("expectedEmailType", ""),
                "\n".join(case["mailVerificationMethod"]),
                "\n".join(case["expectedResult"]),
                case["priority"],
                "Not Run",
                "",
                "",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")

    width_map = {
        "A": 16,
        "B": 10,
        "C": 24,
        "D": 40,
        "E": 14,
        "F": 36,
        "G": 30,
        "H": 14,
        "I": 20,
        "J": 32,
        "K": 28,
        "L": 52,
        "M": 48,
        "N": 10,
        "O": 12,
        "P": 28,
        "Q": 28,
    }
    for col, width in width_map.items():
        ws.column_dimensions[col].width = width

    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-7 test cases")
    print(f"- {OUTPUT_JSON}")
    print(f"- {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
