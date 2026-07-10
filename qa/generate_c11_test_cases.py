import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUTPUT_JSON = Path("qa/C-11-Design-Refresh-Shadcn-V0-Candidates.json")
OUTPUT_XLSX = Path("qa/C-11-Design-Refresh-Shadcn-V0-Candidates.xlsx")

# Grounded in repo: dashboard uses custom globals (btn/card/form-*) not @/components/ui.
# Priority: shell + highest-traffic / densest screens benefit most from shadcn/v0 polish.
SCREENS = [
    {
        "route": "/dashboard/layout (shell)",
        "roles": ["student", "employer", "college_admin", "super_admin"],
        "priority": "P0",
        "designDebtSignals": "Top bar uses inline layout styles; session/year select is raw select; sidebar density varies by role.",
        "shadcnV0Targets": "Sidebar (ScrollArea), DropdownMenu (user), Select (academic year), Separator, Sheet (mobile nav), consistent typography scale.",
    },
    {
        "route": "/dashboard/college/offers",
        "roles": ["college_admin"],
        "priority": "P0",
        "designDebtSignals": "Heavy table + modal-style flows; many btn-secondary clusters; complex form state.",
        "shadcnV0Targets": "DataTable + Column visibility, Dialog/Sheet for edit, Form + Field, Badge for status, Command for student lookup.",
    },
    {
        "route": "/dashboard/employer/offers",
        "roles": ["employer"],
        "priority": "P0",
        "designDebtSignals": "Dense offers grid; multiple action buttons per row; filter UX.",
        "shadcnV0Targets": "Table toolbar (filters), DropdownMenu row actions, Dialog forms, Tabs (by status).",
    },
    {
        "route": "/dashboard/college/students",
        "roles": ["college_admin"],
        "priority": "P0",
        "designDebtSignals": "Directory + bulk actions + CSV; typical 'admin grid' that benefits from table primitives.",
        "shadcnV0Targets": "DataTable, Pagination, Input debounced search, AlertDialog destructive confirm.",
    },
    {
        "route": "/dashboard/employer/applications",
        "roles": ["employer"],
        "priority": "P0",
        "designDebtSignals": "Pipeline table + status changes; needs clear hierarchy and row affordances.",
        "shadcnV0Targets": "DataTable, Select status, HoverCard candidate summary, Skeleton loading.",
    },
    {
        "route": "/dashboard/college/applications",
        "roles": ["college_admin"],
        "priority": "P0",
        "designDebtSignals": "Cross-tenant visibility table; filters and density.",
        "shadcnV0Targets": "DataTable, faceted filters (Popover + Command), Badge.",
    },
    {
        "route": "/dashboard/college/drives",
        "roles": ["college_admin"],
        "priority": "P1",
        "designDebtSignals": "Drive cards/table mix; report download pattern; status chips.",
        "shadcnV0Targets": "Card grid with consistent spacing, AlertDialog approve/reject, Toast patterns unified.",
    },
    {
        "route": "/dashboard/employer/drives",
        "roles": ["employer"],
        "priority": "P1",
        "designDebtSignals": "Create/request flows + lists; form density.",
        "shadcnV0Targets": "Stepper or multi-section Form, CalendarDatePicker for drive date.",
    },
    {
        "route": "/dashboard/employer/interviews",
        "roles": ["employer"],
        "priority": "P1",
        "designDebtSignals": "Scheduling UI often needs calendar + slot list clarity.",
        "shadcnV0Targets": "Tabs (list vs calendar), Sheet for slot detail, Time pickers.",
    },
    {
        "route": "/dashboard/college/interviews",
        "roles": ["college_admin"],
        "priority": "P1",
        "designDebtSignals": "Mirrors employer scheduling complexity for campus ops.",
        "shadcnV0Targets": "Same as employer interviews + conflict warnings (Alert).",
    },
    {
        "route": "/dashboard/student/applications",
        "roles": ["student"],
        "priority": "P1",
        "designDebtSignals": "Status timeline readability; mobile-friendly cards.",
        "shadcnV0Targets": "Accordion or Timeline component pattern, Card list, Progress.",
    },
    {
        "route": "/dashboard/student/offers",
        "roles": ["student"],
        "priority": "P1",
        "designDebtSignals": "High-stakes UI; needs strong visual hierarchy and CTA clarity.",
        "shadcnV0Targets": "Card + Alert for deadlines, Dialog confirm accept/reject.",
    },
    {
        "route": "/dashboard/student/profile",
        "roles": ["student"],
        "priority": "P1",
        "designDebtSignals": "Long form; file inputs; section grouping.",
        "shadcnV0Targets": "Tabs (sections), Form, Field, File upload dropzone styling.",
    },
    {
        "route": "/dashboard/employer/profile",
        "roles": ["employer"],
        "priority": "P1",
        "designDebtSignals": "Company profile + logo upload flows.",
        "shadcnV0Targets": "Card sections, Avatar, Form layout grid.",
    },
    {
        "route": "/dashboard/college/settings",
        "roles": ["college_admin"],
        "priority": "P1",
        "designDebtSignals": "Settings forms + uploads; validation messaging.",
        "shadcnV0Targets": "Tabs, Switch toggles, Form + description helper text.",
    },
    {
        "route": "/dashboard/college/hiring-assessment",
        "roles": ["college_admin"],
        "priority": "P1",
        "designDebtSignals": "Data-heavy assessment views; export buttons.",
        "shadcnV0Targets": "DataTable, Chart placeholder area (optional), Toolbar.",
    },
    {
        "route": "/dashboard/employer/hiring-assessment",
        "roles": ["employer"],
        "priority": "P1",
        "designDebtSignals": "Parallel to college hiring assessment; consistency opportunity.",
        "shadcnV0Targets": "Shared assessment table pattern with college side.",
    },
    {
        "route": "/dashboard/employer/assessment-uploads",
        "roles": ["employer"],
        "priority": "P1",
        "designDebtSignals": "Upload + template UX; error tables.",
        "shadcnV0Targets": "Dropzone, Table for validation errors, Progress.",
    },
    {
        "route": "/dashboard/college/reports",
        "roles": ["college_admin"],
        "priority": "P2",
        "designDebtSignals": "Analytics page; export split button; filters.",
        "shadcnV0Targets": "Tabs per report, Card KPI grid, Chart components (recharts optional).",
    },
    {
        "route": "/dashboard/college/calendar",
        "roles": ["college_admin"],
        "priority": "P2",
        "designDebtSignals": "Calendar views are hard to get to v0 polish without dedicated primitives.",
        "shadcnV0Targets": "Calendar (community) or structured month grid + Sheet event detail.",
    },
    {
        "route": "/dashboard/employer/calendar",
        "roles": ["employer"],
        "priority": "P2",
        "designDebtSignals": "Employer-side calendar grid consistency.",
        "shadcnV0Targets": "Match college calendar component language.",
    },
    {
        "route": "/dashboard/student/calendar",
        "roles": ["student"],
        "priority": "P2",
        "designDebtSignals": "Student calendar readability.",
        "shadcnV0Targets": "Mobile-first list + expandable day groups.",
    },
    {
        "route": "/dashboard/alerts",
        "roles": ["student", "employer", "college_admin"],
        "priority": "P2",
        "designDebtSignals": "Notification list density; preference controls if present.",
        "shadcnV0Targets": "List with Switch preferences, Empty state illustration.",
    },
    {
        "route": "/dashboard/feedback",
        "roles": ["student", "employer", "college_admin"],
        "priority": "P2",
        "designDebtSignals": "Form + thread UI opportunity.",
        "shadcnV0Targets": "Textarea + Card thread, Badge status.",
    },
    {
        "route": "/dashboard/admin/users",
        "roles": ["super_admin"],
        "priority": "P2",
        "designDebtSignals": "Platform directory table.",
        "shadcnV0Targets": "DataTable, role filter Combobox.",
    },
    {
        "route": "/dashboard/admin/colleges",
        "roles": ["super_admin"],
        "priority": "P2",
        "designDebtSignals": "Tenant management table/forms.",
        "shadcnV0Targets": "DataTable + Sheet create/edit.",
    },
    {
        "route": "/dashboard/admin/employers",
        "roles": ["super_admin"],
        "priority": "P2",
        "designDebtSignals": "Directory + actions.",
        "shadcnV0Targets": "DataTable, DropdownMenu actions.",
    },
    {
        "route": "/dashboard/admin/pending-registrations",
        "roles": ["super_admin"],
        "priority": "P2",
        "designDebtSignals": "Queue UI; approve/deny clarity.",
        "shadcnV0Targets": "Card queue or Table, AlertDialog.",
    },
    {
        "route": "/dashboard/admin/feedback",
        "roles": ["super_admin"],
        "priority": "P3",
        "designDebtSignals": "Inbox pattern.",
        "shadcnV0Targets": "Resizable panels (optional) or Table + detail Sheet.",
    },
    {
        "route": "/dashboard/help",
        "roles": ["student", "employer", "college_admin", "super_admin"],
        "priority": "P3",
        "designDebtSignals": "Long-form content presentation.",
        "shadcnV0Targets": "Accordion FAQ, Typography prose styling.",
    },
    {
        "route": "/data-entry/users",
        "roles": ["super_admin"],
        "priority": "P3",
        "designDebtSignals": "Utility screens often lag product chrome.",
        "shadcnV0Targets": "Align with admin DataTable patterns.",
    },
]


def build_cases():
    cases = []
    for i, row in enumerate(SCREENS, start=1):
        cases.append(
            {
                "testCaseId": f"C11-DESIGN-{i:03d}",
                "category": "C-11",
                "type": "design_refresh_candidate",
                "title": f"Refresh UI toward shadcn/v0 quality: {row['route']}",
                "route": row["route"],
                "roles": row["roles"],
                "priority": row["priority"],
                "designDebtSignals": row["designDebtSignals"],
                "suggestedShadcnV0Elements": row["shadcnV0Targets"],
                "acceptanceCriteria": [
                    "Visual density and spacing match a coherent design system (8px grid, consistent radii/shadows).",
                    "Interactive states: focus ring, hover, disabled, loading (Skeleton) on primary paths.",
                    "Tables: sort/filter affordances, sticky header optional, empty and error states.",
                    "Forms: labels, descriptions, inline validation, accessible error association.",
                    "Responsive: Sheet/drawer navigation on small screens where applicable.",
                ],
                "automationHints": [
                    "Screenshot baseline per breakpoint (sm/md/lg) before/after.",
                    "Playwright checks for focus-visible on primary actions.",
                    "Optional: axe accessibility scan on refactored route.",
                ],
            }
        )
    cases.append(
        {
            "testCaseId": "C11-AUDIT-001",
            "category": "C-11",
            "type": "design_system_gap_register",
            "title": "Complete dashboard route inventory vs shadcn adoption plan",
            "route": "all /dashboard/** and /data-entry/**",
            "roles": ["all"],
            "priority": "P0",
            "designDebtSignals": "Repo dashboard pages do not import @/components/ui; global btn/card patterns dominate.",
            "suggestedShadcnV0Elements": "Introduce components/ui primitives and migrate shell first, then P0 tables.",
            "acceptanceCriteria": [
                "Every route classified: Done / In progress / Backlog.",
                "Shared tokens: typography, spacing, component variants documented.",
            ],
            "automationHints": ["Track bundle of routes in CI storybook or visual regression optional."],
        }
    )
    return cases


def write_json(cases):
    payload = {
        "category": "C-11",
        "categoryTitle": "Design refresh candidates (shadcn / v0 level)",
        "description": "Screens and shells prioritized for visual/interaction upgrade using shadcn-style primitives and v0-level polish.",
        "notes": [
            "Current dashboard code relies on custom CSS utility classes (e.g. btn, card) rather than @/components/ui.",
            "Highest ROI: layout shell, then dense operational tables (offers, applications, students, drives).",
        ],
        "testCases": cases,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_xlsx(cases):
    wb = Workbook()
    ws = wb.active
    ws.title = "C-11 Design backlog"
    headers = [
        "Test Case ID",
        "Category",
        "Priority",
        "Route",
        "Roles",
        "Design debt signals",
        "Suggested shadcn/v0 elements",
        "Acceptance criteria (summary)",
        "Status",
        "Owner",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    for c in cases:
        crit = c.get("acceptanceCriteria", [])
        summary = crit[0] if crit else ""
        ws.append(
            [
                c["testCaseId"],
                c["category"],
                c.get("priority", ""),
                c["route"],
                ", ".join(c.get("roles", [])),
                c.get("designDebtSignals", ""),
                c.get("suggestedShadcnV0Elements", ""),
                summary,
                "Backlog",
                "",
                "",
            ]
        )
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    for col, w in zip("ABCDEFGHIJK", [14, 8, 8, 36, 28, 48, 48, 40, 12, 14, 20]):
        ws.column_dimensions[col].width = w
    wb.save(OUTPUT_XLSX)


def main():
    cases = build_cases()
    write_json(cases)
    write_xlsx(cases)
    print(f"Generated {len(cases)} Category-11 items")
    print(OUTPUT_JSON)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
