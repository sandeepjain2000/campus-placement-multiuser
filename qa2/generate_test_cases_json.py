"""
Build qa2/test_cases.json from all *.xlsx in qa2/, doubling row count.

Each source row becomes:
  1) primary case (normalized from the sheet)
  2) companion case (*-COMP) with an extra resilience/a11y/data spot-check angle

Run from repo root:
  python qa2/generate_test_cases_json.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "test_cases.json"

# Rotating companion themes (generic SaaS QA — no domain-specific wording)
COMPANION_THEMES = [
    "Repeat on a narrow mobile viewport (≈390×844): primary actions remain reachable without horizontal scroll.",
    "Repeat with a hard refresh first; authenticated session and expected data still align with preconditions.",
    "Spot-check: sampled visible labels or metrics align with the backing list or detail source (no obvious stale placeholder).",
    "Critical path only using keyboard (Tab/Enter): focus order is logical and actions remain activatable.",
    "If the flow includes an irreversible or bulk action, assert an explicit confirmation or undo affordance exists.",
    "After success, re-open the same view: persisted state matches expectations (no phantom success).",
    "Error path: force invalid input once (if applicable) and confirm user-facing validation (no silent failure).",
    "Empty or edge state: when lists are empty, confirm helpful empty state (not a broken shell).",
]

def row_to_dict(headers: list[str], values: tuple) -> dict[str, str]:
    d: dict[str, str] = {}
    for i, h in enumerate(headers):
        v = values[i] if i < len(values) else None
        d[h.strip()] = "" if v is None else str(v).strip()
    return d


def source_path_to_suite(filename: str) -> str:
    m = re.match(r"C-(\d+)", filename, re.I)
    if not m:
        return "A"
    n = int(m.group(1))
    suites = "ABCDEFGHIJKLMNO"
    if 1 <= n <= len(suites):
        return suites[n - 1]
    return "A"


def load_all_cases() -> list[dict]:
    cases: list[dict] = []
    seq = 0
    for xlsx in sorted(ROOT.glob("C-*.xlsx")):
        suite = source_path_to_suite(xlsx.name)
        wb = load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers_raw = next(rows_iter, None)
        if not headers_raw:
            continue
        headers = [str(h).strip() if h is not None else "" for h in headers_raw]
        for values in rows_iter:
            if not values or not any(v is not None and str(v).strip() for v in values):
                continue
            row = row_to_dict(headers, values)
            tid = row.get("Test Case ID") or row.get("testCaseId") or f"{suite}-{seq:04d}"
            seq += 1
            cases.append(
                {
                    "id": tid,
                    "suite": suite,
                    "sourceFile": xlsx.name,
                    "raw": row,
                    "title": row.get("title") or row.get("Title") or row.get("Scenario") or tid,
                    "persona": row.get("Role") or row.get("role") or row.get("Persona") or row.get("Demo Email") or "",
                    "preconditions": row.get("Preconditions") or row.get("preConditions") or row.get("Pre-conditions") or "",
                    "steps": row.get("Steps") or row.get("steps") or "",
                    "expected": row.get("Expected Result") or row.get("expected") or row.get("Expected") or row.get("checks") or "",
                    "priority": row.get("Priority") or row.get("priority") or "medium",
                    "kind": "primary",
                }
            )
    return cases


def add_companions(cases: list[dict]) -> list[dict]:
    out: list[dict] = []
    for i, c in enumerate(cases):
        out.append(c)
        theme = COMPANION_THEMES[i % len(COMPANION_THEMES)]
        cid = f"{c['id']}-COMP"
        out.append(
            {
                "id": cid,
                "suite": c["suite"],
                "sourceFile": c["sourceFile"],
                "raw": {**(c.get("raw") or {}), "companionTheme": theme},
                "title": f"{c['title']} — companion: resilience/UX spot-check",
                "persona": c["persona"],
                "preconditions": c["preconditions"],
                "steps": f"{c['steps']}\n\nAdditional focus:\n{theme}",
                "expected": f"{c['expected']}\n\nAdditionally: {theme} should hold or be noted as N/A with rationale.",
                "priority": c["priority"],
                "kind": "companion",
                "parentId": c["id"],
            }
        )
    return out


def main() -> None:
    primaries = load_all_cases()
    doubled = add_companions(primaries)
    payload = {
        "version": 2,
        "description": "Expanded from qa2 XLSX: each row plus a generic companion case (a11y, viewport, validation, empty state, etc.).",
        "count": len(doubled),
        "primaryCount": len(primaries),
        "cases": doubled,
    }
    OUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} — {len(primaries)} primary, {len(doubled)} total")


if __name__ == "__main__":
    main()
