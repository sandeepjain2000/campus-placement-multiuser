import openpyxl

file_path = r"C:\Users\sandeep\Downloads\Claudes\Applicant_List_for_Intern_-_CSE_Internship_Drive_M_at_Unique_School_App_India_LLP_Internship_Batch_2027.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("Headers:", [cell.value for cell in ws[1]])
print("\nFirst 3 rows of data:")
for row in range(2, 5):
    print(f"Row {row}: {[cell.value for cell in ws[row]]}")
