import openpyxl
from urllib.parse import quote
import os

file_path = r"C:\Users\sandeep\Downloads\Claudes\Applicant_List_for_Intern_-_CSE_Internship_Drive_M_at_Unique_School_App_India_LLP_Internship_Batch_2027.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Find the Current Address column index
    address_col_idx = None
    for cell in ws[1]:
        if cell.value == 'Current Address':
            address_col_idx = cell.column
            break

    if address_col_idx:
        link_col_idx = address_col_idx + 1
        ws.cell(row=1, column=link_col_idx, value='Google Maps Link')
        
        # Iterate through data rows
        for row in range(2, ws.max_row + 1):
            address = ws.cell(row=row, column=address_col_idx).value
            if address and str(address).strip():
                # Clean up the address string for the URL
                clean_address = str(address).replace('\n', ' ').strip()
                encoded_address = quote(clean_address)
                url = f"https://www.google.com/maps/search/?api=1&query={encoded_address}"
                
                # Create HYPERLINK formula
                # Excel formula: =HYPERLINK("url", "friendly_name")
                # We need to escape double quotes in the URL if any (unlikely with quote)
                formula = f'=HYPERLINK("{url}", "View on Map")'
                ws.cell(row=row, column=link_col_idx, value=formula)
                
                # Apply blue color and underline to make it look like a link
                ws.cell(row=row, column=link_col_idx).font = openpyxl.styles.Font(color="0000FF", underline="single")

        wb.save(file_path)
        print(f"Successfully added 'Google Maps Link' column to {file_path}")
    else:
        print("Error: 'Current Address' column not found.")

except Exception as e:
    print(f"An error occurred: {e}")
