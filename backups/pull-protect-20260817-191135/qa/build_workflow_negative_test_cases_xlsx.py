"""
Generate docs/test-cases/PlacementHub-Test-Cases-Workflow-Negative.xlsx

Focused suite for Placement Drives, Internships, and Offers:
  - Happy-path / state-machine transitions
  - Negative, boundary, and auth cases

Generated with patterns from QASkills:
  test-case-generator-user-stories, negative-test-generator,
  boundary-value-generator, state-machine-test-generator
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "test-cases" / "PlacementHub-Test-Cases-Workflow-Negative.xlsx"

HEADERS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
    "Actual Result",
    "Executed At",
]

CASES: list[tuple[str, str, list[tuple]]] = []


def add(sheet: str, module: str, rows: list[tuple]):
    CASES.append((sheet, module, rows))


# ── 01 Drives happy / state machine ─────────────────────────────────────────
add(
    "01 Drives workflow",
    "01 Placement drives — workflow",
    [
        (
            "Partnership gate",
            "Employer with approved campus can open Create Drive",
            "P0",
            "Functional",
            "Employer",
            "employer_approvals.status=approved for at least one campus",
            "1. Sign in as employer\n2. Open /dashboard/employer/drives\n3. Start create drive for approved campus",
            "Create form loads; campus selectable; no partnership error",
            "Requires approved tie-up",
            "Candidate",
        ),
        (
            "Create → requested",
            "Creating a drive sets status requested",
            "P0",
            "State",
            "Employer",
            "Approved partnership; valid title + driveDate + driveType",
            "1. Submit new drive with on_campus type\n2. Observe list / detail status",
            "Drive status = requested; appears in college Unapproved tab",
            "drive_type ∈ on_campus|off_campus|virtual|hybrid",
            "Candidate",
        ),
        (
            "Approve → approved",
            "College approve moves requested → approved",
            "P0",
            "State",
            "College Admin",
            "Drive in requested; no blocking calendar clash (or force)",
            "1. Open /dashboard/college/drives Unapproved\n2. Approve drive\n3. Check Active tab",
            "Status = approved; employer notified; students can browse",
            "Optional force:true on CALENDAR_CLASH",
            "Candidate",
        ),
        (
            "Reject → cancelled",
            "College reject of requested drive ends lifecycle",
            "P0",
            "State",
            "College Admin",
            "Drive status = requested",
            "1. Reject drive with reason if prompted\n2. Check Rejected tab",
            "Status becomes cancelled (or rejected); not visible as open to students",
            "College tabs treat cancelled|rejected as rejected",
            "Candidate",
        ),
        (
            "Student apply",
            "Eligible student can apply when drive is approved/scheduled",
            "P0",
            "Functional",
            "Student",
            "Drive status ∈ {approved, scheduled}; student meets CGPA/branch/resume gates",
            "1. Open /dashboard/student/drives\n2. Apply to drive\n3. Open My Drives applications",
            "Application status = applied; unique (student_id, drive_id)",
            "Campus student (not alumni)",
            "Candidate",
        ),
        (
            "Pipeline shortlist",
            "Employer shortlist moves application applied → shortlisted",
            "P0",
            "State",
            "Employer",
            "Student application = applied",
            "1. Open Applications for drive\n2. Shortlist candidate (UI or hiring result Shortlist)",
            "Application status = shortlisted; UI shows Shortlisted",
            "Hiring map: Shortlist → shortlisted",
            "Candidate",
        ),
        (
            "Select → formal offer pending",
            "Select sets application selected awaiting formal offer",
            "P0",
            "State",
            "Employer",
            "Application in pipeline (applied/shortlisted/in_progress)",
            "1. Mark Select / selected\n2. Observe student/employer application label",
            "Application = selected; UI indicates formal offer pending",
            "Hiring map: Select → selected",
            "Candidate",
        ),
        (
            "Employer cancel open drive",
            "Employer can cancel drive in requested/approved/scheduled/in_progress",
            "P1",
            "State",
            "Employer",
            "Drive status ∈ {requested, approved, scheduled, in_progress}",
            "1. Cancel drive from employer UI\n2. Refresh college/student views",
            "Status = cancelled; new applies blocked",
            None,
            "Candidate",
        ),
        (
            "Withdrawn terminal",
            "Student withdraw ends employer update path",
            "P1",
            "State",
            "Student / Employer",
            "Application = applied",
            "1. Student withdraws application\n2. Employer attempts status update",
            "Application = withdrawn; employer cannot update withdrawn row",
            None,
            "Candidate",
        ),
    ],
)

# ── 02 Drives negative / boundary ───────────────────────────────────────────
add(
    "02 Drives negative",
    "02 Placement drives — negative & boundary",
    [
        (
            "No partnership",
            "Employer without approved campus cannot create drive",
            "P0",
            "Negative",
            "Employer",
            "No approved employer_approvals (pending/rejected only)",
            "1. Attempt POST /api/employer/drives or Create Drive UI",
            "403 or empty campus / blocked create; clear partnership message",
            "Security + authz",
            "Candidate",
        ),
        (
            "Missing title",
            "Drive create rejects missing title",
            "P0",
            "Negative",
            "Employer",
            "Approved campus",
            "1. Submit create with empty title",
            "Validation error; drive not created",
            "Required: tenantId, title, valid driveDate",
            "Candidate",
        ),
        (
            "Invalid driveType",
            "Drive create rejects driveType outside allowed set",
            "P1",
            "Negative",
            "Employer",
            "Approved campus",
            "1. Submit driveType=bogus (API)",
            "400 validation error",
            "Allowed: on_campus|off_campus|virtual|hybrid",
            "Candidate",
        ),
        (
            "Approve non-requested",
            "College cannot approve already approved/cancelled drive",
            "P0",
            "Negative",
            "College Admin",
            "Drive status already approved or cancelled",
            "1. POST approve on non-requested drive",
            "409 conflict; status unchanged",
            "State guard",
            "Candidate",
        ),
        (
            "Calendar clash",
            "Approve with calendar clash returns 409 unless force",
            "P1",
            "Negative",
            "College Admin",
            "Drive date clashes with college calendar event",
            "1. Approve without force\n2. Retry with force:true",
            "First: 409 CALENDAR_CLASH; second: approve succeeds",
            None,
            "Candidate",
        ),
        (
            "Apply closed status",
            "Student cannot apply when drive is requested/cancelled/completed",
            "P0",
            "Negative",
            "Student",
            "Drive status not in {approved, scheduled}",
            "1. Attempt apply (UI/API)",
            "Blocked with status/eligibility message",
            "Open statuses: approved, scheduled",
            "Candidate",
        ),
        (
            "Duplicate apply",
            "Second apply to same drive is rejected",
            "P0",
            "Negative",
            "Student",
            "Existing application for (student, drive)",
            "1. Apply again to same drive",
            "Unique constraint / clear already-applied error",
            "unique (student_id, drive_id)",
            "Candidate",
        ),
        (
            "No resume",
            "Apply blocked when resume/CV gate fails",
            "P0",
            "Negative",
            "Student",
            "Student profile missing resume or CV verification required",
            "1. Attempt apply",
            "Explicit error about resume/CV verification",
            "College rules may require verified CV",
            "Candidate",
        ),
        (
            "CGPA below min",
            "Apply blocked when CGPA below drive minimum",
            "P0",
            "Boundary",
            "Student",
            "Drive min CGPA = 7.0; student CGPA = 6.99",
            "1. Attempt apply",
            "Eligibility failure citing CGPA",
            "Boundary: min-ε fails; min passes",
            "Candidate",
        ),
        (
            "CGPA at min",
            "Apply allowed when CGPA equals drive minimum",
            "P1",
            "Boundary",
            "Student",
            "Drive min CGPA = 7.0; student CGPA = 7.0; other gates pass",
            "1. Apply",
            "Application created (applied)",
            "Inclusive lower bound",
            "Candidate",
        ),
        (
            "Placement lock",
            "Student with accepted placement offer cannot apply to new drives",
            "P0",
            "Negative",
            "Student",
            "Student already accepted a placement offer (placement lock)",
            "1. Attempt apply to another open drive",
            "Blocked with placement lock message",
            "After offer accept",
            "Candidate",
        ),
        (
            "Alumni hidden",
            "Alumni student does not see campus drives browse",
            "P1",
            "Negative",
            "Student (alumni)",
            "User flagged alumni",
            "1. Open student nav / drives routes",
            "Campus drives hidden or redirected to alumni jobs flows",
            "Campus-only nav",
            "Candidate",
        ),
        (
            "Cross-role create",
            "College/student cannot create employer drive",
            "P0",
            "Security",
            "College Admin / Student",
            "Signed in as non-employer",
            "1. POST /api/employer/drives",
            "401/403",
            None,
            "Candidate",
        ),
        (
            "Committee no approve",
            "Placement committee cannot approve drives",
            "P0",
            "Security",
            "Placement Committee",
            "Committee session; drive requested",
            "1. Attempt college drive approve API/UI",
            "Denied; drives approve not in committee menu",
            "Committee: students + applications read-only subset",
            "Candidate",
        ),
    ],
)

# ── 03 Internships workflow ─────────────────────────────────────────────────
add(
    "03 Internships workflow",
    "03 Internships — workflow",
    [
        (
            "Draft create",
            "Employer can save internship as draft without full publish fields",
            "P0",
            "Functional",
            "Employer",
            "Approved partnership",
            "1. Create internship as draft\n2. Open employer internships list",
            "jp.status = draft; not visible to students",
            "Dates/campuses looser in draft",
            "Candidate",
        ),
        (
            "Publish → pending visibility",
            "Publish creates college visibility pending per campus",
            "P0",
            "State",
            "Employer",
            "Title; start/end; ≥1 approved campus; batch/CGPA/backlogs valid",
            "1. Publish internship\n2. Check college Internships pending",
            "jp.status=published; jpv.college_status=pending for selected campuses",
            "end >= start",
            "Candidate",
        ),
        (
            "College approve visibility",
            "College approve pending → approved makes listing student-visible",
            "P0",
            "State",
            "College Admin",
            "Visibility row college_status=pending",
            "1. Approve internship visibility\n2. Student browses internships",
            "college_status=approved; student sees listing",
            "Must be published AND approved visibility AND approved tie-up",
            "Candidate",
        ),
        (
            "College reject visibility",
            "Reject pending visibility hides from students",
            "P0",
            "State",
            "College Admin",
            "college_status=pending",
            "1. Reject listing\n2. Student refresh",
            "college_status=rejected; not listed for students",
            "Approve also allowed from rejected in some flows",
            "Candidate",
        ),
        (
            "Student apply",
            "Eligible student applies → program_applications applied",
            "P0",
            "Functional",
            "Student",
            "Published + approved visibility; eligibility gates pass; no FCFS lock",
            "1. Apply from /dashboard/student/internships\n2. Open My Internships",
            "Application status=applied; unique (student_id, job_id)",
            None,
            "Candidate",
        ),
        (
            "Select → FCFS internship lock",
            "First selected internship locks student from other internship applies",
            "P0",
            "State",
            "Employer / Student",
            "fcfs_enabled true (default); application selectable",
            "1. Employer Select on internship A\n2. Student tries apply internship B",
            "A=selected; B apply blocked with FCFS / one-internship message",
            "MAX_INTERNSHIPS_PER_STUDENT = 1",
            "Candidate",
        ),
        (
            "Pipeline parity",
            "Shortlist / in_progress / rejected mirror drive application states",
            "P1",
            "State",
            "Employer",
            "Internship application = applied",
            "1. Shortlist then move to in_progress then reject another candidate",
            "Statuses update: shortlisted → in_progress → rejected as applicable",
            "Same enum as drives applications",
            "Candidate",
        ),
    ],
)

# ── 04 Internships negative / boundary ──────────────────────────────────────
add(
    "04 Internships negative",
    "04 Internships — negative & boundary",
    [
        (
            "Publish no campus",
            "Publish rejected when no approved campus selected",
            "P0",
            "Negative",
            "Employer",
            "Form otherwise filled",
            "1. Publish with zero campuses",
            "Error: select at least one approved campus",
            None,
            "Candidate",
        ),
        (
            "End before start",
            "Publish rejects end date before start date",
            "P0",
            "Boundary",
            "Employer",
            "start=2026-08-10; end=2026-08-09",
            "1. Publish",
            "Validation error; end >= start required",
            "Boundary: end=start allowed if product allows same-day",
            "Candidate",
        ),
        (
            "Equal dates",
            "Publish with start=end date (same day internship)",
            "P2",
            "Boundary",
            "Employer",
            "start=end valid calendar day; other fields valid",
            "1. Publish",
            "Accepted if product allows; otherwise clear date-range error",
            "Confirm product rule during exec",
            "Manual",
        ),
        (
            "Approve non-pending",
            "College approve when not awaiting approval returns 409",
            "P1",
            "Negative",
            "College Admin",
            "college_status already approved",
            "1. Approve again",
            "409; status unchanged",
            None,
            "Candidate",
        ),
        (
            "Reject non-pending",
            "Reject only allowed from pending",
            "P1",
            "Negative",
            "College Admin",
            "college_status=approved",
            "1. Attempt reject",
            "409 or denied",
            None,
            "Candidate",
        ),
        (
            "Duplicate internship apply",
            "Cannot apply twice to same job_id",
            "P0",
            "Negative",
            "Student",
            "Existing program_application",
            "1. Apply again",
            "Unique (student_id, job_id) enforced",
            None,
            "Candidate",
        ),
        (
            "FCFS unavailable to employer",
            "Other employers see FCFS unavailable after first Select",
            "P1",
            "Negative",
            "Employer B",
            "Student selected on internship track by Employer A; fcfs_enabled",
            "1. Employer B opens FCFS Unavailable / tries Select",
            "Select blocked; student appears unavailable for internship track",
            "Tracks: internship | placement | jobs",
            "Candidate",
        ),
        (
            "Stipend/vacancies invalid",
            "Negative stipend or zero vacancies rejected on publish",
            "P1",
            "Boundary",
            "Employer",
            "Otherwise valid publish payload",
            "1. Set stipend=-1 or vacancies=0\n2. Publish",
            "Validation error",
            None,
            "Candidate",
        ),
        (
            "Cross-role publish",
            "Student cannot publish internship",
            "P0",
            "Security",
            "Student",
            "Student session",
            "1. POST employer internships/jobs publish API",
            "401/403",
            None,
            "Candidate",
        ),
    ],
)

# ── 05 Offers workflow ──────────────────────────────────────────────────────
add(
    "05 Offers workflow",
    "05 Offers — workflow",
    [
        (
            "Create pending",
            "Employer creates formal offer → pending after selection",
            "P0",
            "State",
            "Employer",
            "Student application selected (drive or internship)",
            "1. Create offer from Offers UI with job title + salary + deadline\n2. Student opens My Offers",
            "Offer status=pending; student can respond",
            "Aliases offered/sent normalize to pending",
            "Candidate",
        ),
        (
            "College create",
            "College admin can create/monitor campus offers",
            "P0",
            "Functional",
            "College Admin",
            "Campus student identifiable; selection context valid",
            "1. Create offer (UI or Offers-page CSV)\n2. Verify pending on student My Offers",
            "Offer pending; college sees it in Offers",
            "CSV link on Offers page (not sidebar)",
            "Candidate",
        ),
        (
            "Accept → accepted",
            "Student accept pending → accepted with accepted_at",
            "P0",
            "State",
            "Student",
            "Offer status=pending; within acceptance window; under max offers",
            "1. Accept on My Offers\n2. Refresh",
            "Status=accepted; accepted_at set; placement lock if placement offer",
            None,
            "Candidate",
        ),
        (
            "Decline → rejected",
            "Student decline maps to rejected",
            "P0",
            "State",
            "Student",
            "Offer pending",
            "1. Decline / reject offer",
            "Status=rejected (API maps reject→decline)",
            "Aliases declined/decline/reject → rejected",
            "Candidate",
        ),
        (
            "Revoke pending",
            "Employer revokes pending → revoked",
            "P0",
            "State",
            "Employer",
            "Offer pending",
            "1. Revoke offer\n2. Student refresh",
            "Status=revoked; student cannot accept",
            "Revoke only from pending",
            "Candidate",
        ),
        (
            "Reopen to pending",
            "Employer/college can reopen accepted/rejected/revoked/expired → pending",
            "P1",
            "State",
            "Employer / College Admin",
            "Offer in terminal status",
            "1. Reopen offer\n2. Student sees pending again",
            "Status=pending; respond enabled again",
            None,
            "Candidate",
        ),
        (
            "Expire path",
            "Past deadline offer becomes expired; respond returns 410",
            "P1",
            "State",
            "Student / System",
            "Offer pending with deadline in the past",
            "1. Attempt accept after deadline",
            "410 or expired status; cannot accept",
            "Date-only deadline valid through end of UTC day",
            "Candidate",
        ),
    ],
)

# ── 06 Offers negative / boundary ───────────────────────────────────────────
add(
    "06 Offers negative",
    "06 Offers — negative & boundary",
    [
        (
            "Accept non-pending",
            "Student cannot accept already accepted/rejected/revoked offer",
            "P0",
            "Negative",
            "Student",
            "Offer not pending",
            "1. POST accept",
            "409 conflict",
            None,
            "Candidate",
        ),
        (
            "Revoke non-pending",
            "Employer revoke on non-pending returns 400",
            "P1",
            "Negative",
            "Employer",
            "Offer accepted",
            "1. Attempt revoke",
            "400; status unchanged",
            None,
            "Candidate",
        ),
        (
            "Max offers",
            "Accept blocked when at max_offers_per_student",
            "P0",
            "Boundary",
            "Student",
            "College max_offers_per_student=1; student already has 1 accepted",
            "1. Accept second pending offer",
            "403 with college rule message",
            "Default max ≥1",
            "Candidate",
        ),
        (
            "Acceptance window",
            "Accept blocked after offer_acceptance_window_days from first accept",
            "P1",
            "Boundary",
            "Student",
            "Window=7 days; first accept was 8 days ago; another pending offer",
            "1. Attempt accept",
            "403 citing acceptance window",
            "Default window 7 days",
            "Candidate",
        ),
        (
            "Missing salary/deadline",
            "Create offer rejects invalid salary or deadline",
            "P0",
            "Negative",
            "Employer",
            "Selected student context",
            "1. Submit without job title or with invalid deadline",
            "Validation error; no offer row",
            None,
            "Candidate",
        ),
        (
            "Deadline today inclusive",
            "Student can accept on deadline calendar day",
            "P1",
            "Boundary",
            "Student",
            "Offer pending; deadline = today (date-only)",
            "1. Accept before end of UTC day",
            "Accept succeeds (not expired yet)",
            "Valid through end of UTC day",
            "Candidate",
        ),
        (
            "Cross-role accept",
            "Employer/college cannot accept as student",
            "P0",
            "Security",
            "Employer / College Admin",
            "Offer pending for a student",
            "1. Call student respond API as employer/college",
            "401/403",
            None,
            "Candidate",
        ),
        (
            "Committee no create",
            "Placement committee cannot create offers",
            "P0",
            "Security",
            "Placement Committee",
            "Committee session",
            "1. Attempt offer create API/UI",
            "Denied; no offers create menu",
            None,
            "Candidate",
        ),
    ],
)

# ── 07 Cross-cutting ────────────────────────────────────────────────────────
add(
    "07 Cross-cutting",
    "07 Partnership, FCFS, assessment — cross-cutting",
    [
        (
            "Partnership pending",
            "Pending partnership blocks drive/internship create for that campus",
            "P0",
            "Negative",
            "Employer",
            "employer_approvals.status=pending",
            "1. Try create drive/internship for that campus",
            "Blocked until college approves partnership",
            "Statuses: approved|pending|rejected (+revoked/blacklisted UI)",
            "Candidate",
        ),
        (
            "Assessment CSV bad roll",
            "Assessment import rejects unknown roll numbers",
            "P1",
            "Negative",
            "Employer",
            "Assessment upload session open",
            "1. Upload CSV with invalid rolls / rounds\n2. Review staging",
            "Rows rejected or flagged for fix; submit blocked until fixed",
            None,
            "Candidate",
        ),
        (
            "Withdrawn no hiring update",
            "Hiring result cannot move withdrawn application",
            "P1",
            "Negative",
            "Employer",
            "Application withdrawn",
            "1. Submit hiring Select/Reject for that row",
            "Update rejected / ignored with clear reason",
            None,
            "Candidate",
        ),
        (
            "Placement FCFS track",
            "First placement Select claims FCFS placement track for student",
            "P1",
            "State",
            "Employer A / Employer B",
            "fcfs_enabled; two employers interviewing same student on drives",
            "1. A Selects\n2. B tries Select on placement track",
            "B blocked; student on Unavailable (FCFS) for placement",
            "Track: placement",
            "Candidate",
        ),
        (
            "Unauthenticated API",
            "Drive/internship/offer APIs require session",
            "P0",
            "Security",
            "Anonymous",
            "No session cookie",
            "1. GET/POST employer drives, college approve, student offers without auth",
            "401 Unauthorized",
            None,
            "Candidate",
        ),
        (
            "Happy E2E smoke",
            "Drive request → college approve → student apply → select → offer accept",
            "P0",
            "E2E",
            "Employer / College / Student",
            "Approved partnership; eligible student; empty FCFS state",
            "1. Employer create drive\n2. College approve\n3. Student apply\n4. Employer select\n5. Employer create offer\n6. Student accept",
            "Final: application selected; offer accepted; placement lock on",
            "Primary smoke for release",
            "Guided / Playwright candidate",
        ),
        (
            "Internship E2E smoke",
            "Publish → college approve visibility → apply → select → offer",
            "P0",
            "E2E",
            "Employer / College / Student",
            "Approved partnership; eligible student",
            "1. Publish internship\n2. College approve visibility\n3. Student apply\n4. Select\n5. Formal offer accept",
            "End states correct; FCFS internship lock engaged after select",
            None,
            "Guided / Playwright candidate",
        ),
    ],
)


def style_header(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row, c, h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws):
    widths = [12, 28, 18, 42, 8, 12, 22, 36, 40, 40, 28, 14, 10, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = Workbook()
    idx = wb.active
    idx.title = "Index"
    title = idx.cell(1, 1, "PlacementHub — Workflow & Negative Test Cases")
    title.font = Font(name="Arial", bold=True, size=14, color="1E3A8A")
    idx.merge_cells("A1:G1")
    sub = idx.cell(
        2,
        1,
        "Drives · Internships · Offers — happy paths, state transitions, negatives, boundaries. "
        "Generated with QASkills patterns (user stories, negative, boundary, state machine).",
    )
    sub.font = Font(name="Arial", size=10, color="475569")
    idx.merge_cells("A2:G2")

    headers_idx = ["#", "Sheet", "Cases", "P0", "P1", "P2+", "Link"]
    head_font = Font(name="Arial", bold=True)
    body = Font(name="Arial", size=10)
    link_font = Font(name="Arial", color="2563EB", underline="single")
    for c, h in enumerate(headers_idx, 1):
        cell = idx.cell(4, c, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")

    total = p0 = p1 = p2 = 0
    for i, (sheet, _module, rows) in enumerate(CASES, 1):
        counts = {"P0": 0, "P1": 0, "P2": 0}
        for r in rows:
            pr = r[2]
            if pr.startswith("P0"):
                counts["P0"] += 1
            elif pr.startswith("P1"):
                counts["P1"] += 1
            else:
                counts["P2"] += 1
        n = len(rows)
        total += n
        p0 += counts["P0"]
        p1 += counts["P1"]
        p2 += counts["P2"]
        row = 4 + i
        idx.cell(row, 1, i).font = body
        idx.cell(row, 2, sheet).font = body
        idx.cell(row, 3, n).font = body
        idx.cell(row, 4, counts["P0"]).font = body
        idx.cell(row, 5, counts["P1"]).font = body
        idx.cell(row, 6, counts["P2"]).font = body
        link = idx.cell(row, 7, "Open ›")
        link.font = link_font
        link.hyperlink = f"#'{sheet}'!A1"

    sum_row = 5 + len(CASES)
    idx.cell(sum_row, 2, "TOTAL").font = head_font
    idx.cell(sum_row, 3, total).font = head_font
    idx.cell(sum_row, 4, p0).font = head_font
    idx.cell(sum_row, 5, p1).font = head_font
    idx.cell(sum_row, 6, p2).font = head_font

    idx.cell(sum_row + 2, 1, "How to use").font = head_font
    idx.cell(
        sum_row + 3,
        1,
        "1) Use demo employer + college + eligible student with approved partnership. "
        "2) Execute P0 first (blockers). 3) Fill Status / Actual / Executed At. "
        "4) Complements docs/test-cases/PlacementHub-Test-Cases.xlsx — do not treat as full regression. "
        "5) Regenerate: npm run qa:workflow-negative-xlsx",
    ).font = body
    idx.merge_cells(start_row=sum_row + 3, start_column=1, end_row=sum_row + 3, end_column=7)
    idx.row_dimensions[sum_row + 3].height = 56

    for c, w in enumerate([6, 28, 12, 8, 8, 8, 10], 1):
        idx.column_dimensions[get_column_letter(c)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    font = Font(name="Arial", size=10)
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for sheet, module, rows in CASES:
        ws = wb.create_sheet(sheet)
        back = ws.cell(1, 1, "« Back to Index")
        back.font = link_font
        back.hyperlink = "#Index!A1"
        style_header(ws, 2)
        prefix = sheet.split(" ", 1)[0]
        for i, r in enumerate(rows, 1):
            feature, title, priority, typ, roles, pre, steps, expected, notes, auto = r
            values = [
                f"TC-WN-{prefix}-{i:03d}",
                module,
                feature,
                title,
                priority,
                typ,
                roles,
                pre,
                steps,
                expected,
                notes,
                auto,
                "",
                "",
                "",
            ]
            row_idx = 2 + i
            for col, val in enumerate(values, 1):
                cell = ws.cell(row_idx, col, val)
                cell.font = font
                cell.alignment = wrap
                cell.border = thin
            ws.row_dimensions[row_idx].height = 72
        autosize(ws)

    wb.save(OUT)
    print(f"Wrote {OUT} with {total} cases ({p0} P0 / {p1} P1 / {p2} P2+)")


if __name__ == "__main__":
    main()
