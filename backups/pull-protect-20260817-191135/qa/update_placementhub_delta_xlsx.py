"""Write Status / Actual Result / Executed At into docs/test-cases/PlacementHub-Test-Cases-Delta-Post-Gen.xlsx."""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "docs" / "test-cases" / "PlacementHub-Test-Cases-Delta-Post-Gen.xlsx"
RESULTS = ROOT / "qa" / "data" / "placementhub_delta_results.json"

STATUS_FILLS = {
    "Pass": PatternFill("solid", fgColor="C6EFCE"),
    "Fail": PatternFill("solid", fgColor="FFC7CE"),
    "Not Run": PatternFill("solid", fgColor="FFEB9C"),
    "Blocked": PatternFill("solid", fgColor="D9D9D9"),
}
STATUS_FONTS = {
    "Pass": Font(name="Arial", bold=True, color="006100"),
    "Fail": Font(name="Arial", bold=True, color="9C0006"),
    "Not Run": Font(name="Arial", bold=True, color="9C5700"),
    "Blocked": Font(name="Arial", bold=True, color="595959"),
}
HEADER_FONT = Font(name="Arial", bold=True)
BODY_FONT = Font(name="Arial")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def clean(text: str) -> str:
    return re.sub(r"[\000-\010\013\014\016-\037]", "", str(text or ""))


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        vals = [str(v or "").lower() for v in row]
        if any(v == "tc id" for v in vals) or ("tc id" in "|".join(vals)):
            return i, list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    return None, None


def ensure_columns(ws, header_row, headers):
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers, start=1) if h}
    status_col = header_map.get("Status")
    actual_col = header_map.get("Actual Result")
    executed_col = header_map.get("Executed At")
    next_col = ws.max_column + 1
    if status_col is None:
        status_col = next_col
        ws.cell(header_row, status_col).value = "Status"
        ws.cell(header_row, status_col).font = HEADER_FONT
        next_col += 1
    if actual_col is None:
        actual_col = next_col
        ws.cell(header_row, actual_col).value = "Actual Result"
        ws.cell(header_row, actual_col).font = HEADER_FONT
        next_col += 1
    if executed_col is None:
        executed_col = next_col
        ws.cell(header_row, executed_col).value = "Executed At"
        ws.cell(header_row, executed_col).font = HEADER_FONT
    return status_col, actual_col, executed_col


def main() -> None:
    if not RESULTS.is_file():
        raise SystemExit(f"Missing {RESULTS}")
    payload = json.loads(RESULTS.read_text(encoding="utf-8"))
    by_id = {r["id"]: r for r in payload.get("results", [])}

    wb = load_workbook(XLSX)
    updated = 0
    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        ws = wb[sheet_name]
        header_row, headers = find_header_row(ws)
        if not header_row:
            continue
        status_col, actual_col, executed_col = ensure_columns(ws, header_row, headers)
        id_col = 1
        for row_idx in range(header_row + 1, ws.max_row + 1):
            tc_id = ws.cell(row_idx, id_col).value
            if not tc_id or tc_id not in by_id:
                continue
            r = by_id[tc_id]
            status = r.get("status") or "Not Run"
            cell = ws.cell(row_idx, status_col)
            cell.value = status
            cell.font = STATUS_FONTS.get(status, BODY_FONT)
            cell.fill = STATUS_FILLS.get(status, PatternFill())
            cell.border = THIN
            ws.cell(row_idx, actual_col).value = clean(r.get("actual", ""))
            ws.cell(row_idx, actual_col).font = BODY_FONT
            ws.cell(row_idx, actual_col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row_idx, executed_col).value = r.get("executedAt") or payload.get("executedAt")
            ws.cell(row_idx, executed_col).font = BODY_FONT
            updated += 1

    # Index summary
    if "Index" in wb.sheetnames:
        ws = wb["Index"]
        start = ws.max_row + 2
        ws.cell(start, 1).value = "Last automated run"
        ws.cell(start, 1).font = Font(name="Arial", bold=True, size=12)
        ws.cell(start + 1, 1).value = "Base URL"
        ws.cell(start + 1, 2).value = payload.get("baseUrl")
        ws.cell(start + 2, 1).value = "Executed at"
        ws.cell(start + 2, 2).value = payload.get("executedAt")
        summary = payload.get("summary") or {}
        ws.cell(start + 3, 1).value = "Pass"
        ws.cell(start + 3, 2).value = summary.get("Pass", 0)
        ws.cell(start + 4, 1).value = "Fail"
        ws.cell(start + 4, 2).value = summary.get("Fail", 0)
        ws.cell(start + 5, 1).value = "Blocked"
        ws.cell(start + 5, 2).value = summary.get("Blocked", 0)

    wb.save(XLSX)
    print(f"Updated {updated} rows in {XLSX}")
    print("Summary:", payload.get("summary"))


if __name__ == "__main__":
    main()
