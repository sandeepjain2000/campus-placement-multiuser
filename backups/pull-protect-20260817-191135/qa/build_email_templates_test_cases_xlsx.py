"""Generate docs/test-cases/PlacementHub-Test-Cases-Email-Templates.xlsx"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "test-cases" / "PlacementHub-Test-Cases-Email-Templates.xlsx"

HEADERS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
    "Actual Result",
    "Executed At",
]

CASES: list[tuple[str, str, list[tuple]]] = []


def add(sheet: str, module: str, rows: list[tuple]):
    CASES.append((sheet, module, rows))


add(
    "01 Admin platform",
    "01 Admin platform templates",
    [
        (
            "Access",
            "Super admin can open Email templates page",
            "P0",
            "Functional",
            "Super Admin",
            "Signed in as super_admin",
            "1. Open /dashboard/admin/email-templates\n2. Observe catalog of editable keys",
            "Page loads; cards for campus_guest_confirmation, sponsorship_thank_you, sponsorship_college_thanks_sponsor, sponsorship_donation_receipt",
            "Menu: Email templates",
            "Candidate",
        ),
        (
            "Access",
            "Non-admin cannot load admin email templates API",
            "P0",
            "Security",
            "College Admin / Employer / Student",
            "Signed in as non-super-admin",
            "1. GET /api/admin/email-templates",
            "401 Unauthorized",
            None,
            "Candidate",
        ),
        (
            "Edit save",
            "Admin can update subject and body of a platform template",
            "P0",
            "Functional",
            "Super Admin",
            "Editable template exists",
            "1. Change subject/body with placeholders intact\n2. Save template\n3. Reload page",
            "Toast success; values persist; updated_at refreshes",
            "PATCH /api/admin/email-templates",
            "Candidate",
        ),
        (
            "Validation",
            "Empty subject or body is rejected",
            "P1",
            "Negative",
            "Super Admin",
            "On email templates page",
            "1. Clear subject or body\n2. Save",
            "API/UI error; template not overwritten",
            None,
            "Candidate",
        ),
        (
            "Versions",
            "First admin publish seeds baseline and creates System v2",
            "P0",
            "Functional",
            "Super Admin",
            "Migration 116 applied; template has only baseline (or will seed on save)",
            "1. Note current wording\n2. Change body and Save\n3. Inspect Retained system versions list",
            "Baseline remains; new System vN appears; live form shows new wording",
            "system_email_template_versions",
            "Candidate",
        ),
        (
            "Versions",
            "Subsequent admin saves add versions without changing baseline",
            "P0",
            "Functional",
            "Super Admin",
            "At least baseline + one later version exist",
            "1. Save another wording change\n2. Compare version list",
            "Baseline content unchanged; new version appended; previous versions still listed",
            None,
            "Candidate",
        ),
        (
            "Versions",
            "Saving identical content does not create a duplicate version",
            "P1",
            "Functional",
            "Super Admin",
            "Template open with unchanged text",
            "1. Click Save without edits",
            "No new version row; optional 'No changes to publish' toast",
            None,
            "Manual",
        ),
        (
            "Unknown key",
            "PATCH with unknown templateKey returns 400",
            "P1",
            "Negative",
            "Super Admin",
            "Valid session",
            "1. PATCH with templateKey=not_a_real_key",
            "400 Unknown template key",
            None,
            "Candidate",
        ),
    ],
)

add(
    "02 College overrides",
    "02 College system email templates",
    [
        (
            "Access",
            "College admin opens Email templates (communication-templates)",
            "P0",
            "Functional",
            "College Admin",
            "College admin with tenant",
            "1. Open /dashboard/college/communication-templates",
            "Shows college-editable templates (thanks sponsor + donation receipt)",
            "Not employer keys",
            "Candidate",
        ),
        (
            "Scope",
            "College cannot edit employer-only template keys via API",
            "P0",
            "Security",
            "College Admin",
            "Valid college session",
            "1. PATCH /api/college/system-email-templates with campus_guest_confirmation",
            "400 This template cannot be edited by your college",
            None,
            "Candidate",
        ),
        (
            "Save override",
            "College save creates campus override (badge Your campus)",
            "P0",
            "Functional",
            "College Admin",
            "Template on Platform default",
            "1. Edit subject/body\n2. Save for campus\n3. Reload",
            "Badge Your campus; wording persists for this tenant only",
            "email_template_overrides scope=college",
            "Candidate",
        ),
        (
            "Isolation",
            "College A override does not affect College B",
            "P0",
            "Functional",
            "College Admin",
            "Two tenants; A has override",
            "1. As College B open same template key",
            "B still sees platform (or B's own) wording — not A's override",
            "Use two demo colleges",
            "Manual",
        ),
        (
            "Reset current",
            "Current platform default removes campus override",
            "P0",
            "Functional",
            "College Admin",
            "Campus override exists",
            "1. Click Current platform default\n2. Confirm form shows live system text",
            "Override deleted; badge Platform default; live system wording restored",
            "resetToPlatform: true",
            "Candidate",
        ),
        (
            "Restore baseline",
            "College can restore System Baseline after editing",
            "P0",
            "Functional",
            "College Admin",
            "Baseline retained; college has modified override; admin may have published later versions",
            "1. Restore system version → Baseline\n2. Observe subject/body",
            "Form matches baseline wording; campus can undo their change",
            "restoreSystemVersionId",
            "Candidate",
        ),
        (
            "Restore prior",
            "College can restore an older system version after platform moved on",
            "P0",
            "Functional",
            "College Admin",
            "Admin published v2+; college had override or wants older system text",
            "1. Choose Restore system version → System v2 (not current)\n2. Reload",
            "Wording matches selected system version; if not equal to live platform, stored as campus override pinned to that text",
            None,
            "Candidate",
        ),
        (
            "Runtime send",
            "Sponsorship auto-email uses college override when present",
            "P1",
            "Integration",
            "College Admin",
            "Override set; SMTP/Zepto configured; sponsorship payment triggers auto mail",
            "1. Trigger sponsorship thank-you/receipt path\n2. Check mail_delivery_logs / inbox",
            "Sent body/subject reflect campus override placeholders filled",
            "sponsorshipAutoEmails",
            "Manual",
        ),
        (
            "Auth",
            "Unauthenticated GET college system-email-templates returns 401",
            "P1",
            "Security",
            "Anonymous",
            "No session",
            "1. GET /api/college/system-email-templates",
            "401",
            None,
            "Candidate",
        ),
    ],
)

add(
    "03 Employer overrides",
    "03 Employer communication templates",
    [
        (
            "Access",
            "Employer opens communication templates page",
            "P0",
            "Functional",
            "Employer",
            "Employer with profile",
            "1. Open /dashboard/employer/communication-templates",
            "Shows employer keys: campus guest confirmation + sponsorship thank-you",
            None,
            "Candidate",
        ),
        (
            "Scope",
            "Employer cannot PATCH college-only template keys",
            "P0",
            "Security",
            "Employer",
            "Valid employer session",
            "1. PATCH /api/employer/email-templates with sponsorship_donation_receipt",
            "400 This template cannot be edited by employers",
            None,
            "Candidate",
        ),
        (
            "Save override",
            "Employer save creates organization override",
            "P0",
            "Functional",
            "Employer",
            "On platform default",
            "1. Edit and Save for my organization\n2. Reload",
            "Override badge/source; content persists for this employer only",
            None,
            "Candidate",
        ),
        (
            "Reset + restore",
            "Employer can reset to current platform or restore baseline/version",
            "P0",
            "Functional",
            "Employer",
            "Override exists; versions listed",
            "1. Current platform default\n2. Re-edit\n3. Restore Baseline from dropdown",
            "Both actions succeed; wording matches chosen system snapshot",
            None,
            "Candidate",
        ),
        (
            "Draft integration",
            "Guest confirmation draft uses employer override when present",
            "P1",
            "Integration",
            "Employer",
            "Override set for campus_guest_confirmation; published listing available",
            "1. Open confirmation draft for a listing\n2. Compare subject/body to template",
            "Draft reflects employer override with listing placeholders filled",
            "confirmation-draft API",
            "Manual",
        ),
    ],
)

add(
    "04 Placeholders resolve",
    "04 Placeholders & resolution",
    [
        (
            "Apply",
            "Known {{placeholders}} are substituted at send/render time",
            "P0",
            "Functional",
            "System",
            "Template contains {{collegeName}} etc.",
            "1. Trigger render/send with known vars",
            "Placeholders replaced with values",
            "applyEmailTemplate",
            "Candidate",
        ),
        (
            "Unknown",
            "Unknown {{placeholders}} are stripped to empty",
            "P1",
            "Functional",
            "System",
            "Template includes {{notARealVar}}",
            "1. Render template",
            "Unknown token removed (empty), no raw braces left",
            None,
            "Candidate",
        ),
        (
            "Order",
            "Override wins over system when scope override exists",
            "P0",
            "Functional",
            "College Admin",
            "System and override differ",
            "1. loadResolvedEmailTemplate with college scope",
            "source=override; body from override table",
            None,
            "Candidate",
        ),
        (
            "Fallback",
            "Code fallback used if system_email_templates row missing",
            "P2",
            "Edge",
            "System",
            "Row deleted or missing in env (dev only)",
            "1. Resolve key with no DB row",
            "EMAIL_TEMPLATE_FALLBACKS content returned",
            "Do not run on prod",
            "Manual",
        ),
    ],
)

add(
    "05 Message templates",
    "05 College custom message templates",
    [
        (
            "CRUD",
            "College can create, edit, and delete custom message templates",
            "P0",
            "Functional",
            "College Admin",
            "On /dashboard/college/message-templates",
            "1. Add template\n2. Edit from table\n3. Delete",
            "Row appears/updates/removes; API success toasts",
            "Separate from system email templates",
            "Candidate",
        ),
        (
            "Preview table",
            "Table eye action opens Preview with sample data",
            "P0",
            "Functional",
            "College Admin",
            "Template with {{studentName}} in body",
            "1. Click preview (eye) in Actions",
            "Modal shows filled subject/body and sample var chips; nothing emailed",
            None,
            "Candidate",
        ),
        (
            "Preview edit",
            "Edit form Preview with sample data uses current unsaved text",
            "P0",
            "Functional",
            "College Admin",
            "Form open with placeholders",
            "1. Change body without saving\n2. Preview with sample data",
            "Preview reflects unsaved form content with demo values",
            None,
            "Candidate",
        ),
        (
            "Types",
            "Template type email / notification / sms can be selected",
            "P1",
            "Functional",
            "College Admin",
            "New template form",
            "1. Create one of each type\n2. Filter/search",
            "Types stored and shown as badges",
            None,
            "Manual",
        ),
    ],
)

add(
    "06 Delivery logs",
    "06 Delivery visibility",
    [
        (
            "Logs",
            "Successful templated send appears in admin email logs",
            "P1",
            "Integration",
            "Super Admin",
            "Zepto/SMTP configured; trigger verification or sponsorship mail",
            "1. Trigger email\n2. Open /dashboard/admin/email-logs\n3. Search context/subject",
            "Row with status sent/failed; recipient trail present",
            "Optional Zepto request id column after migration 115",
            "Manual",
        ),
        (
            "No export mail",
            "Table CSV Export does not create mail_delivery_logs rows",
            "P2",
            "Regression",
            "College Admin",
            "On students list",
            "1. Export CSV full\n2. Check email logs",
            "No new export-related mail log (browser download only)",
            "Confirms product decision",
            "Manual",
        ),
    ],
)


def style_header(ws, row: int = 2):
    fill = PatternFill("solid", fgColor="1E3A8A")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws):
    widths = {
        1: 16,
        2: 28,
        3: 16,
        4: 42,
        5: 10,
        6: 12,
        7: 18,
        8: 28,
        9: 40,
        10: 40,
        11: 28,
        12: 12,
        13: 10,
        14: 24,
        15: 18,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{ws.max_row}"


def main():
    wb = Workbook()
    idx = wb.active
    idx.title = "Index"
    title_font = Font(name="Arial", bold=True, size=16, color="1E3A8A")
    head_font = Font(name="Arial", bold=True, size=11)
    body = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="2563EB", underline="single")
    idx["A1"] = "PlacementHub — Email Templates Test Cases"
    idx["A1"].font = title_font
    idx["A2"] = (
        "Covers platform system email templates, college/employer overrides, system baseline + version restore, "
        "placeholders, custom message-template preview, and delivery-log checks."
    )
    idx["A2"].font = body
    idx["A2"].alignment = Alignment(wrap_text=True)
    idx.merge_cells("A2:G2")
    idx.row_dimensions[2].height = 40

    headers_idx = ["#", "Module / Tab", "Case Count", "P0", "P1", "P2+", "Open"]
    for c, h in enumerate(headers_idx, 1):
        cell = idx.cell(4, c, h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")

    total = p0 = p1 = p2 = 0
    for i, (sheet, _module, rows) in enumerate(CASES, 1):
        counts = {"P0": 0, "P1": 0, "P2": 0}
        for r in rows:
            pr = r[2]
            if pr.startswith("P0"):
                counts["P0"] += 1
            elif pr.startswith("P1"):
                counts["P1"] += 1
            else:
                counts["P2"] += 1
        n = len(rows)
        total += n
        p0 += counts["P0"]
        p1 += counts["P1"]
        p2 += counts["P2"]
        row = 4 + i
        idx.cell(row, 1, i).font = body
        idx.cell(row, 2, sheet).font = body
        idx.cell(row, 3, n).font = body
        idx.cell(row, 4, counts["P0"]).font = body
        idx.cell(row, 5, counts["P1"]).font = body
        idx.cell(row, 6, counts["P2"]).font = body
        link = idx.cell(row, 7, "Open ›")
        link.font = link_font
        link.hyperlink = f"#'{sheet}'!A1"

    sum_row = 5 + len(CASES)
    idx.cell(sum_row, 2, "TOTAL").font = head_font
    idx.cell(sum_row, 3, total).font = head_font
    idx.cell(sum_row, 4, p0).font = head_font
    idx.cell(sum_row, 5, p1).font = head_font
    idx.cell(sum_row, 6, p2).font = head_font

    idx.cell(sum_row + 2, 1, "How to use").font = head_font
    idx.cell(
        sum_row + 3,
        1,
        "1) Apply db:migrate:116 before version-restore cases. "
        "2) Execute by module tab. 3) Fill Status / Actual / Executed At. "
        "4) P0 blocker, P1 important, P2 polish.",
    ).font = body
    idx.merge_cells(start_row=sum_row + 3, start_column=1, end_row=sum_row + 3, end_column=7)
    idx.row_dimensions[sum_row + 3].height = 48

    for c, w in enumerate([6, 28, 12, 8, 8, 8, 10], 1):
        idx.column_dimensions[get_column_letter(c)].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    font = Font(name="Arial", size=10)
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for sheet, module, rows in CASES:
        ws = wb.create_sheet(sheet)
        back = ws.cell(1, 1, "« Back to Index")
        back.font = link_font
        back.hyperlink = "#Index!A1"
        style_header(ws, 2)
        prefix = sheet.split(" ", 1)[0]
        for i, r in enumerate(rows, 1):
            feature, title, priority, typ, roles, pre, steps, expected, notes, auto = r
            values = [
                f"TC-ET-{prefix}-{i:03d}",
                module,
                feature,
                title,
                priority,
                typ,
                roles,
                pre,
                steps,
                expected,
                notes,
                auto,
                "",
                "",
                "",
            ]
            row_idx = 2 + i
            for col, val in enumerate(values, 1):
                cell = ws.cell(row_idx, col, val)
                cell.font = font
                cell.alignment = wrap
                cell.border = thin
            ws.row_dimensions[row_idx].height = 72
        autosize(ws)

    wb.save(OUT)
    print(f"Wrote {OUT} with {total} cases ({p0} P0 / {p1} P1 / {p2} P2+)")


if __name__ == "__main__":
    main()
