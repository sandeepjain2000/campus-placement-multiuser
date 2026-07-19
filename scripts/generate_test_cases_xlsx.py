#!/usr/bin/env python3
"""Generate PlacementHub exhaustive QA test-case workbook (module tabs)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parents[1] / "docs" / "PlacementHub-Test-Cases.xlsx"

HEADERS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
]

# (feature, title, priority, type, roles, preconditions, steps, expected, notes, automation)
# Priority: P0 critical, P1 high, P2 medium, P3 low
# Type: Functional | Negative | Edge | Security | UI | Regression | Integration | Performance
# Automation: Manual | Candidate | Automated

MODULES = {}


def add(module, *rows):
    MODULES.setdefault(module, []).extend(rows)


# ---------------------------------------------------------------------------
# Auth & Access
# ---------------------------------------------------------------------------
add(
    "01 Auth & Access",
    ("Sign-in", "Valid student credentials land on student dashboard", "P0", "Functional", "Student", "Provisioned student account", "1. Open /sign-in\n2. Enter valid email/password\n3. Complete captcha if shown\n4. Submit", "Redirect to student dashboard; session cookie set; role menus visible", "Seed student@placementhub.test", "Candidate"),
    ("Sign-in", "Valid employer credentials land on employer dashboard", "P0", "Functional", "Employer", "Approved employer account", "1. Sign in as employer\n2. Observe home", "Employer dashboard and partnership-scoped menus load", "", "Candidate"),
    ("Sign-in", "Valid college admin credentials land on college dashboard", "P0", "Functional", "College Admin", "College admin account for tenant", "1. Sign in as college admin", "College dashboard loads; tenant branding/name shown", "", "Candidate"),
    ("Sign-in", "Valid super admin credentials land on admin dashboard", "P0", "Functional", "Super Admin", "Super admin provisioned", "1. Sign in as super admin", "Admin dashboard and platform menus load", "", "Candidate"),
    ("Sign-in", "Invalid password shows error and does not create session", "P0", "Negative", "All", "Existing account", "1. Enter correct email, wrong password\n2. Submit", "Error message; remain on sign-in; no dashboard access", "", "Candidate"),
    ("Sign-in", "Unknown email rejected", "P1", "Negative", "All", "None", "1. Enter non-existent email\n2. Submit", "Generic auth failure; no user enumeration leak preferred", "", "Manual"),
    ("Sign-in", "Empty email/password blocked by validation", "P1", "Negative", "All", "None", "1. Submit blank form", "Client/server validation; no request or 400", "", "Candidate"),
    ("Sign-in", "Captcha required when enabled", "P1", "Security", "All", "Captcha configured", "1. Attempt sign-in without captcha token", "Blocked with verification message", "", "Manual"),
    ("Sign-in", "Account pending approval cannot access dashboard", "P0", "Negative", "Employer, College Admin", "Registration pending super-admin approval", "1. Sign in with pending account", "Denied or pending message; no protected routes", "", "Manual"),
    ("Sign-out", "Sign out clears session and returns to public area", "P0", "Functional", "All", "Signed in", "1. Click Sign out", "Session ended; protected /dashboard redirects to sign-in", "", "Candidate"),
    ("Password reset", "Request reset email for valid account", "P1", "Functional", "All", "SMTP configured", "1. Open forgot password\n2. Enter registered email\n3. Submit", "Success acknowledgement; email logged; reset link works", "Check Email logs / yopmail in sandbox", "Manual"),
    ("Password reset", "Reset with expired/invalid token fails", "P1", "Negative", "All", "Expired token", "1. Open reset URL with bad token\n2. Submit new password", "Error; password unchanged", "", "Manual"),
    ("Password reset", "Successful reset allows new password login", "P0", "Functional", "All", "Valid reset token", "1. Set new password\n2. Sign in with new password", "Login succeeds; old password fails", "", "Manual"),
    ("Registration", "Employer can register via /register", "P0", "Functional", "Employer", "Captcha available", "1. Open /register\n2. Complete employer form\n3. Submit", "Registration created; pending or approved per config; confirmation shown", "", "Manual"),
    ("Registration", "College can register via /register", "P0", "Functional", "College Admin", "Captcha available", "1. Register as college", "Tenant/registration row created; pending approval if required", "", "Manual"),
    ("Registration", "Student cannot self-register on public register", "P0", "Security", "Student", "None", "1. Inspect /register for student self-signup", "No public student self-registration path; college provisioning only", "Product rule Jun 2026", "Manual"),
    ("Registration", "Duplicate email rejected", "P1", "Negative", "Employer, College Admin", "Email already exists", "1. Register with existing email", "Clear error; no duplicate user", "", "Candidate"),
    ("Session", "Unauthenticated /dashboard request redirects to sign-in", "P0", "Security", "All", "Logged out", "1. Open /dashboard/student directly", "Redirect to sign-in with return awareness", "", "Candidate"),
    ("Session", "Role cannot open another role's dashboard routes", "P0", "Security", "Student", "Student session", "1. Open /dashboard/admin and /dashboard/employer URLs", "403/redirect; no data leak", "", "Candidate"),
    ("Session", "Browser back after sign-out does not resurrect session", "P1", "Security", "All", "Just signed out", "1. Sign out\n2. Browser Back", "Still unauthenticated or forced re-auth", "", "Manual"),
    ("Landing", "Landing page loads with Sign In and Register", "P1", "UI", "Public", "None", "1. Open /\n2. Click Sign In and Register", "Pages load; marketing links work", "", "Candidate"),
    ("Landing", "Sandbox banner visible on landing when enabled", "P2", "UI", "Public", "NEXT_PUBLIC_SANDBOX_BANNER not false", "1. Open /", "Sandbox environment line shown at top", "", "Manual"),
    ("Mobile nav", "Mobile Landing link from dashboard returns to /", "P2", "UI", "All", "Signed in; viewport ≤768px", "1. Open any dashboard screen\n2. Tap Landing beside menu", "Navigates to landing page", "Mobile only", "Manual"),
)

# ---------------------------------------------------------------------------
# Student
# ---------------------------------------------------------------------------
add(
    "02 Student",
    ("Dashboard", "Student home loads without fatal errors", "P0", "Functional", "Student", "Active student", "1. Sign in as student\n2. Open home", "Dashboard widgets load; no 500 toast", "", "Candidate"),
    ("Profile", "View My Profile with academic fields", "P0", "Functional", "Student", "Student profile exists", "1. Open My Profile", "Name, department, CGPA, skills visible/editable per rules", "", "Candidate"),
    ("Profile", "Update profile and save successfully", "P0", "Functional", "Student", "Editable fields allowed", "1. Change skills/phone\n2. Save", "Success toast; persisted on reload", "", "Candidate"),
    ("Profile", "Invalid CGPA rejected", "P1", "Negative", "Student", "Editable CGPA", "1. Enter CGPA > 10 or negative\n2. Save", "Validation error; no save", "", "Candidate"),
    ("Documents", "Upload resume/CV when feature available", "P0", "Functional", "Student", "CV management migrated", "1. Open My CVs / documents\n2. Upload PDF within size limit", "CV listed with label; default flag works", "", "Manual"),
    ("Documents", "Missing CV file shows missing-file state not crash", "P1", "Edge", "Student", "CV row without S3 object", "1. Open My CVs\n2. Try view missing file", "Friendly missing-file message; page still usable", "PH-CV-S3-MISSING", "Manual"),
    ("Documents", "CV list soft-failure still loads page", "P1", "Edge", "Student", "Force CV API failure in sandbox", "1. Open My CVs when API soft-fails", "Empty/unavailable state with Ref; rest of page works; Error log entry", "", "Manual"),
    ("Browse drives", "Browse placement drives shows approved visible drives", "P0", "Functional", "Student", "Approved drive for tenant", "1. Open Browse drives", "Drive cards/list; eligibility hints; apply CTA when eligible", "", "Candidate"),
    ("Browse drives", "Ineligible student blocked from apply (CGPA/backlogs)", "P0", "Negative", "Student", "Drive min CGPA above student", "1. Attempt Apply", "Blocked with clear eligibility reason", "", "Manual"),
    ("Browse jobs", "Browse jobs shows approved published jobs only", "P0", "Functional", "Student", "Approved job visibility", "1. Open Browse jobs", "Only approved listings for campus; pending not shown", "", "Candidate"),
    ("Browse internships", "Browse internships lists approved internships", "P0", "Functional", "Student", "Approved internship", "1. Open internships browse", "List loads; apply available before deadline", "", "Candidate"),
    ("Browse projects", "Browse short projects", "P1", "Functional", "Student", "Approved short project", "1. Open projects", "Listings visible; apply flow works", "", "Manual"),
    ("Browse hackathons", "Browse hackathons", "P1", "Functional", "Student", "Approved hackathon", "1. Open hackathons", "Listings visible; apply flow works", "", "Manual"),
    ("Apply", "Apply to drive before deadline succeeds", "P0", "Functional", "Student", "Eligible; not already applied", "1. Apply with required CV\n2. Confirm", "Application created; status Applied; alert/email as configured", "", "Candidate"),
    ("Apply", "Apply after deadline rejected", "P0", "Negative", "Student", "Past deadline drive", "1. Attempt apply", "Rejected with deadline message", "", "Manual"),
    ("Apply", "Duplicate apply prevented", "P1", "Negative", "Student", "Already applied", "1. Apply again", "Idempotent block; no duplicate rows", "", "Candidate"),
    ("Apply", "Withdraw application when allowed", "P1", "Functional", "Student", "Applied; withdrawal enabled", "1. Withdraw", "Status updated; employer sees withdrawn", "", "Manual"),
    ("My Applications", "Applications listed per type tabs", "P0", "Functional", "Student", "Mixed applications", "1. Open My Applications\n2. Switch type tabs", "Correct apps per tab; status badges accurate", "", "Candidate"),
    ("Interviews", "My Interviews shows scheduled interviews", "P0", "Functional", "Student", "Interview scheduled", "1. Open My Interviews", "Date/time/mode visible; no crash", "", "Manual"),
    ("Offers", "View offer and accept", "P0", "Functional", "Student", "Pending offer", "1. Open My Offers\n2. Accept", "Offer accepted; rules enforced (limits)", "", "Manual"),
    ("Offers", "Decline offer", "P1", "Functional", "Student", "Pending offer", "1. Decline", "Offer declined; confirmation", "", "Manual"),
    ("Offers", "Offer letter view when attached", "P1", "Functional", "Student", "Offer with letter file", "1. Open offer letter", "File opens or download works; unavailable state if missing", "", "Manual"),
    ("Calendar", "Placement calendar shows relevant events", "P1", "Functional", "Student", "Campus events exist", "1. Open calendar", "Events render; navigation works", "", "Manual"),
    ("Mentorship", "Student can post mentorship request", "P1", "Functional", "Student", "Feature enabled", "1. Create mentorship need\n2. Submit", "Request pending college approval", "", "Manual"),
    ("Mentorship", "Student sees approval/volunteer status", "P2", "Functional", "Student", "Request approved; employer volunteered", "1. Open mentorship requests", "Status reflects college + volunteers", "", "Manual"),
    ("Clarifications", "Student can read published clarifications", "P1", "Functional", "Student", "College published batch", "1. Open clarifications", "Q&A visible read-only", "", "Manual"),
    ("Alerts", "Student receives and opens alerts", "P0", "Functional", "Student", "Notification exists", "1. Open Inbox & Alerts\n2. Open unread", "Marked read; unread count decreases", "", "Candidate"),
    ("Data export", "My data export downloads when supported", "P2", "Functional", "Student", "Export enabled", "1. Request data export", "File downloads or job queued; no 500", "", "Manual"),
    ("Alumni", "Alumni student sees alumni jobs not campus drives menu", "P0", "Functional", "Student (Alumni)", "isAlumni flag set", "1. Sign in as alumni\n2. Inspect nav", "Alumni browse/my jobs; campus drive browse hidden/replaced", "", "Manual"),
    ("Feedback", "Student can submit product feedback", "P2", "Functional", "Student", "None", "1. Open Feedback\n2. Submit", "Success; appears in admin inbox", "", "Manual"),
)

# ---------------------------------------------------------------------------
# Employer
# ---------------------------------------------------------------------------
add(
    "03 Employer",
    ("Dashboard", "Employer home loads", "P0", "Functional", "Employer", "Approved employer", "1. Sign in", "Dashboard loads; campus context shown", "", "Candidate"),
    ("Profile", "Update company profile", "P0", "Functional", "Employer", "Employer profile", "1. Edit company name/website\n2. Save", "Persisted; logo upload if supported", "", "Manual"),
    ("Partnerships", "Request campus partnership", "P0", "Functional", "Employer", "College visible in directory", "1. Campus Partnerships\n2. Request college", "Pending request created; college sees it", "", "Candidate"),
    ("Partnerships", "Cannot publish to campus without approval", "P0", "Security", "Employer", "No approval for campus A", "1. Attempt create drive targeting A", "Blocked or campus not selectable", "", "Manual"),
    ("Partnerships", "Approved campus available for publishing", "P0", "Functional", "Employer", "Approved partnership", "1. Create internship selecting approved campus", "Campus selectable; save succeeds", "", "Candidate"),
    ("Drives", "Request new placement drive", "P0", "Functional", "Employer", "Approved campus", "1. Create drive with required fields\n2. Submit", "Drive pending college approval", "", "Candidate"),
    ("Drives", "Edit draft/pending drive", "P1", "Functional", "Employer", "Own pending drive", "1. Edit title/dates\n2. Save", "Updates persisted", "", "Manual"),
    ("Drives", "Cancel drive when allowed", "P1", "Functional", "Employer", "Cancellable drive", "1. Cancel", "Status cancelled; students cannot apply", "", "Manual"),
    ("Internships", "Publish internship to campuses", "P0", "Functional", "Employer", "Approved campuses", "1. Create internship\n2. Select campuses\n3. Publish", "Visibility rows pending/approved per campus", "", "Candidate"),
    ("Jobs", "Publish job listing", "P0", "Functional", "Employer", "Approved campuses", "1. Create job\n2. Publish", "Job created with campus visibility", "", "Manual"),
    ("Projects", "Publish short project", "P1", "Functional", "Employer", "Approved campus", "1. Create short project", "Listing created", "", "Manual"),
    ("Hackathons", "Publish hackathon", "P1", "Functional", "Employer", "Approved campus", "1. Create hackathon", "Listing created", "", "Manual"),
    ("Applications", "View applications list with filters", "P0", "Functional", "Employer", "Applications exist", "1. Open Applications\n2. Filter by status/campus", "Correct filtered rows", "", "Candidate"),
    ("Applications", "Update application status", "P0", "Functional", "Employer", "Applied candidate", "1. Change status", "Status saved; student sees update where exposed", "", "Manual"),
    ("Assessments CSV", "Export assessment CSV for campus target", "P0", "Functional", "Employer", "Campus + drive/job", "1. Assessment uploads\n2. Export CSV", "CSV downloads with expected headers", "", "Manual"),
    ("Assessments CSV", "Upload valid hiring_result CSV commits rows", "P0", "Functional", "Employer", "Valid export filled", "1. Upload CSV\n2. Confirm commit", "Results stored; dashboard reflects Shortlist/Reject/Select", "hiring_result values", "Manual"),
    ("Assessments CSV", "Invalid CSV rows show import review errors", "P0", "Negative", "Employer", "Malformed CSV", "1. Upload bad CSV", "Row-level errors; no partial corrupt commit without review", "", "Manual"),
    ("Assessments CSV", "Submit results locks further edits", "P0", "Functional", "Employer", "Committed results", "1. Submit results", "Locked; further upload blocked or read-only", "", "Manual"),
    ("Assessments Online", "Assessment Update Online edits hiring_result", "P0", "Functional", "Employer", "Online assessment screen", "1. Edit result inline\n2. Save", "Same data model as CSV path", "", "Manual"),
    ("Assessments", "FCFS unavailable students listed when enabled", "P1", "Functional", "Employer", "College FCFS on; student claimed", "1. Open FCFS unavailable", "Claimed students listed; cannot select them", "", "Manual"),
    ("Hiring Results", "Hiring Results Dashboard read view", "P1", "Functional", "Employer", "Results exist", "1. Open dashboard", "Aggregated results visible", "", "Manual"),
    ("Interviews", "Schedule interview for candidate", "P0", "Functional", "Employer", "Shortlisted student", "1. Schedule interview\n2. Save", "Student sees interview; calendar entry", "", "Manual"),
    ("Interviews", "Reschedule / cancel interview", "P1", "Functional", "Employer", "Existing interview", "1. Reschedule then cancel", "Updates propagate to student", "", "Manual"),
    ("Offers", "Create offer for selected student", "P0", "Functional", "Employer", "Selected student; partnership ok", "1. Create offer\n2. Submit", "Offer visible to student/college; rules checked", "", "Manual"),
    ("Offers", "Upload offers CSV", "P1", "Functional", "Employer", "Valid offers CSV", "1. Upload from Offers page", "Offers created or validation errors listed", "", "Manual"),
    ("Offers", "Attach offer letter file", "P1", "Functional", "Employer", "Offer exists", "1. Attach PDF", "Student can open letter", "", "Manual"),
    ("Alumni jobs", "Publish alumni job", "P1", "Functional", "Employer", "Alumni feature", "1. Create alumni job", "Visible to alumni students", "", "Manual"),
    ("Alumni assessment", "Alumni Job Assessment Online updates", "P2", "Functional", "Employer", "Alumni applications", "1. Update results online", "Statuses saved", "", "Manual"),
    ("Mentorship", "Volunteer on approved mentorship request", "P1", "Functional", "Employer", "Approved request; partnership", "1. Volunteer", "Student/college see volunteer", "", "Manual"),
    ("Guest needs", "Respond to campus guest lecture need", "P2", "Functional", "Employer", "College listed need", "1. Confirm/respond", "Status updates", "", "Manual"),
    ("Sponsorships", "Create sponsorship offer", "P2", "Functional", "Employer", "Feature enabled", "1. Submit sponsorship", "College can view", "", "Manual"),
    ("Seed funding", "Submit startup seed funding interest", "P2", "Functional", "Employer", "Feature enabled", "1. Submit", "Tracked for college review", "", "Manual"),
    ("Calendar", "Employer calendar shows events", "P1", "Functional", "Employer", "Events exist", "1. Open calendar", "Events load", "", "Manual"),
    ("Clarifications", "Employer participates in clarification threads", "P1", "Functional", "Employer", "College published", "1. Open clarifications\n2. Respond if allowed", "Responses saved", "", "Manual"),
    ("Templates", "Email/message templates CRUD", "P2", "Functional", "Employer", "Templates feature", "1. Create template\n2. Edit\n3. Delete", "CRUD works", "", "Manual"),
    ("Alerts", "Employer alerts inbox works", "P1", "Functional", "Employer", "Notifications", "1. Open Alerts", "List, star, trash work", "", "Manual"),
    ("Marketplace", "Marketplace orders/browse if enabled", "P2", "Functional", "Employer", "Marketplace on", "1. Open marketplace", "Pages load without 500", "", "Manual"),
)

# ---------------------------------------------------------------------------
# College Admin
# ---------------------------------------------------------------------------
add(
    "04 College Admin",
    ("Dashboard", "College home loads with tenant context", "P0", "Functional", "College Admin", "College admin session", "1. Sign in", "Tenant name/logo; academic year selector works", "", "Candidate"),
    ("Settings", "Update college settings / season label", "P0", "Functional", "College Admin", "Settings access", "1. Change setting\n2. Save", "Persisted; affects season scoping", "", "Manual"),
    ("Academic year", "Switch active academic year", "P0", "Functional", "College Admin", "Multiple years", "1. Change year in topbar", "Lists filter to selected year", "", "Manual"),
    ("Students list", "Master student list loads with search/filter", "P0", "Functional", "College Admin", "Students exist", "1. Open Students\n2. Search by name/email", "Correct rows; pagination if any", "", "Candidate"),
    ("Students import", "Download students CSV import template", "P0", "Functional", "College Admin", "None", "1. Download template", "Headers match export; example row present", "Remarks optional; Photo URL blank ok", "Candidate"),
    ("Students import", "Import valid students CSV", "P0", "Functional", "College Admin", "Valid template filled", "1. Import CSV", "Students created/updated; success count", "", "Manual"),
    ("Students import", "Import rejects blank required CGPA", "P0", "Negative", "College Admin", "CSV with blank CGPA", "1. Import", "Row error; no partial bad row", "collegeStudentsCsv tests", "Candidate"),
    ("Students import", "Import rejects blank Department", "P0", "Negative", "College Admin", "CSV blank dept", "1. Import", "Row rejected", "", "Candidate"),
    ("Students import", "Blank Remarks and Photo URL allowed", "P1", "Edge", "College Admin", "CSV with blanks", "1. Import", "Accepted; photo not imported when blank", "", "Candidate"),
    ("Students", "Add single student", "P0", "Functional", "College Admin", "None", "1. Add student form\n2. Save", "Student created; login email sent if SMTP on", "", "Manual"),
    ("Students", "Verify / unverify student profile", "P1", "Functional", "College Admin", "Unverified student", "1. Mark verified", "Flag persisted; apply gates respect setting", "", "Manual"),
    ("Students", "View student CVs panel", "P1", "Functional", "College Admin", "Student with CVs", "1. Open student detail CVs", "List loads; verify toggle if permitted", "", "Manual"),
    ("Students", "College CV list failure soft-empty + error log", "P1", "Edge", "College Admin", "Induce CV list failure", "1. Open student CVs", "Friendly empty; Ref; platform_error_logs row", "", "Manual"),
    ("Enrollment key", "Rotate enrollment key if exposed", "P2", "Functional", "College Admin", "Legacy key UI", "1. Rotate key", "New key stored; old invalid", "Legacy", "Manual"),
    ("Employers", "Approve employer partnership request", "P0", "Functional", "College Admin", "Pending request", "1. Approve", "Employer can publish to campus", "", "Candidate"),
    ("Employers", "Reject employer partnership request", "P0", "Functional", "College Admin", "Pending request", "1. Reject", "Employer remains blocked for campus", "", "Candidate"),
    ("Employers", "Employer directory lists partners", "P1", "Functional", "College Admin", "Approved employers", "1. Open employers", "Directory accurate", "", "Manual"),
    ("Drives", "Approve placement drive", "P0", "Functional", "College Admin", "Pending drive", "1. Approve drive", "Students can see/apply when published", "", "Candidate"),
    ("Drives", "Reject placement drive", "P1", "Functional", "College Admin", "Pending drive", "1. Reject", "Not visible to students", "", "Manual"),
    ("Visibility", "Approve job/internship visibility", "P0", "Functional", "College Admin", "Pending visibility", "1. Approve listing", "Students see listing", "", "Candidate"),
    ("Visibility", "Reject job/internship visibility", "P1", "Functional", "College Admin", "Pending visibility", "1. Reject", "Hidden from students", "", "Manual"),
    ("Applications", "Campus applications board monitors pipeline", "P0", "Functional", "College Admin", "Applications exist", "1. Open Applications", "Filters/status work; export if present", "", "Manual"),
    ("Offers", "Monitor offers board", "P0", "Functional", "College Admin", "Offers exist", "1. Open Offers", "Offers listed; statuses correct", "", "Manual"),
    ("Hiring Assessment", "Read-only hiring assessment mirror", "P1", "Functional", "College Admin", "Employer submitted results", "1. Open Hiring Assessment", "Read-only; matches employer results", "", "Manual"),
    ("Interviews", "College interview scheduling coordination", "P1", "Functional", "College Admin", "Permission", "1. Schedule/view interviews", "Works without overwriting employer data incorrectly", "", "Manual"),
    ("Placement rules", "Configure CGPA minimum", "P0", "Functional", "College Admin", "Rules UI", "1. Set min CGPA\n2. Save", "Apply blocked for lower CGPA students", "", "Manual"),
    ("Placement rules", "Configure offer limits", "P0", "Functional", "College Admin", "Rules UI", "1. Set max offers\n2. Student with limit tries accept", "Blocked per rule", "", "Manual"),
    ("Placement rules", "Enable FCFS", "P0", "Functional", "College Admin", "Rules UI", "1. Enable FCFS\n2. Two employers select same student track", "First wins; second blocked; unavailable list updates", "", "Manual"),
    ("Placement rules", "Season date windows enforce", "P1", "Functional", "College Admin", "Season dates set", "1. Act outside window", "Blocked or warned per design", "", "Manual"),
    ("Reports", "Generate/export placement reports", "P1", "Functional", "College Admin", "Season data", "1. Open Reports\n2. Export", "File downloads; numbers plausible", "", "Manual"),
    ("Audit", "Audit reports load log entries", "P0", "Functional", "College Admin", "Audit data", "1. Open Audit Reports", "Logs load; soft-fail shows Ref + Error log if fails", "log-entries path", "Candidate"),
    ("Audit", "Audit export downloads", "P1", "Functional", "College Admin", "Logs exist", "1. Export", "File generated", "", "Manual"),
    ("Clarifications", "Publish clarification batch", "P1", "Functional", "College Admin", "None", "1. Create Q&A batch\n2. Publish", "Employers/students see content", "", "Manual"),
    ("Mentorship", "Approve student mentorship request", "P1", "Functional", "College Admin", "Pending request", "1. Approve", "Visible to partnered employers", "", "Manual"),
    ("Mentorship", "Reject mentorship request", "P2", "Functional", "College Admin", "Pending request", "1. Reject", "Not visible to employers", "", "Manual"),
    ("Calendar/Events", "Create campus calendar event", "P1", "Functional", "College Admin", "None", "1. Create event", "Visible on calendars", "", "Manual"),
    ("Guest faculty", "List campus guest needs", "P2", "Functional", "College Admin", "None", "1. Create guest need", "Employers can respond", "", "Manual"),
    ("Sponsorships", "View employer sponsorships", "P2", "Functional", "College Admin", "Sponsorship exists", "1. Open sponsorships", "List loads", "", "Manual"),
    ("Logo", "Upload college logo", "P1", "Functional", "College Admin", "Settings", "1. Upload logo", "Topbar/sidebar show logo", "", "Manual"),
    ("Committee", "Placement committee role sees scoped college menus", "P1", "Security", "Placement Committee", "Committee account", "1. Sign in", "College ops menus; no super-admin", "", "Manual"),
)

# ---------------------------------------------------------------------------
# Super Admin
# ---------------------------------------------------------------------------
add(
    "05 Super Admin",
    ("Dashboard", "Admin dashboard loads", "P0", "Functional", "Super Admin", "Super admin", "1. Sign in", "Platform overview loads", "", "Candidate"),
    ("Colleges", "List colleges", "P0", "Functional", "Super Admin", "Tenants exist", "1. Open Colleges", "Directory loads; search works", "", "Candidate"),
    ("Colleges", "Add college", "P0", "Functional", "Super Admin", "None", "1. Add college form\n2. Save", "Tenant created", "", "Manual"),
    ("Colleges", "View/edit college profile", "P1", "Functional", "Super Admin", "College exists", "1. Open profile\n2. Edit\n3. Save", "Changes persisted", "", "Manual"),
    ("Employers", "List employers", "P0", "Functional", "Super Admin", "Employers exist", "1. Open Employers", "List loads", "", "Candidate"),
    ("Users", "Manage users list/search", "P0", "Functional", "Super Admin", "Users exist", "1. Open Manage Users\n2. Search", "Users found; roles shown", "", "Candidate"),
    ("Users", "Update user role/status carefully", "P1", "Functional", "Super Admin", "Target user", "1. Change allowed field\n2. Save", "Persisted; user can/cannot access per role", "", "Manual"),
    ("Onboarding", "Approve pending college registration", "P0", "Functional", "Super Admin", "Pending college", "1. Approve", "College can sign in", "", "Manual"),
    ("Onboarding", "Approve pending employer registration", "P0", "Functional", "Super Admin", "Pending employer", "1. Approve", "Employer can sign in", "", "Manual"),
    ("Onboarding", "Reject pending registration", "P1", "Functional", "Super Admin", "Pending registration", "1. Reject", "Cannot access dashboard", "", "Manual"),
    ("Listings", "Admin placement listings overview", "P1", "Functional", "Super Admin", "Listings exist", "1. Open listings", "Cross-tenant view loads", "", "Manual"),
    ("Archived students", "View archived students", "P2", "Functional", "Super Admin", "Archived exist", "1. Open archived", "List loads", "", "Manual"),
    ("Error logs", "Platform error logs list and detail", "P0", "Functional", "Super Admin", "Errors logged", "1. Open Error logs\n2. Open row", "Ref, context, stack/details visible; System Defined Unique Code shown", "", "Manual"),
    ("Error logs", "Soft API failures appear in error logs", "P0", "Functional", "Super Admin", "Trigger soft-fail API", "1. Cause soft failure\n2. Check Error logs", "Entry with Ref and technical details", "api_soft_failure / cv_soft_failure", "Manual"),
    ("Email logs", "Email logs list delivery attempts", "P1", "Functional", "Super Admin", "Emails sent", "1. Open Email logs", "Entries with status", "", "Manual"),
    ("Audit", "Super-admin audit reports", "P1", "Functional", "Super Admin", "Audit events", "1. Open Audit", "Logs/filter/export", "", "Manual"),
    ("Feedback", "Feedback inbox lists submissions", "P1", "Functional", "Super Admin", "Feedback exists", "1. Open Feedback inbox", "Items listed", "", "Manual"),
    ("Settings", "Platform settings SMTP/branding save", "P0", "Functional", "Super Admin", "Settings access", "1. Update setting\n2. Save", "Persisted; affects outbound mail/branding", "", "Manual"),
    ("Isolation", "Super admin actions audit-logged", "P1", "Security", "Super Admin", "Audit on", "1. Perform sensitive update", "Audit row created", "", "Manual"),
)

# ---------------------------------------------------------------------------
# Shared / Cross-cutting modules
# ---------------------------------------------------------------------------
add(
    "06 Partnerships",
    ("Lifecycle", "Request → Pending → Approved end-to-end", "P0", "Integration", "Employer, College Admin", "Both accounts", "1. Employer requests\n2. College approves\n3. Employer publishes", "Full path succeeds", "", "Manual"),
    ("Lifecycle", "Request → Rejected blocks publishing", "P0", "Integration", "Employer, College Admin", "Both accounts", "1. Request\n2. Reject\n3. Employer tries publish", "Blocked", "", "Manual"),
    ("Directory", "Employer sees all colleges for requests", "P1", "Functional", "Employer", "Multiple tenants", "1. Open partnerships directory", "All colleges listed for request", "", "Manual"),
    ("Scope", "API denies cross-campus data without approval", "P0", "Security", "Employer", "Token for employer", "1. Call sensitive API for unapproved tenant", "403/empty", "", "Candidate"),
    ("Multi-campus", "Employer with multiple approvals works across campuses", "P1", "Functional", "Employer", "≥2 approvals", "1. Publish selecting two campuses", "Visibility per campus independent", "Campus switcher disabled; all campuses scope", "Manual"),
)

add(
    "07 Drives",
    ("Create", "Required fields validation on drive create", "P1", "Negative", "Employer", "Approved campus", "1. Submit empty required fields", "Validation errors", "", "Candidate"),
    ("Lifecycle", "Scheduled → In progress → Completed transitions", "P1", "Functional", "Employer, College Admin", "Drive exists", "1. Advance statuses per UI", "Status badges update; students see correct state", "", "Manual"),
    ("Venue", "Venue warning when date near and venue unconfirmed", "P2", "UI", "Employer, College Admin", "Drive soon; venue empty", "1. Open drive", "Warning surfaced", "", "Manual"),
    ("Eligibility", "Branch/dept display without false apply block until taxonomy", "P2", "Edge", "Student", "Drive with branch display", "1. View drive", "Branch shown; apply uses implemented eligibility only", "Product note", "Manual"),
    ("Apply window", "Students cannot apply to cancelled drive", "P0", "Negative", "Student", "Cancelled drive", "1. Attempt apply", "Blocked", "", "Manual"),
)

add(
    "08 Jobs Internships Projects",
    ("Visibility", "Pending visibility not shown to students", "P0", "Security", "Student", "Pending job", "1. Browse jobs", "Job absent", "", "Candidate"),
    ("Visibility", "Rejected visibility not shown to students", "P0", "Security", "Student", "Rejected job", "1. Browse", "Job absent", "", "Manual"),
    ("Edit", "Employer edits published listing", "P1", "Functional", "Employer", "Own listing", "1. Edit description\n2. Save", "Updates visible after college rules", "", "Manual"),
    ("Deadline", "Internship application closed after deadline", "P0", "Negative", "Student", "Past deadline", "1. Apply", "Blocked", "", "Manual"),
    ("Campus multi", "Same listing different campus visibility states", "P1", "Edge", "Student A/B", "Approved on A, pending on B", "1. Student A browses\n2. Student B browses", "Only A sees listing", "", "Manual"),
)

add(
    "09 Applications Pipeline",
    ("Status", "Applied → Shortlisted → Selected happy path", "P0", "Integration", "Student, Employer", "Eligible student", "1. Apply\n2. Employer shortlist\n3. Select", "Statuses sync both sides", "", "Manual"),
    ("Status", "Reject application", "P1", "Functional", "Employer", "Applied", "1. Reject", "Student sees rejected where exposed", "", "Manual"),
    ("Filters", "Filter by type/status/date", "P1", "Functional", "Employer, College Admin", "Mixed data", "1. Apply filters", "Result set correct", "", "Manual"),
    ("Export", "Export applications CSV", "P1", "Functional", "Employer, College Admin", "Apps exist", "1. Export", "CSV opens; columns sensible", "", "Manual"),
    ("Concurrency", "Two tabs updating same application", "P2", "Edge", "Employer", "Same app", "1. Update in two browsers", "Last write wins or conflict handled; no crash", "", "Manual"),
)

add(
    "10 Assessments",
    ("CSV", "Unknown hiring_result value rejected", "P0", "Negative", "Employer", "CSV with bad enum", "1. Upload", "Validation error", "Shortlist,Reject,Select,Decline,Withdraw", "Manual"),
    ("CSV", "Student ID mismatch handled", "P1", "Negative", "Employer", "Wrong student id", "1. Upload", "Row error", "", "Manual"),
    ("Online", "Decline and Withdraw results", "P1", "Functional", "Employer", "Online screen", "1. Set Decline\n2. Set Withdraw", "Stored correctly", "", "Manual"),
    ("Lock", "Post-submit cannot change results", "P0", "Negative", "Employer", "Submitted", "1. Try edit/upload", "Blocked", "", "Manual"),
    ("College mirror", "College view matches employer after submit", "P1", "Integration", "College Admin, Employer", "Submitted", "1. Compare both screens", "Parity", "", "Manual"),
    ("No email", "Assessment update does not email by default", "P2", "Functional", "Employer", "SMTP on", "1. Update result\n2. Check email logs", "No student email unless configured otherwise", "Product doc", "Manual"),
)

add(
    "11 Interviews",
    ("Create", "Required datetime validation", "P1", "Negative", "Employer", "Candidate", "1. Save without time", "Validation error", "", "Manual"),
    ("Notify", "Student alert on interview scheduled", "P1", "Functional", "Student, Employer", "Alerts on", "1. Schedule\n2. Check student alerts", "Alert present", "", "Manual"),
    ("Timezone", "Interview time displays consistently", "P2", "Edge", "Student, Employer", "Different TZ browsers", "1. Compare displayed times", "Consistent institution policy", "", "Manual"),
    ("Past", "Scheduling in the past warned or blocked", "P2", "Negative", "Employer", "None", "1. Pick past datetime", "Warning/block", "", "Manual"),
)

add(
    "12 Offers",
    ("Create", "Offer requires package/role fields per form", "P1", "Negative", "Employer", "Selected student", "1. Submit incomplete", "Validation", "", "Manual"),
    ("Rules", "College offer limit blocks extra accept", "P0", "Integration", "Student, College Admin", "Limit=1; one accepted", "1. Accept second offer", "Blocked", "", "Manual"),
    ("Letter", "Missing offer letter file graceful", "P1", "Edge", "Student", "Broken file_url", "1. Open letter", "Unavailable message; no 500", "", "Manual"),
    ("CSV", "Offers CSV duplicate rows handled", "P2", "Negative", "Employer", "Dup CSV", "1. Upload", "Errors or upsert rules clear", "", "Manual"),
)

add(
    "13 Placement Rules FCFS",
    ("FCFS", "First select locks student on track", "P0", "Integration", "Employer A/B, College", "FCFS enabled", "1. A selects\n2. B tries select same track", "B blocked; unavailable list", "", "Manual"),
    ("FCFS", "Different tracks independent", "P1", "Edge", "Employer", "FCFS on", "1. Select internship\n2. Other employer selects placement track", "Allowed if tracks differ per rules", "", "Manual"),
    ("CGPA", "Apply-time CGPA enforcement", "P0", "Functional", "Student", "Min CGPA set", "1. Low CGPA apply", "Blocked", "", "Manual"),
    ("Backlogs", "Backlog rule enforcement", "P1", "Functional", "Student", "Max backlogs=0; student has backlog", "1. Apply", "Blocked", "", "Manual"),
    ("Season", "Postings outside season handled", "P2", "Functional", "Employer, College", "Season closed", "1. Publish/apply", "Per configured behavior", "", "Manual"),
)

add(
    "14 Alerts Notifications",
    ("Inbox", "Unread and total shown at top of list", "P1", "UI", "All", "Alerts exist", "1. Open Inbox & Alerts", "Summary shows unread · total", "", "Manual"),
    ("Inbox", "Starred shows total beside Starred", "P1", "UI", "All", "Starred alerts", "1. View nav", "Count beside Starred label", "", "Manual"),
    ("Inbox", "Preview snippet max 50 characters", "P1", "UI", "All", "Long message", "1. View list row", "Preview ≤50 chars + ellipsis; full on open", "", "Manual"),
    ("Inbox", "Delete and date visible without horizontal scroll", "P1", "UI", "All", "Mobile/desktop", "1. View list", "Actions/time visible without scroll", "", "Manual"),
    ("Inbox", "Star / unstar persists", "P1", "Functional", "All", "Alert", "1. Star\n2. Open Starred\n3. Unstar", "Lists update", "", "Candidate"),
    ("Inbox", "Move to trash and restore", "P1", "Functional", "All", "Alert", "1. Trash\n2. Restore", "Back in inbox", "", "Candidate"),
    ("Inbox", "Delete forever from trash", "P1", "Functional", "All", "Trashed alert", "1. Delete permanently", "Gone; empty trash works", "", "Manual"),
    ("Inbox", "Opening unread marks read", "P0", "Functional", "All", "Unread", "1. Open message", "Unread count decrements", "", "Candidate"),
    ("Bell", "Notification dropdown unread badge", "P1", "UI", "All", "Unread >0", "1. View topbar bell", "Badge shows count; Open alerts inbox link works", "", "Manual"),
    ("Bell", "Mark all read", "P1", "Functional", "All", "Unread", "1. Mark all read", "Badge clears", "", "Manual"),
    ("Mobile back", "Browser back leaves Alerts to previous screen", "P2", "UI", "All", "Navigated from another page", "1. Open Alerts\n2. System Back", "Returns to previous route (open message not history)", "", "Manual"),
)

add(
    "15 Communication",
    ("Clarifications", "College publish → employer/student read", "P1", "Integration", "College, Employer, Student", "Accounts", "1. Publish\n2. Others open", "Content visible", "", "Manual"),
    ("Discussions", "Create and moderate discussion if enabled", "P2", "Functional", "College Admin", "Feature on", "1. Create thread", "Visible to allowed roles", "", "Manual"),
    ("Templates", "College templates CRUD", "P2", "Functional", "College Admin", "Feature", "1. CRUD template", "Works", "", "Manual"),
    ("Email", "Transactional welcome email on student create", "P1", "Functional", "College Admin", "SMTP", "1. Add student\n2. Check mail/logs", "Email sent or logged", "yopmail sandbox", "Manual"),
    ("Email", "Password reset email delivered", "P1", "Functional", "All", "SMTP", "1. Request reset", "Email + log entry", "", "Manual"),
)

add(
    "16 Mentorship",
    ("Flow", "Post → Approve → Volunteer E2E", "P1", "Integration", "Student, College, Employer", "Partnership", "1. Student posts\n2. College approves\n3. Employer volunteers", "All statuses update; no offer side effects", "", "Manual"),
    ("Scope", "Unapproved request hidden from employers", "P0", "Security", "Employer", "Pending request", "1. Browse mentorship", "Not listed", "", "Manual"),
    ("Scope", "Non-partnered employer cannot see campus requests", "P0", "Security", "Employer", "No partnership", "1. Browse", "Empty/blocked", "", "Manual"),
    ("Contrast", "Formal mentorship job still uses apply pipeline", "P2", "Functional", "Employer, Student", "Mentorship program listing", "1. Publish as opportunity\n2. Student apply", "Tracked as application not volunteer", "", "Manual"),
)

add(
    "17 Alumni",
    ("Nav", "Alumni menus replace campus drives", "P0", "Functional", "Student (Alumni)", "Alumni flag", "1. Inspect nav", "Alumni jobs browse/my jobs", "", "Manual"),
    ("Apply", "Alumni applies to alumni job", "P0", "Functional", "Student (Alumni), Employer", "Published alumni job", "1. Apply", "Application created", "", "Manual"),
    ("Non-alumni", "Regular student does not see alumni jobs as primary", "P1", "Security", "Student", "Non-alumni", "1. Inspect menus", "Campus placement menus; not alumni primary", "", "Manual"),
)

add(
    "18 Engagement Calendar",
    ("Calendar", "College creates event visible to students", "P1", "Integration", "College, Student", "None", "1. Create event\n2. Student calendar", "Event visible", "", "Manual"),
    ("Guest", "College need → employer response", "P2", "Integration", "College, Employer", "Partnership", "1. Create need\n2. Employer respond", "Status sync", "", "Manual"),
    ("Sponsorship", "Employer offer visible to college", "P2", "Integration", "Employer, College", "None", "1. Create\n2. College view", "Visible", "", "Manual"),
    ("Seed funding", "Employer submission tracked", "P2", "Functional", "Employer, College", "Feature", "1. Submit\n2. College review screen", "Listed", "", "Manual"),
)

add(
    "19 Reports Audit Export",
    ("Reports", "College reports empty season shows empty state", "P2", "Edge", "College Admin", "No data", "1. Open reports", "Friendly empty; no 500", "", "Manual"),
    ("Audit", "Filter by action DEMO_PURGE", "P2", "Functional", "College, Super Admin", "Demo purge events", "1. Filter DEMO_PURGE", "Matching rows", "", "Manual"),
    ("Audit", "404 log-entries path not used; log-entries works", "P0", "Regression", "College Admin", "None", "1. Load Audit Reports network", "Calls /api/audit/log-entries; 200", "Vercel logs/ quirk", "Candidate"),
    ("Export", "My data export authorization", "P1", "Security", "Student", "Other user id", "1. Attempt export another user via API", "Denied", "", "Candidate"),
    ("CSV", "Large export performance acceptable", "P2", "Performance", "College Admin", "Large dataset", "1. Export", "Completes within reasonable time or streams", "", "Manual"),
)

add(
    "20 CVs Documents Storage",
    ("Upload", "Reject non-PDF/oversize CV", "P1", "Negative", "Student", "CV upload UI", "1. Upload exe/huge file", "Rejected", "", "Manual"),
    ("S3", "Misconfigured S3 surfaces PH-CV-S3-CONFIG in logs", "P1", "Edge", "Student, Super Admin", "Bad AWS config", "1. View CV\n2. Check Error logs", "System code + details", "", "Manual"),
    ("S3", "AccessDenied not treated as missing file", "P1", "Regression", "Student", "HeadObject AccessDenied", "1. View CV", "Access error code; not missing-file copy", "PH-CV-S3-ACCESS", "Manual"),
    ("Verify", "College verify CV when setting requires", "P1", "Functional", "College Admin, Student", "Verification required", "1. Unverified CV\n2. Student try apply\n3. College verify\n4. Apply", "Blocked then allowed", "", "Manual"),
    ("Default", "Set default CV used on apply", "P1", "Functional", "Student", "Multiple CVs", "1. Set default\n2. Apply", "Default attached", "", "Manual"),
)

add(
    "21 Multi-tenant Security",
    ("Isolation", "College A cannot see College B students", "P0", "Security", "College Admin", "Two tenants", "1. API/UI as A for B ids", "Denied/empty", "", "Candidate"),
    ("Isolation", "Student of A cannot apply as B", "P0", "Security", "Student", "Two tenants", "1. Tamper tenant in request", "Denied", "", "Candidate"),
    ("Isolation", "Employer cannot read other employer applications", "P0", "Security", "Employer", "Two employers", "1. Access other employer app id", "Denied", "", "Candidate"),
    ("RBAC", "Student cannot open college students API", "P0", "Security", "Student", "Student session", "1. GET college students API", "401/403", "", "Candidate"),
    ("RBAC", "Employer cannot open platform error logs", "P0", "Security", "Employer", "Employer session", "1. Open admin error logs", "Denied", "", "Candidate"),
    ("CSRF/Auth", "Mutating API without session fails", "P0", "Security", "Public", "Logged out", "1. PATCH notifications", "401", "", "Candidate"),
)

add(
    "22 Mobile UI",
    ("Layout", "Dashboard usable at 375px width", "P1", "UI", "Student", "Mobile viewport", "1. Open key student flows", "No clipped primary actions; hamburger works", "", "Manual"),
    ("Layout", "College students table usable on mobile", "P1", "UI", "College Admin", "Mobile", "1. Open Students", "Stacked/cards or scroll intentional; actions reachable", "", "Manual"),
    ("Landing link", "Landing link hidden on desktop topbar", "P2", "UI", "All", "Desktop >768", "1. Inspect topbar", "Landing link not shown", "", "Manual"),
    ("Theme", "Dark mode toggle persists", "P2", "UI", "All", "None", "1. Toggle theme\n2. Reload", "Theme persists", "", "Manual"),
    ("Search", "Screen search Ctrl+K / mobile search", "P2", "UI", "All", "Dashboard", "1. Search 'alerts'", "Navigates to matching screen", "", "Manual"),
)

add(
    "23 Error Handling Logging",
    ("API", "Thrown 500 persists platform_error_logs + Ref", "P0", "Functional", "Super Admin", "Induce 500", "1. Trigger\n2. Check toast Ref\n3. Error logs", "Matching row with stack", "withApiHandlers", "Manual"),
    ("API", "Soft 200 unavailable auto-logged if no Ref", "P0", "Functional", "Super Admin", "Soft fail without prior log", "1. Trigger\n2. Error logs", "api_soft_failure source", "", "Manual"),
    ("Client", "fetchJson network failure reports client error", "P1", "Functional", "Super Admin", "Offline", "1. Break network mid-fetch", "Client report in logs", "", "Manual"),
    ("Client", "404 missing route reports PH-HTTP-404", "P1", "Functional", "Super Admin", "None", "1. Hit missing API from UI", "Error log with 404 code", "", "Manual"),
    ("UX", "User-facing message strips ops boilerplate", "P2", "UI", "College Admin", "CV failure", "1. View CV error", "No 'Full details were saved…' in college panel", "", "Manual"),
)

add(
    "24 Marketplace Help Feedback",
    ("Help", "Help widget answers from docs", "P2", "Functional", "All", "Help knowledge synced", "1. Ask known question", "Relevant answer", "", "Manual"),
    ("Feedback", "Feedback reaches admin inbox", "P1", "Integration", "Student, Super Admin", "None", "1. Submit feedback\n2. Admin inbox", "Present", "", "Manual"),
    ("Marketplace", "College marketplace browse", "P2", "Functional", "College Admin", "Feature on", "1. Open marketplace", "Loads", "", "Manual"),
    ("Developer", "Developer notes password gate", "P1", "Security", "Public", "None", "1. Open /developer without auth", "Password/gate required", "", "Manual"),
    ("Guided runner", "Guided runner hidden when env disables", "P2", "UI", "All", "HIDE_GUIDED_RUNNER=true", "1. Load dashboard", "Runner not shown", "", "Manual"),
)

add(
    "25 Sandbox Demo QA",
    ("Demo logins", "Demo login chips work when enabled", "P1", "Functional", "Public", "Demo logins enabled", "1. Use demo login on sign-in", "Signs into correct role", "", "Manual"),
    ("Demo purge", "Demo purge removes sandbox rows only", "P1", "Functional", "Developer", "Demo data present", "1. Run purge\n2. Verify seed untouched", "Only sandbox markers removed; audit DEMO_PURGE", "", "Manual"),
    ("Data tester", "Data Tester API creates sandbox entities", "P2", "Functional", "Developer", "DEMO_DATA_API_ENABLED", "1. Create via tester", "Entities usable on screens", "", "Manual"),
    ("Screen IDs", "Dev screen ID pills show when enabled", "P2", "UI", "All", "SHOW_DEV_SCREEN_IDS", "1. Visit screens", "S-xx pills visible", "", "Manual"),
    ("Banner", "Sandbox banner only on landing", "P1", "UI", "Public", "Banner enabled", "1. Open /\n2. Open /dashboard", "Banner on landing only", "", "Manual"),
)


def style_header(ws):
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    # freeze_panes set by add_back_to_index after inserting the Index link row



def write_cases(ws, module_name, rows):
    style_header(ws)
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    body = Font(name="Arial", size=9)
    wrap = Alignment(vertical="top", wrap_text=True)
    prio_fills = {
        "P0": PatternFill("solid", fgColor="FEE2E2"),
        "P1": PatternFill("solid", fgColor="FFEDD5"),
        "P2": PatternFill("solid", fgColor="FEF9C3"),
        "P3": PatternFill("solid", fgColor="F1F5F9"),
    }
    prefix = "".join(ch for ch in module_name.split(" ", 1)[0] if ch.isdigit()).zfill(2) or "00"
    slug = "".join(ch for ch in module_name if ch.isalnum())[:8].upper()

    for i, row in enumerate(rows, 1):
        feature, title, priority, typ, roles, pre, steps, expected, notes, auto = row
        tc_id = f"TC-{prefix}-{i:03d}"
        values = [tc_id, module_name, feature, title, priority, typ, roles, pre, steps, expected, notes, auto, ""]
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 1, col, val)
            cell.font = body
            cell.alignment = wrap
            cell.border = thin
            if col == 5 and priority in prio_fills:
                cell.fill = prio_fills[priority]
        ws.row_dimensions[i + 1].height = max(45, min(90, 18 + steps.count("\n") * 12))

    widths = [12, 22, 16, 42, 9, 12, 22, 28, 40, 36, 28, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Status dropdown
    dv = DataValidation(type="list", formula1='"Not Started,Pass,Fail,Blocked,Skipped,N/A"', allow_blank=True)
    ws.add_data_validation(dv)
    if rows:
        dv.add(f"M2:M{len(rows)+1}")

    # Extend autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows)+1}"


def sheet_jump(sheet_name):
    """Excel internal hyperlink target for a worksheet name."""
    return f"#'{sheet_name}'!A1"


def write_index(wb, sheet_names):
    """First tab: Index with clickable links to every other sheet."""
    ws = wb.create_sheet("Index", 0)
    title_font = Font(name="Arial", bold=True, size=16, color="1E3A5F")
    sub = Font(name="Arial", size=10, color="475569")
    head_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    body = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws["A1"] = "PlacementHub — Test Case Index"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Click a module name to open that tab. "
        "Source: docs/product/placementhub-functionality.md + in-app modules (sandbox / pre-launch)."
    )
    ws["A2"].font = sub
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 36

    headers = ["#", "Module / Tab (click to open)", "Case Count", "P0", "P1", "P2+", "Link"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font = head_font
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
    ws.row_dimensions[4].height = 30

    # Coverage matrix row
    cov_name = "Coverage Matrix"
    ws.cell(5, 1, 0).font = body
    ws.cell(5, 1).border = thin
    ws.cell(5, 1).alignment = Alignment(horizontal="center")
    link_cell = ws.cell(5, 2, cov_name)
    link_cell.font = link_font
    link_cell.hyperlink = sheet_jump(cov_name)
    link_cell.border = thin
    for col in range(3, 7):
        cell = ws.cell(5, col, "—" if col > 3 else "Checklist")
        cell.font = body
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")
        go = ws.cell(5, 7, "Open >")
        go.font = link_font
        go.hyperlink = sheet_jump(cov_name)
        go.border = thin
        go.alignment = Alignment(horizontal="center")

    total = 0
    p0 = p1 = p2 = 0
    for idx, (name, rows) in enumerate(MODULES.items(), 1):
        sheet_name = sheet_names[name]
        c0 = sum(1 for r in rows if r[2] == "P0")
        c1 = sum(1 for r in rows if r[2] == "P1")
        c2 = sum(1 for r in rows if r[2] in ("P2", "P3"))
        total += len(rows)
        p0 += c0
        p1 += c1
        p2 += c2
        row_num = 5 + idx

        num = ws.cell(row_num, 1, idx)
        num.font = body
        num.border = thin
        num.alignment = Alignment(horizontal="center")

        name_cell = ws.cell(row_num, 2, name)
        name_cell.font = link_font
        name_cell.hyperlink = sheet_jump(sheet_name)
        name_cell.border = thin

        for col, v in enumerate([len(rows), c0, c1, c2], 3):
            cell = ws.cell(row_num, col, v)
            cell.font = body
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

        go = ws.cell(row_num, 7, "Open >")
        go.font = link_font
        go.hyperlink = sheet_jump(sheet_name)
        go.border = thin
        go.alignment = Alignment(horizontal="center")

    r = 6 + len(MODULES)
    ws.cell(r, 1, "").border = thin
    tot_label = ws.cell(r, 2, "TOTAL")
    tot_label.font = Font(name="Arial", bold=True, size=10)
    tot_label.border = thin
    for col, v in enumerate([total, p0, p1, p2], 3):
        cell = ws.cell(r, col, v)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")
    ws.cell(r, 7, "").border = thin

    ws["A" + str(r + 2)] = "Priority legend"
    ws["A" + str(r + 2)].font = Font(name="Arial", bold=True, size=10)
    ws["A" + str(r + 3)] = "P0 = Critical path / security / data integrity"
    ws["A" + str(r + 4)] = "P1 = High — core role workflows"
    ws["A" + str(r + 5)] = "P2/P3 = Medium/Low — engagement, polish, edge"
    ws["A" + str(r + 7)] = "Columns on each module tab"
    ws["A" + str(r + 7)].font = Font(name="Arial", bold=True, size=10)
    ws["A" + str(r + 8)] = ", ".join(HEADERS)
    ws["A" + str(r + 10)] = "Tip: Use Open > or click the module name. Each module tab has a << Back to Index link."
    ws["A" + str(r + 10)].font = sub

    for i, w in enumerate([6, 36, 12, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_back_to_index(ws):
    """Place a back-link to Index above the case table header row."""
    ws.insert_rows(1)
    link = ws.cell(1, 1, "<< Back to Index")
    link.font = Font(name="Arial", size=10, color="0563C1", underline="single", bold=True)
    link.hyperlink = sheet_jump("Index")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A3"


def write_coverage(wb):
    ws = wb.create_sheet("Coverage Matrix", 1)
    font = Font(name="Arial", size=9)
    head = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1E3A5F")
    yes = PatternFill("solid", fgColor="DCFCE7")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    link = ws.cell(1, 1, "<< Back to Index")
    link.font = Font(name="Arial", size=10, color="0563C1", underline="single", bold=True)
    link.hyperlink = sheet_jump("Index")

    ws["A2"] = "Role x capability coverage checklist (mark during planning)"
    ws["A2"].font = Font(name="Arial", bold=True, size=12, color="1E3A5F")
    ws.merge_cells("A2:G2")

    roles = ["Student", "Employer", "College Admin", "Placement Committee", "Super Admin", "Public"]
    caps = [
        "Sign-in / session",
        "Profile / settings",
        "Browse opportunities",
        "Apply / withdraw",
        "Publish listings / drives",
        "Approve partnerships / visibility",
        "Assessments / hiring results",
        "Interviews",
        "Offers accept/create",
        "Placement rules / FCFS",
        "Students master / CSV",
        "Alerts inbox",
        "Clarifications",
        "Mentorship requests",
        "Alumni jobs",
        "Reports / audit",
        "Error / email logs",
        "Platform tenant admin",
    ]
    matrix = {
        "Sign-in / session": set(roles),
        "Profile / settings": {"Student", "Employer", "College Admin", "Super Admin", "Placement Committee"},
        "Browse opportunities": {"Student"},
        "Apply / withdraw": {"Student"},
        "Publish listings / drives": {"Employer"},
        "Approve partnerships / visibility": {"College Admin", "Placement Committee"},
        "Assessments / hiring results": {"Employer", "College Admin"},
        "Interviews": {"Student", "Employer", "College Admin"},
        "Offers accept/create": {"Student", "Employer", "College Admin"},
        "Placement rules / FCFS": {"College Admin", "Employer", "Student"},
        "Students master / CSV": {"College Admin", "Placement Committee"},
        "Alerts inbox": {"Student", "Employer", "College Admin", "Placement Committee", "Super Admin"},
        "Clarifications": {"Student", "Employer", "College Admin"},
        "Mentorship requests": {"Student", "Employer", "College Admin"},
        "Alumni jobs": {"Student", "Employer"},
        "Reports / audit": {"College Admin", "Super Admin"},
        "Error / email logs": {"Super Admin"},
        "Platform tenant admin": {"Super Admin"},
    }
    headers = ["Capability"] + roles
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font = head
        c.fill = fill
        c.border = thin
    for r, cap in enumerate(caps, 5):
        cell = ws.cell(r, 1, cap)
        cell.font = font
        cell.border = thin
        for c, role in enumerate(roles, 2):
            val = "Y" if role in matrix.get(cap, set()) else "—"
            cell = ws.cell(r, c, val)
            cell.font = font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if val == "Y":
                cell.fill = yes
    ws.column_dimensions["A"].width = 36
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "A5"


def main():
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheet_names = {}
    for name, rows in MODULES.items():
        sheet_name = name[:31]
        sheet_names[name] = sheet_name
        ws = wb.create_sheet(sheet_name)
        write_cases(ws, name, rows)
        add_back_to_index(ws)
        n = len(rows)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{n + 2}"
        for dv in ws.data_validations.dataValidation:
            dv.sqref = f"M3:M{n + 2}"

    write_coverage(wb)
    write_index(wb, sheet_names)

    # Order: Index, Coverage Matrix, then modules
    desired = ["Index", "Coverage Matrix"] + [sheet_names[n] for n in MODULES]
    for i, name in enumerate(desired):
        current = wb.sheetnames.index(name)
        if current != i:
            wb.move_sheet(name, offset=i - current)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    total = sum(len(v) for v in MODULES.values())
    print(f"Wrote {OUT}")
    print(f"First sheet: {wb.sheetnames[0]}")
    print(f"Modules: {len(MODULES)} | Cases: {total}")
    print("Order:", " > ".join(wb.sheetnames[:4]), "...")


if __name__ == "__main__":
    main()
