#!/usr/bin/env python3
"""
Delta test cases for features added AFTER the last full PlacementHub-Test-Cases generation.

Does NOT duplicate the main suite in docs/PlacementHub-Test-Cases.xlsx.
Output: docs/PlacementHub-Test-Cases-Delta-Post-Gen.xlsx
         docs/PlacementHub-Test-Cases-Delta-Post-Gen.csv
"""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "docs" / "PlacementHub-Test-Cases-Delta-Post-Gen.xlsx"
OUT_CSV = ROOT / "docs" / "PlacementHub-Test-Cases-Delta-Post-Gen.csv"

HEADERS = [
    "TC ID",
    "Module",
    "Feature",
    "Title",
    "Priority",
    "Type",
    "Role(s)",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Test Data / Notes",
    "Automation",
    "Status",
]

# Scope note for Index sheet
SCOPE = (
    "Delta suite — features shipped after the last full test-case generation. "
    "Covers: College Calendar ICS import/export/delete/filter/clash; Feature Ideas (college); "
    "db:clear-placement Alerts/Audit/non-core colleges; Demo college (Demo) suffix; "
    "Developer unlock show-password; Hub Home titles."
)

MODULES: dict[str, list[tuple]] = {}


def add(module: str, *rows: tuple) -> None:
    MODULES.setdefault(module, []).extend(rows)


# ---------------------------------------------------------------------------
# 01 College Calendar ICS & clash
# ---------------------------------------------------------------------------
add(
    "01 College Calendar ICS",
    (
        "ICS Import",
        "College can open Import calendar and upload a valid .ics",
        "P0",
        "Functional",
        "College Admin",
        "Signed in as college admin; Calendar page open; sample file docs/sample-import-july26.ics available",
        "1. Open /dashboard/college/calendar\n2. Click Import (ICS)\n3. Select sample-import-july26.ics\n4. Keep default “from today onward”\n5. Confirm import",
        "Import succeeds; events for Jul 26–28 2026 appear on calendar; toast/success confirmation shown",
        "docs/sample-import-july26.ics",
        "Manual",
    ),
    (
        "ICS Import",
        "Dry-run / preview shows clash warnings when imported dates overlap placement drives",
        "P1",
        "Functional",
        "College Admin",
        "Campus has a placement drive on/near Jul 26; sample ICS ready",
        "1. Open Import ICS\n2. Upload sample with Jul 26 events\n3. Review preview/clash section before confirm (if shown) or after import banner",
        "Clash summary mentions imported/holiday/exam vs placement drive; import can still complete",
        "Uses calendarClashDetection + import API warnings",
        "Manual",
    ),
    (
        "ICS Import",
        "Re-import same UID does not create duplicate rows",
        "P0",
        "Regression",
        "College Admin",
        "Already imported sample ICS once",
        "1. Import the same .ics again\n2. Count events for those dates / check DB source_uid",
        "No duplicate events for same source_uid; count stable or update-in-place only",
        "source_uid unique per tenant (migration 112)",
        "Candidate",
    ),
    (
        "ICS Import",
        "Invalid / empty .ics is rejected with clear error",
        "P1",
        "Negative",
        "College Admin",
        "Calendar open",
        "1. Attempt import of empty file or non-ICS text renamed .ics",
        "Error message; no calendar rows created",
        "",
        "Manual",
    ),
    (
        "ICS Export",
        "Unified Export menu offers CSV and ICS for this month and full range",
        "P0",
        "Functional",
        "College Admin",
        "Calendar has at least one event",
        "1. Open calendar Export control\n2. Confirm options: CSV this month, CSV full, ICS this month, ICS full\n3. Download ICS this month\n4. Open file in text editor",
        "File downloads; begins with BEGIN:VCALENDAR; contains VEVENT rows for selected range",
        "ExportCollegeCalendarButton",
        "Manual",
    ),
    (
        "ICS Export",
        "ICS full export includes imported events",
        "P1",
        "Functional",
        "College Admin",
        "Imported ICS events exist",
        "1. Export ICS full\n2. Search for an imported SUMMARY (e.g. Founders Day)",
        "Imported event present in .ics output",
        "",
        "Manual",
    ),
    (
        "Delete imported",
        "Delete imported removes only source_uid rows in selected range",
        "P0",
        "Functional",
        "College Admin",
        "Mix of placement_drive calendar rows and imported (source_uid) rows",
        "1. Open Delete imported\n2. Choose date range covering Jul 26\n3. Confirm delete\n4. Refresh calendar",
        "Imported events gone; manually created / placement_drive events remain",
        "DELETE /api/college/calendar/imported",
        "Manual",
    ),
    (
        "Delete imported",
        "Delete all imported clears every source_uid event for tenant",
        "P1",
        "Functional",
        "College Admin",
        "Multiple imported events across months",
        "1. Choose delete all imported\n2. Confirm\n3. Filter category Imported",
        "Imported category empty; other categories still show events",
        "",
        "Manual",
    ),
    (
        "Category filter",
        "Filter All / Placement / Imported / Programs changes visible events",
        "P0",
        "Functional",
        "College Admin",
        "Calendar has placement, imported, and program-type items (or at least two categories)",
        "1. Select Imported — only imported rows\n2. Select Placement — drive-related\n3. Select All — union",
        "List/month view matches filter; no crash; empty state when none match",
        "CollegeCalendarCategoryFilter",
        "Candidate",
    ),
    (
        "Clash banner",
        "Clash banner surfaces placement vs imported blocking dates",
        "P1",
        "Functional",
        "College Admin",
        "Imported blocking/holiday/exam overlaps a drive date",
        "1. Open calendar with overlapping data\n2. Observe clash banner / warnings",
        "Banner lists clash with imported item labeled clearly; actionable or informational copy shown",
        "CollegeCalendarClashBanner",
        "Manual",
    ),
    (
        "Auth",
        "Non-college roles cannot call calendar import/export/imported APIs",
        "P0",
        "Security",
        "Student, Employer",
        "Student and employer sessions",
        "1. POST /api/college/calendar/import as student\n2. GET /api/college/calendar/export as employer",
        "401/403; no data mutation",
        "",
        "Candidate",
    ),
)

# ---------------------------------------------------------------------------
# 02 Feature Ideas (college)
# ---------------------------------------------------------------------------
add(
    "02 Feature Ideas",
    (
        "Navigation",
        "Feature Ideas appears under college Communication & Support and loads",
        "P0",
        "Functional",
        "College Admin",
        "Migration 114 applied; college admin session (e.g. admin@iitm.edu)",
        "1. Sign in as college admin\n2. Open Communication & Support → Feature Ideas\n3. Or open /dashboard/college/feature-ideas",
        "Page loads with status/topic filters, Submit Idea, and list (or empty state)",
        "npm run db:migrate:114",
        "Candidate",
    ),
    (
        "Submit",
        "Submit idea with title, description, and topic(s)",
        "P0",
        "Functional",
        "College Admin",
        "Feature Ideas page open",
        "1. Click + Submit Idea\n2. Enter title and description\n3. Select 1–3 topics\n4. Submit",
        "Success dialog (“Idea submitted”); idea appears in list with Pending approval; author auto-upvoted (vote_count ≥ 1)",
        "",
        "Manual",
    ),
    (
        "Submit",
        "Submit blocked without topic or empty title/description",
        "P1",
        "Negative",
        "College Admin",
        "Submit modal open",
        "1. Clear topics\n2. Attempt submit\n3. Retry with blank title",
        "Validation warning/error; no new row created",
        "",
        "Candidate",
    ),
    (
        "Vote",
        "College admin can upvote and remove vote on another idea",
        "P0",
        "Functional",
        "College Admin",
        "At least one idea from another user (or second college admin)",
        "1. Click vote chevron on an idea\n2. Observe count +1 and voted state\n3. Click again",
        "Vote toggles; count increments then decrements; UI reflects voted_by_me",
        "POST /api/college/feature-ideas/:id/vote",
        "Manual",
    ),
    (
        "Filters",
        "Status and topic filters narrow the list; search finds by title",
        "P1",
        "Functional",
        "College Admin",
        "Multiple ideas with varied topics/statuses",
        "1. Filter by New Feature topic\n2. Filter by Pending approval\n3. Search unique title fragment",
        "Only matching ideas shown; counts in sidebar update from API aggregates",
        "",
        "Manual",
    ),
    (
        "Sort",
        "Trending vs Newest sort order",
        "P2",
        "Functional",
        "College Admin",
        "≥2 ideas with different vote_count / created_at",
        "1. Sort Trending\n2. Sort Newest",
        "Trending orders by votes then recency; Newest by created_at desc",
        "",
        "Manual",
    ),
    (
        "Auth",
        "Student/employer cannot open Feature Ideas or APIs",
        "P0",
        "Security",
        "Student, Employer",
        "Non-college sessions",
        "1. Open /dashboard/college/feature-ideas as student\n2. GET /api/college/feature-ideas as employer",
        "Route blocked/redirect; API 401",
        "College-only MVP",
        "Candidate",
    ),
    (
        "Auth",
        "Feature Ideas menu not shown for student/employer",
        "P1",
        "UI",
        "Student, Employer",
        "Signed in as those roles",
        "1. Inspect sidebar menus",
        "No Feature Ideas entry; Feedback may still exist where shared",
        "",
        "Manual",
    ),
)

# ---------------------------------------------------------------------------
# 03 Clear placement & demo data hygiene
# ---------------------------------------------------------------------------
add(
    "03 Clear Placement Cleanup",
    (
        "Alerts",
        "db:clear-placement deletes all notifications (Alerts)",
        "P0",
        "Functional",
        "Developer / QA",
        "At least one in-app notification exists; .env.local DATABASE_URL set",
        "1. Note Alerts count in UI or DB\n2. Run npm run db:clear-placement\n3. Confirm terminal After: Alerts (notifications): 0\n4. Hard-refresh /dashboard/alerts",
        "Inbox and trash empty; terminal lists Alerts line",
        "scripts/clear_all_placement_data.js + SQL DELETE FROM notifications",
        "Manual",
    ),
    (
        "Audit logs",
        "db:clear-placement deletes audit_logs",
        "P0",
        "Functional",
        "Developer / QA",
        "audit_logs has rows; college Audit Reports previously showed activity",
        "1. Run npm run db:clear-placement\n2. Check After: Audit logs: 0\n3. Open college Audit Reports",
        "Audit log list empty / zero rows",
        "DELETE FROM audit_logs",
        "Manual",
    ),
    (
        "Test colleges",
        "Non-core college tenants deleted; seed campuses retained",
        "P0",
        "Functional",
        "Developer / QA",
        "At least one registration/test college exists besides seed slugs",
        "1. Run npm run db:clear-placement\n2. Read “Test colleges to delete” and “Remaining colleges”\n3. Verify seed campuses still present",
        "Only core seed colleges remain (iit-madras, nit-trichy, bits-pilani, jadavpur, vit, dtu, iiith); test tenants gone",
        "CORE_COLLEGE_SLUGS in clear_all_placement_data.js",
        "Manual",
    ),
    (
        "Preserve",
        "Core users/students/employers on seed campuses survive clear",
        "P0",
        "Regression",
        "Developer / QA",
        "Demo logins work before clear",
        "1. Run clear-placement\n2. Sign in as admin@iitm.edu and hr@techcorp.com",
        "Login succeeds; jobs/drives cleared but accounts intact",
        "May need restore demo tie-ups after wipe",
        "Manual",
    ),
    (
        "Terminal",
        "Before/After snapshot prints Alerts and Audit logs counts",
        "P2",
        "UI",
        "Developer / QA",
        "None",
        "1. Run npm run db:clear-placement\n2. Inspect console output",
        "Lines include “Alerts (notifications)” and “Audit logs” in Before and After blocks",
        "",
        "Manual",
    ),
)

# ---------------------------------------------------------------------------
# 04 Demo naming & hub titles
# ---------------------------------------------------------------------------
add(
    "04 Demo Naming Hub Titles",
    (
        "Demo suffix",
        "Seed college tenant names end with (Demo)",
        "P0",
        "Functional",
        "College Admin, Super Admin",
        "Migration 113 applied",
        "1. Sign in as IITM college admin\n2. Observe hub/header tenant name\n3. Check college list / partnership UIs showing tenant name",
        "Name shows e.g. “Indian Institute of Technology, Madras (Demo)” — not bare name",
        "npm run db:migrate:113; seed.sql updated",
        "Manual",
    ),
    (
        "Demo suffix",
        "Re-running migrate 113 does not double-suffix",
        "P1",
        "Regression",
        "Developer / QA",
        "Names already end with (Demo)",
        "1. npm run db:migrate:113 again\n2. SELECT name FROM tenants WHERE slug='iit-madras'",
        "Still a single “ (Demo)” suffix",
        "Idempotent WHERE name !~* '(Demo)\\s*$'",
        "Candidate",
    ),
    (
        "Hub title",
        "College hub browser/page title uses “{TenantName} Home”",
        "P1",
        "UI",
        "College Admin",
        "College admin with tenantName",
        "1. Open college full-screen hub / home\n2. Check document title or hub heading",
        "Shows tenant name + Home (includes Demo suffix if present)",
        "DashboardFullScreenHub getHubPageTitle",
        "Manual",
    ),
    (
        "Hub title",
        "Student hub title uses “{FirstName} — Home”",
        "P1",
        "UI",
        "Student",
        "Student session with first name",
        "1. Open student hub\n2. Check title",
        "First name then — Home",
        "",
        "Manual",
    ),
    (
        "Demo logins",
        "Login demo chips for colleges show (Demo) in labels where updated",
        "P2",
        "UI",
        "Public",
        "Demo logins enabled on /login",
        "1. Open /login\n2. Inspect college admin / committee demo account names",
        "Demo college labels include (Demo)",
        "src/lib/demoLogins.js",
        "Manual",
    ),
)

# ---------------------------------------------------------------------------
# 05 Developer unlock password UX
# ---------------------------------------------------------------------------
add(
    "05 Developer Unlock",
    (
        "Show password",
        "Internal tools unlock page has show/hide password toggle",
        "P0",
        "UI",
        "Public / team",
        "None",
        "1. Open /developer/unlock\n2. Type a password\n3. Click eye icon\n4. Click again",
        "Password revealed as text then masked again; aria-label updates Show/Hide",
        "Matches login page Eye/EyeOff pattern",
        "Candidate",
    ),
    (
        "Unlock",
        "Correct team password unlocks and navigates to developer notes",
        "P0",
        "Functional",
        "Public / team",
        "Known developer notes password",
        "1. Enter correct password\n2. Click Unlock",
        "Full navigation to /developer (or from= target); subsequent visit allowed via cookie",
        "Wolfe123@# / env override per team docs",
        "Manual",
    ),
    (
        "Unlock",
        "Wrong password shows error and does not grant access",
        "P0",
        "Negative",
        "Public / team",
        "None",
        "1. Enter wrong password\n2. Submit\n3. Open /developer",
        "Incorrect password message; /developer still gated to unlock",
        "",
        "Manual",
    ),
)


def sheet_jump(name: str) -> str:
    return f"#'{name}'!A1"


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28


def write_cases(ws, module_name: str, rows: list[tuple]) -> list[list]:
    style_header(ws)
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    body = Font(name="Arial", size=9)
    wrap = Alignment(vertical="top", wrap_text=True)
    prio_fills = {
        "P0": PatternFill("solid", fgColor="FEE2E2"),
        "P1": PatternFill("solid", fgColor="FFEDD5"),
        "P2": PatternFill("solid", fgColor="FEF9C3"),
        "P3": PatternFill("solid", fgColor="F1F5F9"),
    }
    prefix = "".join(ch for ch in module_name.split(" ", 1)[0] if ch.isdigit()).zfill(2) or "00"
    flat_rows: list[list] = []

    widths = [12, 26, 16, 42, 8, 12, 18, 36, 44, 40, 28, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, row in enumerate(rows, 1):
        feature, title, priority, typ, roles, pre, steps, expected, notes, auto = row
        tc_id = f"TC-D{prefix}-{i:03d}"
        values = [tc_id, module_name, feature, title, priority, typ, roles, pre, steps, expected, notes, auto, ""]
        flat_rows.append(values)
        for col, val in enumerate(values, 1):
            cell = ws.cell(i + 1, col, val)
            cell.font = body
            cell.alignment = wrap
            cell.border = thin
            if col == 5:
                cell.fill = prio_fills.get(priority, prio_fills["P3"])
                cell.alignment = Alignment(horizontal="center", vertical="top")
        ws.row_dimensions[i + 1].height = 72

    n = len(rows)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{n + 1}"
    ws.freeze_panes = "A2"
    return flat_rows


def add_back_to_index(ws) -> None:
    ws.insert_rows(1)
    link = ws.cell(1, 1, "<< Back to Index")
    link.font = Font(name="Arial", size=10, color="0563C1", underline="single", bold=True)
    link.hyperlink = sheet_jump("Index")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A3"


def write_index(wb, sheet_names: dict[str, str], all_flat: list[list]) -> None:
    ws = wb.create_sheet("Index", 0)
    title_font = Font(name="Arial", bold=True, size=16, color="1E3A5F")
    sub = Font(name="Arial", size=10, color="475569")
    head_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    body = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    ws["A1"] = "PlacementHub — Delta Test Cases (Post last generation)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    ws["A2"] = SCOPE
    ws["A2"].font = sub
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 48

    headers = ["#", "Module / Tab", "Case Count", "P0", "P1", "P2+", "Link"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(4, col, h)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = thin

    total = p0 = p1 = p2 = 0
    for idx, (name, rows) in enumerate(MODULES.items(), 1):
        sheet_name = sheet_names[name]
        c0 = sum(1 for r in rows if r[2] == "P0")
        c1 = sum(1 for r in rows if r[2] == "P1")
        c2 = sum(1 for r in rows if r[2] in ("P2", "P3"))
        total += len(rows)
        p0 += c0
        p1 += c1
        p2 += c2
        row_num = 4 + idx
        ws.cell(row_num, 1, idx).border = thin
        name_cell = ws.cell(row_num, 2, name)
        name_cell.font = link_font
        name_cell.hyperlink = sheet_jump(sheet_name)
        name_cell.border = thin
        for col, v in enumerate([len(rows), c0, c1, c2], 3):
            cell = ws.cell(row_num, col, v)
            cell.font = body
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
        go = ws.cell(row_num, 7, "Open >")
        go.font = link_font
        go.hyperlink = sheet_jump(sheet_name)
        go.border = thin

    r = 5 + len(MODULES)
    ws.cell(r, 2, "TOTAL").font = Font(name="Arial", bold=True, size=10)
    for col, v in enumerate([total, p0, p1, p2], 3):
        cell = ws.cell(r, col, v)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")

    ws[f"A{r + 2}"] = "Companion to docs/PlacementHub-Test-Cases.xlsx — do not merge blindly; keep as delta pack."
    ws[f"A{r + 2}"].font = sub
    ws[f"A{r + 3}"] = f"CSV export: {OUT_CSV.name} ({len(all_flat)} rows)"
    ws[f"A{r + 3}"].font = sub

    for i, w in enumerate([6, 36, 12, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    sheet_names: dict[str, str] = {}
    all_flat: list[list] = []

    for name, rows in MODULES.items():
        sheet_name = name[:31]
        sheet_names[name] = sheet_name
        ws = wb.create_sheet(sheet_name)
        flat = write_cases(ws, name, rows)
        all_flat.extend(flat)
        add_back_to_index(ws)
        n = len(rows)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{n + 2}"

    write_index(wb, sheet_names, all_flat)

    desired = ["Index"] + [sheet_names[n] for n in MODULES]
    for i, name in enumerate(desired):
        current = wb.sheetnames.index(name)
        if current != i:
            wb.move_sheet(name, offset=i - current)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(all_flat)

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    print(f"Modules: {len(MODULES)} | Cases: {len(all_flat)}")


if __name__ == "__main__":
    main()
